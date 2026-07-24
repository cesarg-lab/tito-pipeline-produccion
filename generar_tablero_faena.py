#!/usr/bin/env python3
"""
generar_tablero_faena.py — Tablero Arauco pre-llenado por faena (SOLO productividad).
Módulo de productividad · spec 09. Salida que el jefe TRASPASA a la pizarra física.

Fuentes: reporte PG del NOC (agrupado por hora_inicio = día trabajado, NO fecha_noc),
metas del Excel CONFIGURACIÓN, capacidad = trozado p90.
Buffers: SIN_DATO hasta que el jefe declare (avance_faena en el CMMS) — nunca verde por defecto.

Uso:  python3 generar_tablero_faena.py [pg_historico.json]
      (sin argumento baja PG del mes por API vía descargar_noc_api)
"""
import json, sys, base64, os
from pathlib import Path
import pandas as pd, numpy as np

BASE = Path(__file__).parent
TEAM_MAP = {'S123':'M1.1','S58':'M1.2','S223':'M1.3','S246':'M1.4','MG5':'M5','TEA02':'M7','TEA08':'M9','T125':'M11','TEA30':'M1.3'}
NOMBRE = {'M1.1':'Millalemu 1.1','M1.2':'Millalemu 1.2','M1.3':'Millalemu 1.3','M1.4':'Millalemu 1.4','M5':'Millalemu 5','M7':'Millalemu 7','M9':'Millalemu 9','M11':'Millalemu 11'}
ESP = {'PIRA':'Pino radiata','EUGL':'Euca globulus','EUNI':'Euca nitens'}
METAS_DEFAULT = {'M1.1':8000,'M1.2':8000,'M1.3':8000,'M1.4':8000,'M5':7000,'M7':8060,'M9':5940,'M11':6000}
USO, HDISP = 0.90, 10.5
TR=[0,.15,.25,.35,.50,.75,9]; LB=['<0,15','0,15-0,25','0,25-0,35','0,35-0,50','0,50-0,75','>0,75']

# Tecnologías reales de la flota (3): el SKIDDER/GRAPPLE/FORWINCH del NOC son el MISMO
# skidder 6×6 de garra (FORWINCH = con winche para pendiente) → una sola tecnología de
# madereo. La torre (Alpine) es la otra. El volteo (feller/shovel) no tiene volumen en el
# NOC (solo trozado), se calibra con lo que informa el jefe/operador.
TEC_NORM={'SKIDDER':'SKIDDER 6X6 GRAPPLE','GRAPPLE':'SKIDDER 6X6 GRAPPLE',
          'FORWINCH':'SKIDDER 6X6 GRAPPLE','TORRE500':'TORRE'}

def num(s): return pd.to_numeric(pd.Series(s).astype(str).str.replace(',','.'),errors='coerce')

def cargar_pg():
    """PG por API (hora_inicio completa) o desde json local para pruebas."""
    if len(sys.argv)>1 and Path(sys.argv[1]).exists():
        return pd.DataFrame(json.loads(Path(sys.argv[1]).read_text()))
    try:
        import descargar_noc_api as noc, requests
        from datetime import datetime
        hoy=datetime.now(); fi=hoy.replace(day=1).strftime('%Y-%m-%d'); ff=hoy.strftime('%Y-%m-%d')
        s=requests.Session(); s.verify=False; s.headers.update({'User-Agent':'Mozilla/5.0'})
        tok=noc.obtener_token_arcgis(s)
        try: noc.establecer_sesion(s,tok)
        except Exception: pass
        return pd.DataFrame(noc.descargar_reporte(s,tok,'PG',fi,ff))
    except Exception as e:
        print(f"⚠️  no se pudo bajar PG ({e})"); return None

def metas_excel():
    m=dict(METAS_DEFAULT); xl=BASE/"Dashboard_CosechaForestal.xlsx"
    if xl.exists():
        try:
            from openpyxl import load_workbook
            wb=load_workbook(str(xl),data_only=True)
            if "CONFIGURACIÓN" in wb.sheetnames:
                ws=wb["CONFIGURACIÓN"]; teams=['M1.1','M1.2','M1.3','M1.4','M5','M7','M9','M11']
                for i,t in enumerate(teams):
                    v=ws.cell(24+i,5).value
                    if v: m[t]=float(v)
            wb.close()
        except Exception as e: print(f"⚠️  metas Excel: {e}; uso defaults")
    return m

def base_diaria(df):
    d=pd.DataFrame({
      'dia': df['hora_inicio'].astype(str).str[:10],   # DIA TRABAJADO
      'faena': df['equipo'].astype(str).str.strip().map(TEAM_MAP),
      'tec': df['tipo_equipo'].astype(str).str.strip().str.upper()
               .map(TEC_NORM).fillna(df['tipo_equipo'].astype(str).str.strip().str.upper()),
      'especie': df['desc_especie'].astype(str).str.strip().str.upper(),
      'predio': df['codigo_predio'].astype(str).str.strip(),
      'vol': num(df['m3ssc_pu']).fillna(0)+num(df['m3ssc_as']).fillna(0),
      'arb': num(df['arboles_madereados']).fillna(0),
      'ciclos': num(df['numero_ciclos']).fillna(0),
      'te': num(df['tiempo_efectivo']).fillna(0)})
    d=d[d.faena.notna()].copy()
    pos=d.te[d.te>0]; d['hrs']=d.te/(3600 if pos.median()>1000 else 60)
    def moda(s):
        s=s[s.astype(str).str.len()>0]; return s.mode().iat[0] if len(s.mode()) else ''
    g=d.groupby(['faena','dia']).agg(predio=('predio',moda),especie=('especie',moda),tec=('tec',moda),
       m3=('vol','sum'),arb=('arb','sum'),hrs=('hrs','sum'),ciclos=('ciclos','sum')).reset_index()
    g['VMA']=np.where(g.arb>0,g.m3/g.arb,np.nan).round(3)
    g['rend']=np.where(g.hrs>0,g.m3/g.hrs,np.nan).round(1)
    g['carga']=np.where(g.ciclos>0,g.m3/g.ciclos,np.nan).round(2)
    g['ritmo']=np.where(g.hrs>0,g.ciclos/g.hrs,np.nan).round(2)
    return g

def tabla_p75(g):
    d=g.dropna(subset=['VMA','rend']).copy(); d=d[(d.hrs>=3)&(d.carga.between(.3,15))&(d.ciclos>=3)]
    d['tramo']=pd.cut(d.VMA,TR,labels=LB)
    cell={}
    for (tec,esp,tr),x in d.groupby(['tec','especie','tramo'],observed=True):
        if len(x)>=3:
            cell[(tec,esp,tr)]=dict(p75=round(x.rend.quantile(.75),1),faenas=x.faena.nunique(),
                p50=round(x.rend.median(),1),carga=round(x.carga.median(),2),
                ritmo=round(x.ritmo.median(),2),n=len(x))
    return d,cell


TRAMO_MID={'<0,15':0.10,'0,15-0,25':0.20,'0,25-0,35':0.30,'0,35-0,50':0.42,'0,50-0,75':0.62,'>0,75':0.90}

def teorico(d):
    """Modelo teórico: rendimiento esperado (m³/h) por tec × especie × tramo, para LAS 3
    especies y TODOS los tramos — incluso donde no hay dato medido. Ajusta rend = a + b·VMA
    por (tec, especie) con nuestros datos (la relación árbol→rendimiento de la bibliografía);
    donde una especie tiene pocos puntos, usa el ajuste de la tecnología escalado por el
    factor de esa especie. Devuelve {(tec,esp,tramo): m³/h}."""
    import numpy as _np
    out={}
    for tec in ['SKIDDER 6X6 GRAPPLE','TORRE']:
        dt=d[d.tec==tec]
        if len(dt)<5: continue
        # ajuste base de la tecnología (todas las especies)
        ab=_np.polyfit(dt.VMA.values, dt.rend.values, 1)
        rend_tec=dt.rend.median()
        for esp in ['PIRA','EUGL','EUNI']:
            de=dt[dt.especie==esp]
            if len(de)>=5:
                a=_np.polyfit(de.VMA.values, de.rend.values, 1)
                pred=lambda v,a=a: a[0]*v+a[1]
            else:
                # escala el ajuste de la tecnología por el nivel de la especie (o 1 si no hay dato)
                factor=(de.rend.median()/rend_tec) if len(de)>0 and rend_tec>0 else 1.0
                pred=lambda v,f=factor: (ab[0]*v+ab[1])*f
            for tr,mid in TRAMO_MID.items():
                out[(tec,esp,tr)]=max(0.0, round(float(pred(mid)),1))
    return out


ESPN={'PIRA':'Pino radiata','EUGL':'Eucalipto globulus','EUNI':'Eucalipto nitens'}
TECN={'SKIDDER 6X6 GRAPPLE':'Skidder 6×6 grapple','TORRE':'Torre de madereo'}
# Referencia BIBLIOGRÁFICA por tecnología (el "teórico" de los estudios). La literatura da
# relaciones y rangos, no una tabla por celda para nuestro rodal chileno de pendiente.
BIBLIO={
 'SKIDDER 6X6 GRAPPLE':("Eriksson & Lindroos (2014): el <b>tamaño del árbol explica ~57%</b> del "
   "rendimiento — casi se DUPLICA del árbol chico al grande. Esa es la relación que se ve en la "
   "columna p75. <b>OJO:</b> los estudios de skidder de garra dan 80-114 m³/h, pero en plantación "
   "PLANA y distancias cortas → <b>NO comparable</b> con nuestro terreno de pendiente. Vale la "
   "relación con el VMA, no el número absoluto."),
 'TORRE':("Alpine + carro garra en pino radiata (NZ): <b>25-100 m³/h</b>, mandado por el tamaño de "
   "pieza (Koller K602 = 32 m³/PMH a pieza de 2,33 m³). M11 trabaja euca (pieza chica) → cae al "
   "extremo BAJO del rango, que es lo esperado. Al ser el único yarder, su techo se mide contra sus "
   "propios mejores días."),
}

def html_tablas(cell, metas, cap, teo, d):
    """Tablas de productividad con filtro General / por faena. Columnas auto-explicativas:
    Habitual (mediana), Meta (mejor cuartil), Carga (m³/viaje), Ritmo (viajes/h), Días (muestra)."""
    THR="<tr><th class=l>Tramo VMA (m³/árbol)</th><th>Días</th><th>Habitual</th><th>Meta</th><th>Carga (m³/viaje)</th><th>Ritmo (viajes/h)</th>"
    # ---------- VISTA GENERAL (todas las faenas juntas) ----------
    gen=""
    for tec in ['SKIDDER 6X6 GRAPPLE','TORRE']:
        hay=False
        for esp in ['PIRA','EUGL','EUNI']:
            filas=[(tr,cell[(tec,esp,tr)]) for tr in LB if (tec,esp,tr) in cell]
            if not filas: continue
            hay=True
            gen+=f"<h3>{TECN.get(tec,tec)} · {ESPN.get(esp,esp)} <span style='font-weight:400;color:#889'>· REAL</span></h3><table>{THR}<th>Faenas</th></tr>"
            for tr,c in filas:
                par="1 (propio récord)" if c['faenas']==1 else str(c['faenas'])
                gen+=f"<tr><td class=l>{tr}</td><td>{c['n']}</td><td>{c['p50']}</td><td class=hi>{c['p75']}</td><td>{c['carga']}</td><td>{c['ritmo']}</td><td class=sm>{par}</td></tr>"
            gen+="</table>"
        if any((tec,e,tr) in teo for e in ['PIRA','EUGL','EUNI'] for tr in LB):
            gen+=f"<h3>{TECN.get(tec,tec)} <span style='font-weight:400;color:#1A5276'>· TEÓRICO (m³/h esperado)</span></h3><table>"
            gen+="<tr><th class=l>Tramo VMA (m³/árbol)</th><th>Pino</th><th>Euca globulus</th><th>Euca nitens</th></tr>"
            for tr in LB:
                if not any((tec,e,tr) in teo for e in ['PIRA','EUGL','EUNI']): continue
                def cel(e): return str(teo[(tec,e,tr)]) if (tec,e,tr) in teo else '—'
                gen+=f"<tr><td class=l>{tr}</td><td class=teo>{cel('PIRA')}</td><td class=teo>{cel('EUGL')}</td><td class=teo>{cel('EUNI')}</td></tr>"
            gen+="</table>"
        if hay and tec in BIBLIO:
            gen+=f"<div class=biblio><b>📖 Referencia bibliográfica ({TECN.get(tec,tec)}):</b> {BIBLIO[tec]}</div>"
    # ---------- VISTAS POR FAENA (solo los datos de esa faena) ----------
    faenas=[f for f in ['M1.1','M1.2','M1.3','M1.4','M5','M7','M9','M11'] if f in set(d.faena)]
    fv=""
    for fa in faenas:
        df=d[d.faena==fa]
        tec=df.tec.mode().iat[0]; esp=df.especie.mode().iat[0]
        fv+=f"<div data-view='{fa}' hidden><h3>{NOMBRE.get(fa,fa)} · {TECN.get(tec,tec)} · {ESPN.get(esp,esp)} <span style='font-weight:400;color:#889'>· REAL (solo esta faena)</span></h3><table>{THR}</tr>"
        for tr in LB:
            x=df[df.tramo.astype(str)==tr]
            if len(x)==0: continue
            fv+=f"<tr><td class=l>{tr}</td><td>{len(x)}</td><td>{x.rend.median():.1f}</td><td class=hi>{x.rend.quantile(.75):.1f}</td><td>{x.carga.median():.2f}</td><td>{x.ritmo.median():.2f}</td></tr>"
        fv+="</table>"
        capfa=cap.get(fa)
        if capfa: fv+=f"<div class=biblio>Ritmo del procesador de esta faena: <b>{capfa:.0f} m³/día</b> → 3 días de volteo = {capfa*3:,.0f} m³ · 2 días de madereo = {capfa*2:,.0f} m³.</div>"
        fv+="</div>"
    opts="<option value='general'>General — todas las faenas</option>"+"".join(f"<option value='{fa}'>{NOMBRE.get(fa,fa)}</option>" for fa in faenas)
    filtro=(f"<div style='margin:8px 0 14px'><label style='font-size:13px;font-weight:600;color:#3a4a5a'>Ver: "
            f"<select id=selfaena onchange=filtrar() style='font-size:14px;padding:6px 10px;border-radius:8px;border:1px solid #cfd6dd'>{opts}</select>"
            "</label></div>")
    secc=filtro+f"<div data-view='general'>{gen}</div>"+fv
    JS="<script>function filtrar(){var v=document.getElementById('selfaena').value;document.querySelectorAll('[data-view]').forEach(function(e){e.hidden=(e.getAttribute('data-view')!==v)});}</script>"
    return (f"<!doctype html><html lang=es><head><meta charset=utf-8><title>Tablas de Productividad</title><style>{CSS}"
            "h3{font-size:13px;color:#1b3a05;margin:14px 0 4px}"
            ".wrap{max-width:900px;margin:14px auto;padding:16px 20px;background:#fff;border-radius:10px;color:#233}"
            "td{background:#fff;color:#233} td.hi{background:#eaf3e0;font-weight:700;color:#2d5202} td.sm{font-size:11px;color:#667}"
            "td.teo{background:#f4f7fb;color:#1A5276;font-weight:600}"
            "h2{font-size:13px;text-transform:uppercase;letter-spacing:.5px;color:#417505;margin:18px 0 6px;border-bottom:1px solid #cdd;padding-bottom:3px}"
            ".nota{background:#fff;border:1px solid #e0e5ea;border-radius:8px;padding:10px 14px;font-size:12.5px;color:#44505e;line-height:1.5}"
            ".biblio{background:#f4f7fb;border:1px solid #d3ddea;border-left:4px solid #1A5276;border-radius:8px;padding:10px 14px;font-size:12px;color:#33475b;line-height:1.5;margin:6px 0 16px}</style></head><body>"
            f"<div class=wrap><h2>Tablas de Productividad por VMA · especie · tecnología</h2>"
            "<div class=nota>Tres miradas del rendimiento (m³/h) según el rodal — <b>tamaño de árbol (VMA)</b>, "
            "<b>especie</b> y <b>tecnología</b>:<br>"
            "• <b>REAL</b> (verde) = nuestros datos del NOC. <b>Habitual</b> = lo típico (mediana). "
            "<b>Meta</b> = lo mejor que ya se logró en ese rodal (cuartil superior).<br>"
            "• <b>TEÓRICO</b> (azul claro) = modelo esperado por tramo y especie, para las 3 especies "
            "y todos los tramos (llena donde no hay dato medido).<br>"
            "• <b>📖 Bibliografía</b> (recuadro azul) = lo que dicen los estudios de ingeniería.<br>"
            "<b>rendimiento = carga × ritmo</b> · <b>Carga</b> = m³ que mueve por viaje · <b>Ritmo</b> = viajes por "
            "hora · <b>Días</b> = cuántos días de datos hay detrás (más días = más confiable). Día por hora de "
            "inicio del turno.</div>"
            f"{secc}{JS}</div></body></html>")

LOGO=""
_p=BASE/"millalemu-logo.png"
if _p.exists(): LOGO="data:image/png;base64,"+base64.b64encode(_p.read_bytes()).decode()

CSS="""*{box-sizing:border-box;font-family:'IBM Plex Sans',Segoe UI,sans-serif}
body{margin:0;background:#eef1f6;color:#233;font-size:12px}
.sheet{max-width:194mm;margin:0 auto 14px;background:#fff;padding:12px 16px}
header{display:flex;align-items:center;gap:10px;border-bottom:3px solid #417505;padding-bottom:7px}
header img{height:34px} h1{font-size:16px;margin:0;color:#1b3a05}
.sub{color:#5a6675;font-size:11px}
h2{font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:#417505;margin:11px 0 4px;border-bottom:1px solid #cdd4dc;padding-bottom:2px}
table{width:100%;border-collapse:collapse;font-variant-numeric:tabular-nums}
td,th{border:1px solid #d3d8e0;padding:3px 7px;text-align:right;font-size:11px}
th{background:#f2f5f8;text-align:center} td.l,th.l{text-align:left}
.kpi{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:4px}
.kpi div{flex:1;min-width:82px;background:#f6f8fa;border:1px solid #e0e5ea;border-radius:5px;padding:5px 8px}
.kpi b{display:block;font-size:15px;color:#1b3a05} .kpi span{font-size:9px;color:#778;text-transform:uppercase}
.buf{display:flex;gap:8px} .buf div{flex:1;border-radius:6px;padding:7px 10px;color:#fff}
.buf .lab{font-size:9px;text-transform:uppercase;opacity:.9} .buf .big{font-size:17px;font-weight:700}
.porrep{color:#a06000;font-style:italic} .foot{color:#889;font-size:9px;margin-top:6px}
.pal{background:#fff7e6;border:1px solid #e08e0b;border-radius:5px;padding:4px 8px;color:#8a5a00;font-size:11px;margin-top:3px}
@media print{@page{size:A4;margin:8mm}.sheet{page-break-after:always;margin:0 auto}}"""

def tablero(fa,g,cell,meta_mes,cap,dia=None,volteado=None,cancha=None):
    jul=g[(g.faena==fa)]; jul=jul[jul.dia.str[:7]==g.dia.str[:7].max()]
    if dia is None: dia=jul.dia.max()
    row=jul[jul.dia==dia]; last=row.iloc[0] if len(row) else jul.iloc[-1]
    acum=jul.m3.sum(); diast=jul.dia.nunique(); real_dia=last.m3
    predio=last.predio; esp=ESP.get(last.especie,last.especie)
    cumpl=acum/meta_mes*100 if meta_mes else 0
    dh=24; plan_dia=meta_mes/dh; proy=acum/max(diast,1)*dh
    # meta rend por p75 anclado por dia
    jd=jul.dropna(subset=['VMA','rend']); jd=jd[jd.hrs>=3].copy()
    jd['tramo']=pd.cut(jd.VMA,TR,labels=LB)
    p75s=[cell[k]['p75'] for k in ((r.tec,r.especie,r.tramo) for _,r in jd.iterrows()) if k in cell]
    meta_rend=round(np.median(p75s),1) if p75s else round(jul.rend.median(),1)
    meta_dia=round(max(meta_rend*USO*HDISP, jul.rend.median()*USO*HDISP))
    carga_r=jul.carga.median(); ritmo_r=jul.ritmo.median(); rend_r=jul.rend.median()
    palanca="Carga" if (meta_rend>rend_r and carga_r< (cell and 3) ) else "Ritmo"
    def buf(m3,rojo,verde):
        if m3 is None or cap is None: return('#9aa','—','sin reporte del jefe')
        d=m3/cap; col='#c0392b' if d<rojo else ('#e08e0b' if d<verde else '#417505')
        return(col,f'{d:.1f} d',f'alcanza {d:.1f} días de procesador')
    bm=buf(cancha,1,2); bv=buf(volteado,2,3)
    return f"""<div class=sheet>
<header>{'<img src="'+LOGO+'">' if LOGO else ''}<div><h1>Tablero de Faena · {NOMBRE.get(fa,fa)}</h1>
<div class=sub>{dia} · Predio {predio} · {esp} · <b>solo productividad</b></div></div></header>
<h2>Producción General</h2>
<div class=kpi>
<div><span>Meta mes</span><b>{meta_mes:,.0f}</b></div><div><span>Avance real</span><b>{acum:,.0f}</b></div>
<div><span>Cumplimiento</span><b>{cumpl:.0f}%</b></div><div><span>Proyección</span><b>{proy:,.0f}</b></div>
<div><span>Plan diario</span><b>{plan_dia:.0f}</b></div><div><span>Real diario</span><b>{real_dia:,.0f}</b></div></div>
<h2>Buffers — que el procesador no pare {'(capacidad '+str(int(cap))+' m³/día)' if cap else '(sin capacidad cargada)'}</h2>
<div class=buf>
<div style="background:{bv[0]}"><div class=lab>Volteado adelantado</div><div class=big>{bv[1]}</div><div style="font-size:9px">{bv[2]}</div></div>
<div style="background:{bm[0]}"><div class=lab>Madera en cancha</div><div class=big>{bm[1]}</div><div style="font-size:9px">{bm[2]}</div></div></div>
<h2>Productividad — Plan (VMA+especie) vs Real</h2>
<table><tr><th class=l>Palanca</th><th>Plan (p75)</th><th>Real</th></tr>
<tr><td class=l>Rendimiento (m³/h)</td><td>{meta_rend}</td><td>{rend_r:.1f}</td></tr>
<tr><td class=l>Carga (m³/ciclo)</td><td>ref VMA</td><td>{carga_r:.2f}</td></tr>
<tr><td class=l>Ritmo (ciclos/h)</td><td>ref VMA</td><td>{ritmo_r:.2f}</td></tr></table>
<div class=pal>Meta {meta_dia:,.0f} m³/día al p75. Palanca a vigilar: <b>{palanca}</b>.</div>
<h2>Tabla diaria por proceso</h2>
<table><tr><th class=l>Proceso</th><th>Meta día</th><th>Real día</th><th>T. Perdido</th></tr>
<tr><td class=l>Volteo</td><td class=porrep>por reportar</td><td class=porrep>jefe (m³)</td><td>—</td></tr>
<tr><td class=l>Madereo</td><td class=porrep>por reportar</td><td class=porrep>jefe (m³)</td><td>—</td></tr>
<tr><td class=l>Procesado (trozado)</td><td>{meta_dia:,.0f}</td><td>{real_dia:,.0f}</td><td>NOC</td></tr>
<tr><td class=l>Clasificado</td><td>—</td><td>por grado (NOC)</td><td>—</td></tr></table>
<div class=foot>Trozado y clasificado de GeoNOC. Volteo y cancha los informa el jefe en m³ (app CMMS · Avance del día).
Buffer gris = sin reporte del jefe (nunca verde por defecto). Día = hora de inicio del turno.</div></div>"""

def main():
    df=cargar_pg()
    if df is None or len(df)==0: print("❌ sin datos PG"); return 1
    g=base_diaria(df); d,cell=tabla_p75(g); metas=metas_excel()
    # capacidad = trozado p90 del mes por faena
    mes=g.dia.str[:7].max(); gm=g[g.dia.str[:7]==mes]
    cap={fa:round(x.m3.quantile(.90)) for fa,x in gm.groupby('faena')}
    faenas=[f for f in ['M1.1','M1.2','M1.3','M1.4','M5','M7','M9','M11'] if f in set(g.faena)]
    sheets=""
    for fa in faenas:
        sheets+=tablero(fa,g,cell,metas.get(fa,METAS_DEFAULT.get(fa,0)),cap.get(fa))
    idx=f"""<div class=sheet style="page-break-after:auto"><header>{'<img src="'+LOGO+'">' if LOGO else ''}
<div><h1>Tableros de Faena — {mes}</h1><div class=sub>Módulo de productividad · pre-llenado para la pizarra · {len(faenas)} faenas</div></div></header>
<div class=foot>Cada hoja se imprime aparte (A4). El jefe traspasa a la pizarra y agrega lo cualitativo (SSO, riesgos, IAP) que va fuera de este módulo.</div></div>"""
    html=f"<!doctype html><html lang=es><head><meta charset=utf-8><title>Tableros de Faena {mes}</title><style>{CSS}</style></head><body>{idx}{sheets}</body></html>"
    out=BASE/"Tableros_Faena.html"; out.write_text(html,encoding="utf-8")
    print(f"✅ Tableros_Faena.html — {len(faenas)} faenas, mes {mes}, {len(html):,} bytes")
    print(f"   capacidades (trozado p90): {cap}")
    # Tablas de productividad teóricas (VMA×especie) — página de consulta.
    tablas=html_tablas(cell, metas, cap, teorico(d), d)
    (BASE/"Tablas_Productividad.html").write_text(tablas,encoding="utf-8")
    print(f"✅ Tablas_Productividad.html — {len(cell)} celdas VMA×especie, {len(tablas):,} bytes")
    return 0

if __name__=="__main__":
    sys.exit(main())
