#!/usr/bin/env python3
"""
compute_kpis.py — KPIs Uso / Ritmo / Carga / VMA por faena.
Lee ProductividadGenerico.csv + TiemposPerdidos.csv (los que ya baja el pipeline,
reporte PG del NOC — trae árboles madereados) y produce:
  - kpis.json          (por faena: Uso/Ritmo/Carga/VMA + proyección de cierre)
  - tm_por_faena.json  (top causas de tiempo perdido por faena)

Descomposición: Productividad (m³/h disponible) = Uso × Ritmo × Carga
  Uso   = horas efectivas / horas DISPONIBLES (turno − 1 h almuerzo − 0,5 h pausa activa)
  Ritmo = ciclos / hora efectiva
  Carga = m³ / ciclo
  VMA   = m³ / árboles madereados (rodal asignado; driver estructural)

Reusa la lógica autoritativa de GENERAR_RESUMEN.py (folio-dedup, TEAM_MAP, CLASIF).
Auto-detecta el mes en curso; marca parcial=True si el mes no está completo.
"""
import pandas as pd, numpy as np, json, calendar
from pathlib import Path
import sys

BASE = Path(__file__).parent
sys.path.insert(0, str(BASE))
from normalizar_produccion import normalizar

TEAM_MAP = {'S123':'M1.1','S58':'M1.2','S223':'M1.3','S246':'M1.4','MG5':'M5','TEA02':'M7','TEA08':'M9','T125':'M11','TEA30':'M1.3'}
NOMBRE   = {'M1.1':'Millalemu 1.1','M1.2':'Millalemu 1.2','M1.3':'Millalemu 1.3','M1.4':'Millalemu 1.4','M5':'Millalemu 5','M7':'Millalemu 7','M9':'Millalemu 9','M11':'Millalemu 11'}
TERRESTRE = {'M1.1','M1.2','M1.3','M1.4'}   # ZONA (para el display de tablas): Terrestre vs Aéreo
# TECNOLOGÍA (para el potencial): skidder/madereo vs torre. M7/M9/M5 son skidder aunque estén en zona aérea;
# el único sistema distinto es M11 (torre). El potencial se compara solo dentro de la misma tecnología.
GRUPO_TEC = {'M1.1':'Skidder','M1.2':'Skidder','M1.3':'Skidder','M1.4':'Skidder',
             'M5':'Skidder','M7':'Skidder','M9':'Skidder','M11':'Torre'}
# Metas: mismo origen que GENERAR_HTML/GENERAR_RESUMEN (Excel CONFIGURACIÓN, celdas E24-E31),
# con los mismos defaults, para que el cumplimiento calce EXACTO con el dashboard de producción.
_TEAMS_FULL = ['Millalemu 1.1','Millalemu 1.2','Millalemu 1.3','Millalemu 1.4','Millalemu 5','Millalemu 7','Millalemu 9','Millalemu 11']
_ABR = {n: n.replace('Millalemu ', 'M') for n in _TEAMS_FULL}
_METAS_NOMBRE = {'Millalemu 1.1':8000.0,'Millalemu 1.2':8000.0,'Millalemu 1.3':8000.0,'Millalemu 1.4':8000.0,
                 'Millalemu 5':4500.0,'Millalemu 7':8060.0,'Millalemu 9':5940.0,'Millalemu 11':6000.0}
_EXCEL_META = Path(__file__).parent / "Dashboard_CosechaForestal.xlsx"
if _EXCEL_META.exists():
    try:
        from openpyxl import load_workbook
        _wb = load_workbook(str(_EXCEL_META), data_only=True)
        if "CONFIGURACIÓN" in _wb.sheetnames:
            _ws = _wb["CONFIGURACIÓN"]
            for _i, _t in enumerate(_TEAMS_FULL):
                _v = _ws.cell(24 + _i, 5).value
                if _v: _METAS_NOMBRE[_t] = float(_v)
        _wb.close()
    except Exception as _e:
        print(f"⚠️  metas Excel: {_e}; usando defaults")
METAS = {_ABR[n]: v for n, v in _METAS_NOMBRE.items()}
CLASIF = {1:'Mantención',2:'Mantención',3:'Mantención',4:'Mantención',5:'Mantención',6:'Mantención',7:'Mantención',8:'Mantención',10:'Mantención',12:'Mantención',58:'Mantención',69:'Mantención',
    13:'Operacional',14:'Operacional',15:'Operacional',20:'Operacional',21:'Operacional',22:'Operacional',31:'Operacional',32:'Operacional',33:'Operacional',38:'Operacional',41:'Operacional',
    16:'Proceso',17:'Proceso',18:'Proceso',25:'Proceso',26:'Proceso',61:'Proceso',65:'Proceso',66:'Proceso',68:'Proceso',42:'Programado',43:'Programado'}
PAUSA_DIA = 1.5
MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
_FERIADOS_IRR = {'01-01','05-01','09-18','09-19','12-25'}

# Preferir Manual.csv/Base2NOC no aplica aquí: usamos el PG (única fuente con árboles).
CSV_PROD = BASE / "ProductividadGenerico.csv"
if not CSV_PROD.exists():                       # fallback al alias Base2NOC si el pipeline renombró
    CSV_PROD = BASE / "Base2NOC.csv"
CSV_TM = BASE / "TiemposPerdidos.csv"


def _num(s):
    return pd.to_numeric(s.astype(str).str.replace(',', '.'), errors='coerce')


def _hora_seg(v):
    """Convierte hora (ISO datetime / 'HH:MM[:SS]' / número) a segundos desde medianoche."""
    if v is None or v == '':
        return 0
    s = str(v)
    try:
        if 'T' in s:
            from datetime import datetime as _dt
            t = _dt.fromisoformat(s)
            return t.hour * 3600 + t.minute * 60 + t.second
        if ':' in s:
            p = s.split(':')
            return int(p[0]) * 3600 + int(p[1]) * 60 + (int(p[2]) if len(p) > 2 else 0)
        return int(float(s))
    except Exception:
        return 0


def _fetch_pg_api():
    """Baja el reporte Productividad Genérico del NOC (única fuente con árboles
    madereados; Base2NOC no los trae) → DataFrame con las columnas que espera el
    cómputo. Devuelve None si falla (para caer al CSV)."""
    try:
        import descargar_noc_api as noc
        import requests
        from datetime import datetime
        hoy = datetime.now()
        fi = hoy.replace(day=1).strftime('%Y-%m-%d')
        ff = hoy.strftime('%Y-%m-%d')
        s = requests.Session(); s.verify = False
        s.headers.update({'User-Agent': 'Mozilla/5.0'})
        tok = noc.obtener_token_arcgis(s)
        try:
            noc.establecer_sesion(s, tok)
        except Exception:
            pass
        datos = noc.descargar_reporte(s, tok, 'PG', fi, ff)
        if not datos:
            return None
        rows = [{
            'Equipo': str(r.get('equipo', '')).strip(),
            'Fecha NOC': str(r.get('hora_inicio', ''))[:10],
            'Volumen SSC PU': r.get('m3ssc_pu') or 0,
            'Volumen SSC AS': r.get('m3ssc_as') or 0,
            'Tiempo Efectivo': r.get('tiempo_efectivo') or 0,
            'Hora Inicio': _hora_seg(r.get('hora_inicio')),
            'Hora Fin': _hora_seg(r.get('hora_fin')),
            'Número Ciclos': r.get('numero_ciclos') or 0,
            'Árboles Madereados': r.get('arboles_madereados') or 0,
            'Número Noc': r.get('numero_noc') or 0,
        } for r in datos]
        print(f"   🌲 PG API: {len(rows)} folios con árboles madereados")
        return pd.DataFrame(rows)
    except Exception as e:
        print(f"   ⚠️  PG API no disponible ({e}); uso CSV local (VMA quedará en 0 si no hay árboles)")
        return None


def _turnos_schedule(anio, mes):
    """Genera el calendario de turnos 7×7 del mes desde turnos_config.json (rotación
    CORRELATIVA — se auto-continúa para cualquier mes, no hay que recargar Excel).
    Ciclo de 14 días desde la referencia; pos<7 → jefeA, si no → jefeB.
    Devuelve {faena: [{jefe, full:set, half:set}]}. La config se derivó del Excel una vez."""
    try:
        from datetime import date
        cfg_f = BASE / "turnos_config.json"
        if not cfg_f.exists():
            return {}
        cfg = json.loads(cfg_f.read_text(encoding='utf-8'))
        ref = date.fromisoformat(cfg.get('ref', '2026-07-01')); ciclo = int(cfg.get('ciclo_dias', 14))
        mitad = ciclo // 2
        dm = calendar.monthrange(anio, mes)[1]
        out = {}
        for fa, c in cfg.get('faenas', {}).items():
            da, db = set(), set()
            for d in range(1, dm + 1):
                pos = (c['posR'] + (date(anio, mes, d) - ref).days) % ciclo
                (da if pos < mitad else db).add(d)
            out[fa] = [{'jefe': c['jefeA'], 'full': da, 'half': set()},
                       {'jefe': c['jefeB'], 'full': db, 'half': set()}]
        return out
    except Exception as e:
        print(f"   ⚠️  turnos config: {e}")
        return {}


def main():
    prod = _fetch_pg_api()
    if prod is None:
        prod = normalizar(pd.read_csv(CSV_PROD, sep=';', encoding='utf-8-sig'))
    for c in ['Volumen SSC PU', 'Volumen SSC AS']:
        if c not in prod.columns:
            prod[c] = 0
        prod[c] = _num(prod[c])
    prod['Vol'] = prod['Volumen SSC PU'].fillna(0) + prod['Volumen SSC AS'].fillna(0)
    prod['Arb'] = _num(prod['Árboles Madereados']).fillna(0) if 'Árboles Madereados' in prod.columns else 0.0
    _te = _num(prod['Tiempo Efectivo']).fillna(0)
    _pos = _te[_te > 0]
    prod['HrsEf'] = _te / (3600 if (len(_pos) > 0 and _pos.median() > 1000) else 60)
    for c in ['Hora Inicio', 'Hora Fin']:
        prod[c] = _num(prod[c]).fillna(0)
    prod['Turno_seg'] = (prod['Hora Fin'] - prod['Hora Inicio']).clip(lower=0)
    prod['Ciclos'] = _num(prod['Número Ciclos']).fillna(0)
    prod['Team'] = prod['Equipo'].map(TEAM_MAP)
    prod['Fecha_dt'] = pd.to_datetime(prod['Fecha NOC'], dayfirst=True, errors='coerce')

    ultimo = prod['Fecha_dt'].max()
    MES, ANIO = int(ultimo.month), int(ultimo.year)
    TURNOS = _turnos_schedule(ANIO, MES)   # rotación 7×7 correlativa para el mes en curso
    prod_mes = prod[(prod['Fecha_dt'].dt.month == MES) & (prod['Fecha_dt'].dt.year == ANIO) & prod['Team'].notna()].copy()

    tm = pd.read_csv(CSV_TM, sep=';', encoding='utf-8-sig')
    tm['Tiempo (Min)'] = _num(tm['Tiempo (Min)']).fillna(0)
    tm['Clasif'] = tm['Código Tiempo Perdido'].map(CLASIF).fillna('Operacional')
    tm = tm[tm['Clasif'] != 'Programado']
    tm['Team'] = tm['Código Equipo'].map(TEAM_MAP)
    tm['Fecha_dt'] = pd.to_datetime(tm['Fecha'], dayfirst=True, errors='coerce')
    tm_mes = tm[(tm['Fecha_dt'].dt.month == MES) & tm['Team'].notna()].copy()

    DM = calendar.monthrange(ANIO, MES)[1]
    ult_dia = int(prod_mes['Fecha_dt'].dt.day.max())
    DT = sum(1 for d in range(1, DM + 1) if f"{MES:02d}-{d:02d}" not in _FERIADOS_IRR)
    DD = sum(1 for d in range(1, ult_dia + 1) if f"{MES:02d}-{d:02d}" not in _FERIADOS_IRR)
    DR = max(DT - DD, 0)
    parcial = ult_dia < DM

    faenas, tmf = [], {}
    for t in ['M1.1','M1.2','M1.3','M1.4','M5','M7','M9','M11']:
        d = prod_mes[prod_mes['Team'] == t]
        if len(d) == 0:
            continue
        vol, arb = d['Vol'].sum(), d['Arb'].sum()
        g = d.groupby('Número Noc')
        hrs = g['HrsEf'].first().sum(); turno_bruto = g['Turno_seg'].first().sum() / 3600.0; ciclos = g['Ciclos'].first().sum()
        ndias = d['Fecha_dt'].dt.day.nunique(); nfolios = d['Número Noc'].nunique()
        disp = max(turno_bruto - PAUSA_DIA * nfolios, 0.1)
        uso = hrs / disp if disp > 0 else 0
        ritmo = ciclos / hrs if hrs > 0 else 0
        carga = vol / ciclos if ciclos > 0 else 0
        vma = vol / arb if arb > 0 else 0
        prom = vol / ndias if ndias > 0 else 0
        proy = vol + prom * DR
        tmt = tm_mes[tm_mes['Team'] == t]
        # Detalle diario para la cartilla (jornada a jornada)
        dias_det = []
        for _day, _x in d.groupby(d['Fecha_dt'].dt.day):
            _gx = _x.groupby('Número Noc'); _v = _x['Vol'].sum(); _he = _gx['HrsEf'].first().sum(); _ci = _gx['Ciclos'].first().sum()
            dias_det.append({'d': int(_day), 'vol': round(_v), 'hrs': round(_he, 1), 'cic': int(_ci), 'rend': round(_v / _he, 1) if _he else 0})
        dias_det.sort(key=lambda z: z['d'])
        # Anotar el jefe/turno (0=jefeA, 1=jefeB) de cada jornada — para colorear la grilla.
        _tsched = TURNOS.get(NOMBRE[t], [])
        jefes_orden = [_tsched[0]['jefe'], _tsched[1]['jefe']] if len(_tsched) == 2 else []
        for x in dias_det:
            x['turno'] = (0 if _tsched and x['d'] in _tsched[0]['full']
                          else (1 if len(_tsched) == 2 and x['d'] in _tsched[1]['full'] else None))
        # Atribución por turno/jefe (7×7): cada día se asigna al jefe que estaba.
        turnos_faena = []
        for i_t, jt in enumerate(_tsched):
            jv = jh = jc = jd = 0.0
            for x in dias_det:
                w = 1.0 if x['d'] in jt['full'] else (0.5 if x['d'] in jt['half'] else 0.0)
                if w:
                    jv += x['vol'] * w; jh += x['hrs'] * w; jc += x['cic'] * w; jd += w
            if jd > 0:
                turnos_faena.append(dict(jefe=jt['jefe'], turno=i_t, dias=round(jd, 1), m3=round(jv), hrs=round(jh, 1),
                    ciclos=int(round(jc)), m3_dia=round(jv / jd, 1), m3_hr=round(jv / jh, 1) if jh else 0))
        turnos_faena.sort(key=lambda z: -z['m3_dia'])   # mejor turno primero
        faenas.append(dict(team=t, nombre=NOMBRE[t], tipo=('Terrestre' if t in TERRESTRE else 'Aéreo'),
            grupo_tec=GRUPO_TEC.get(t, 'Skidder'), dias_detalle=dias_det, turnos=turnos_faena, jefes_orden=jefes_orden,
            meta_m3=METAS[t], vol_m3=round(vol, 1), cumpl_pct=round(vol / METAS[t] * 100, 1),
            proy_cierre_m3=round(proy, 0), proy_cumpl_pct=round(proy / METAS[t] * 100, 1), prom_diario_m3=round(prom, 1),
            dias=int(ndias), folios=int(nfolios), arboles=int(arb),
            hrs_efectivas=round(hrs, 1), hrs_disponible=round(disp, 1), hrs_turno_bruto=round(turno_bruto, 1), ciclos=int(ciclos),
            uso_pct=round(uso * 100, 1), ritmo_ciclos_h=round(ritmo, 2), carga_m3_ciclo=round(carga, 3),
            vma_m3_arbol=round(vma, 3), arboles_por_ciclo=round(arb / ciclos, 1) if ciclos else None,
            prod_m3_h_disp=round(vol / disp, 2) if disp > 0 else 0, prod_m3_h_efec=round(vol / hrs, 2) if hrs > 0 else 0,
            tm_mant_min=int(tmt[tmt['Clasif'] == 'Mantención']['Tiempo (Min)'].sum()),
            tm_oper_min=int(tmt[tmt['Clasif'] == 'Operacional']['Tiempo (Min)'].sum()),
            tm_proc_min=int(tmt[tmt['Clasif'] == 'Proceso']['Tiempo (Min)'].sum())))
        top = tmt.groupby('Descripción')['Tiempo (Min)'].sum().sort_values(ascending=False).head(6)
        tmf[NOMBRE[t]] = {'total_min': int(tmt['Tiempo (Min)'].sum()), 'top_causas': {k: int(v) for k, v in top.items()}}

    for tec in ['Skidder', 'Torre']:   # potencial = mejor de tu MISMA tecnología
        grp = [f for f in faenas if f['grupo_tec'] == tec]
        if not grp:
            continue
        # Potencial por palanca = máx(mejor mes propio, líder de su tipo). Con 1 mes de
        # datos = líder de su tipo (referencia MENSUAL, confiable; el mejor-día es ruidoso).
        bu = min(max(f['uso_pct'] for f in grp), 105)  # cap Uso a 105% (sobre-jornada no es potencial)
        br = max(f['ritmo_ciclos_h'] for f in grp); bc = max(f['carga_m3_ciclo'] for f in grp)
        for f in grp:
            uso_ref = min(f['uso_pct'], 105)
            gaps = {'Uso': (bu - uso_ref) / bu if bu else 0, 'Ritmo': (br - f['ritmo_ciclos_h']) / br if br else 0, 'Carga': (bc - f['carga_m3_ciclo']) / bc if bc else 0}
            gaps = {k: max(v, 0) for k, v in gaps.items()}
            f['palanca_limitante'] = max(gaps, key=gaps.get); f['brecha_palanca_pct'] = round(gaps[f['palanca_limitante']] * 100, 1)
            f['benchmark_tipo'] = {'uso_pct': round(bu, 1), 'ritmo_ciclos_h': round(br, 2), 'carga_m3_ciclo': round(bc, 3)}
            f['gaps_palanca'] = {k: round(v * 100, 1) for k, v in gaps.items()}
            # Desafío = cerrar la palanca limitante a su potencial (una sola palanca), el resto del mes.
            pal = f['palanca_limitante']; uf = uso_ref / 100; rit = f['ritmo_ciclos_h']; car = f['carga_m3_ciclo']
            new_prod = (bu / 100) * rit * car if pal == 'Uso' else (uf * br * car if pal == 'Ritmo' else uf * rit * bc)
            f['potencial_prod_m3h'] = round(new_prod, 2)
            proy_pot = f['vol_m3'] + new_prod * 10.5 * DR
            f['proy_potencial_pct'] = round(proy_pot / f['meta_m3'] * 100, 1) if f['meta_m3'] else 0
            f['oportunidad_palanca_m3'] = round(max(new_prod - f['prod_m3_h_disp'], 0) * 10.5 * DR, 0)
            # ¿el potencial de tipo (todas las palancas al mejor) llega a la meta?
            pot_full = (bu / 100) * br * bc * 10.5 * DR + f['vol_m3']
            if len(grp) == 1:   # singleton (M11 torre) — sin par de su tecnología
                f['meta_status'] = 'solo'
            else:
                f['meta_status'] = ('gestion' if f['proy_potencial_pct'] >= 100 else ('multipalanca' if pot_full >= f['meta_m3'] else 'rodal'))

    acum = sum(f['vol_m3'] for f in faenas); meta = sum(f['meta_m3'] for f in faenas)
    proy = sum(f['proy_cierre_m3'] for f in faenas); arbt = sum(f['arboles'] for f in faenas)
    out = dict(mes=f"{MESES[MES]} {ANIO}", mes_num=MES, anio=ANIO, parcial=bool(parcial), ultimo_dia=ult_dia,
        dias_con_datos=DD, dias_habiles_mes=DT, dias_restantes=DR,
        totales=dict(acum_m3=round(acum, 0), meta_m3=meta, cumpl_pct=round(acum / meta * 100, 1) if meta else 0,
            proy_cierre_m3=round(proy, 0), proy_cumpl_pct=round(proy / meta * 100, 1) if meta else 0,
            vma_ponderado=round(acum / arbt, 3) if arbt else 0, arboles=arbt,
            desafio=dict(
                por_gestion=sum(1 for f in faenas if f.get('meta_status') == 'gestion'),
                multipalanca=sum(1 for f in faenas if f.get('meta_status') == 'multipalanca'),
                requiere_rodal=sum(1 for f in faenas if f.get('meta_status') == 'rodal'),
                sin_par=sum(1 for f in faenas if f.get('meta_status') == 'solo'))),
        fuente="NOC · Productividad Genérico (Uso/Ritmo/Carga + VMA, una sola fuente)", faenas=faenas)
    (BASE / "kpis.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')
    (BASE / "tm_por_faena.json").write_text(json.dumps(tmf, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"✅ KPIs {out['mes']} ({'parcial' if parcial else 'cerrado'}) · {len(faenas)} faenas · acum {acum:.0f}/{meta} = {acum/meta*100:.1f}% · proy {proy/meta*100:.0f}% · VMA {out['totales']['vma_ponderado']}")


if __name__ == "__main__":
    main()
