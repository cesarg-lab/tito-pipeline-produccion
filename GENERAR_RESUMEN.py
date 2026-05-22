"""
GENERAR_RESUMEN.py — v1.0 (2026-04-16)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Genera el texto del resumen diario completo para WhatsApp.
Análisis detallado: producción, TM, disponibilidad,
rendimiento, proyección, tendencias y recomendaciones.

Salida: resumen_diario.txt (texto con formato WhatsApp)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
import pandas as pd
import numpy as np
import calendar
import sys
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent
CSV_PROD = BASE_DIR / "Base2NOC.csv"
CSV_TM   = BASE_DIR / "TiemposPerdidos.csv"
import sys as _sys
_sys.path.insert(0, str(BASE_DIR))
from normalizar_produccion import normalizar  # noqa: E402

TEAM_MAP = {
    'S123':'M1.1','S58':'M1.2','S223':'M1.3','S246':'M1.4',
    'MG5':'M5','TEA02':'M7','TEA08':'M9','T125':'M11','TEA30':'M1.3',
}
GRUPOS = {
    'general':   {'teams': ['M1.1','M1.2','M1.3','M1.4','M5','M7','M9','M11'],
                  'titulo': 'CONTROL DE COSECHA FORESTAL',
                  'subtitulo': 'Forestal Millalemu',
                  'output': 'resumen_diario.txt'},
    'aereo':     {'teams': ['M5','M7','M9','M11'],
                  'titulo': 'MILLALEMU AÉREO',
                  'subtitulo': 'Cosecha Aérea · Forestal Millalemu',
                  'output': 'resumen_diario_aereo.txt'},
    'terrestre': {'teams': ['M1.1','M1.2','M1.3','M1.4'],
                  'titulo': 'MILLALEMU TERRESTRE',
                  'subtitulo': 'Cosecha Terrestre · Forestal Millalemu',
                  'output': 'resumen_diario_terrestre.txt'},
}
# Leer argumento --grupo
GRUPO = 'general'
for i, a in enumerate(sys.argv):
    if a == '--grupo' and i+1 < len(sys.argv):
        g = sys.argv[i+1].lower()
        if g in GRUPOS: GRUPO = g
TEAMS = GRUPOS[GRUPO]['teams']
TITULO = GRUPOS[GRUPO]['titulo']
SUBTITULO = GRUPOS[GRUPO]['subtitulo']
OUTPUT = BASE_DIR / GRUPOS[GRUPO]['output']

METAS = {'M1.1':7000,'M1.2':7000,'M1.3':7000,'M1.4':7000,
         'M5':4500,'M7':7000,'M9':7000,'M11':6000}

CLASIF = {
    1:'Mantención',2:'Mantención',3:'Mantención',4:'Mantención',
    5:'Mantención',6:'Mantención',58:'Mantención',
    13:'Operacional',14:'Operacional',18:'Operacional',38:'Operacional',
    42:'Proceso',43:'Proceso'
}

EXCEL = BASE_DIR / "Dashboard_CosechaForestal.xlsx"
if EXCEL.exists():
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(EXCEL), data_only=True)
        if "CONFIGURACIÓN" in wb.sheetnames:
            ws = wb["CONFIGURACIÓN"]
            team_list = ['Millalemu 1.1','Millalemu 1.2','Millalemu 1.3','Millalemu 1.4',
                         'Millalemu 5','Millalemu 7','Millalemu 9','Millalemu 11']
            abrev_map = {t: t.replace('Millalemu ', 'M') for t in team_list}
            for i, t in enumerate(team_list):
                v = ws.cell(24 + i, 5).value
                if v: METAS[abrev_map[t]] = float(v)
        wb.close()
    except Exception:
        pass

DOW_NAMES = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']

def fmt(n):
    if n is None or (isinstance(n, float) and np.isnan(n)):
        return '—'
    return f"{n:,.0f}".replace(',','.')

def generate():
    # ══════════════════════════════════════════════════════
    # CARGAR DATOS
    # ══════════════════════════════════════════════════════
    prod = pd.read_csv(CSV_PROD, sep=';', encoding='utf-8-sig')
    prod = normalizar(prod)
    tm   = pd.read_csv(CSV_TM,   sep=';', encoding='utf-8-sig')

    for c in ['Volumen SSC PU','Volumen SSC AS']:
        prod[c] = pd.to_numeric(prod[c].astype(str).str.replace(',','.'), errors='coerce')
    prod['Vol'] = prod['Volumen SSC PU'].fillna(0) + prod['Volumen SSC AS'].fillna(0)

    # ── Manual.csv DESACTIVADO (Cesar 2026-05-02) — Selenium fuente única ────
    from pathlib import Path
    MANUAL_CSV = Path("/__manual_desactivado__")  # path imposible
    vol_oficial_diario = None
    if MANUAL_CSV.exists():
        try:
            mdf = pd.read_csv(MANUAL_CSV, sep=';', encoding='utf-8-sig', decimal=',', thousands='.')
            vol_oficial_diario = {}
            for _, r in mdf.iterrows():
                eq_full = str(r.get('Equipo','')).strip()
                if '-' not in eq_full: continue
                sigla = eq_full.split('-')[0]
                team = TEAM_MAP.get(sigla)
                if not team: continue
                for d in range(1, 32):
                    v = r.get(str(d))
                    if pd.notna(v) and v != 0:
                        vol_oficial_diario[(team, int(d))] = float(v)
            # Filtrar por equipos del grupo
            vol_oficial_diario = {(t, d): v for (t, d), v in vol_oficial_diario.items() if t in TEAMS}
            print(f"   📋 Manual.csv cargado: {len(vol_oficial_diario)} día/equipo oficiales (grupo: {GRUPO})")
        except Exception as e:
            print(f"   ⚠️  No se pudo leer Manual.csv: {e}")
            vol_oficial_diario = None
    # Tiempo Efectivo: detección automática de unidad (Selenium=segundos, API=minutos)
    _te_raw = pd.to_numeric(prod['Tiempo Efectivo'].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    prod['HrsEf'] = _te_raw / (3600 if _te_raw.max() > 1000 else 60)
    for c in ['Hora Inicio', 'Hora Fin']:
        prod[c] = pd.to_numeric(prod[c].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    prod['Turno_seg'] = (prod['Hora Fin'] - prod['Hora Inicio']).clip(lower=0)
    prod['Team'] = prod['Equipo'].map(TEAM_MAP)
    prod['Fecha_dt'] = pd.to_datetime(prod['Fecha NOC'], dayfirst=True, errors='coerce')
    prod['Dia'] = prod['Fecha_dt'].dt.day

    tm['Tiempo (Min)'] = pd.to_numeric(tm['Tiempo (Min)'].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    tm['Clasif'] = tm['Código Tiempo Perdido'].map(CLASIF).fillna('Operacional')
    tm['Team'] = tm['Código Equipo'].map(TEAM_MAP)
    tm['Fecha_dt'] = pd.to_datetime(tm['Fecha'], dayfirst=True, errors='coerce')
    tm = tm.dropna(subset=['Fecha_dt'])
    tm['Dia'] = tm['Fecha_dt'].dt.day

    # Filtrar mes actual
    ultimo = prod['Fecha_dt'].max()
    MES = ultimo.month; ANIO = ultimo.year
    DM = calendar.monthrange(ANIO, MES)[1]
    MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

    prod_mes = prod[prod['Fecha_dt'].dt.month == MES]
    tm_mes = tm[tm['Fecha_dt'].dt.month == MES]
    # Filtrar por equipos del grupo seleccionado
    prod_mes = prod_mes[prod_mes['Team'].isin(TEAMS)]
    tm_mes = tm_mes[tm_mes['Team'].isin(TEAMS)]
    DD = len(prod_mes['Dia'].unique())
    DR = max(DM - DD, 1)

    prod_dia = prod_mes[prod_mes['Fecha_dt'] == ultimo]
    tm_dia = tm_mes[tm_mes['Fecha_dt'] == ultimo]

    fecha_str = ultimo.strftime('%d/%m/%Y')
    dow = DOW_NAMES[ultimo.weekday()]

    # ══════════════════════════════════════════════════════
    # CALCULAR DATOS DEL DÍA
    # ══════════════════════════════════════════════════════
    dia_data = []
    vol_dia_total = 0
    dia_actual = int(ultimo.day)
    # Si el Manual no tiene datos del día actual para ningún equipo, está desactualizado
    # → fallback al detallado para el día actual
    manual_cubre_dia = False
    if vol_oficial_diario is not None:
        manual_cubre_dia = any(dia == dia_actual for (_, dia) in vol_oficial_diario.keys())
        if not manual_cubre_dia:
            print(f"   ℹ️  Manual.csv no tiene datos del día {dia_actual} — usando detallado para el día (acumulado mes sigue del Manual)")

    for t in TEAMS:
        td_p = prod_dia[prod_dia['Team'] == t]
        td_tm = tm_dia[tm_dia['Team'] == t]
        # Vol del día: Manual si tiene ese día, sino detallado
        if vol_oficial_diario is not None and manual_cubre_dia:
            vol = vol_oficial_diario.get((t, dia_actual), 0)
        else:
            vol = td_p['Vol'].sum()
        hrs = td_p['HrsEf'].sum()
        rend = vol / hrs if hrs > 0 else 0
        turno_seg = td_p['Turno_seg'].sum()
        turno_min = turno_seg / 60

        tm_mant = td_tm[td_tm['Clasif'] == 'Mantención']['Tiempo (Min)'].sum()
        tm_oper = td_tm[td_tm['Clasif'] == 'Operacional']['Tiempo (Min)'].sum()
        tm_proc = td_tm[td_tm['Clasif'] == 'Proceso']['Tiempo (Min)'].sum()
        tm_total = tm_mant + tm_oper + tm_proc

        disp = (1 - tm_mant / turno_min) * 100 if turno_min > 0 else 0
        disp = max(0, min(100, disp))

        vol_dia_total += vol
        dia_data.append({
            'team': t, 'vol': vol, 'hrs': hrs, 'rend': rend,
            'turno_min': turno_min, 'tm_mant': tm_mant, 'tm_oper': tm_oper,
            'tm_proc': tm_proc, 'tm_total': tm_total, 'disp': disp,
        })

    # Acumulado mensual: Manual donde tenga, Base2NOC para los días que Manual no cubre
    if vol_oficial_diario is not None:
        # Empezar con totales por equipo desde Base2NOC
        vol_mes_total = prod_mes.groupby('Team')['Vol'].sum().to_dict()
        # Días cubiertos por Manual: reemplazar el aporte del detallado por el oficial
        dias_manual = set(d for (_, d) in vol_oficial_diario.keys())
        for t in TEAMS:
            # Restar lo que aportó Base2NOC en días cubiertos por Manual
            aporte_b2n = prod_mes[(prod_mes['Team']==t) & (prod_mes['Dia'].isin(dias_manual))]['Vol'].sum()
            # Sumar lo que dice Manual para esos días
            aporte_man = sum(v for (tt,d), v in vol_oficial_diario.items() if tt==t)
            base = vol_mes_total.get(t, 0)
            vol_mes_total[t] = base - aporte_b2n + aporte_man
        vol_mes_total = pd.Series(vol_mes_total)
        # DD: total de días distintos (Manual ∪ Base2NOC)
        dias_b2n = set(int(d) for d in prod_mes['Dia'].unique())
        DD_oficial = len(dias_manual | dias_b2n)
    else:
        vol_mes_total = prod_mes.groupby('Team')['Vol'].sum()
        DD_oficial = DD
    acum_total = vol_mes_total.sum()
    prom_dia_total = acum_total / DD_oficial if DD_oficial > 0 else 0
    delta = ((vol_dia_total / prom_dia_total) - 1) * 100 if prom_dia_total > 0 else 0

    # Ranking del día
    dia_data.sort(key=lambda x: x['vol'], reverse=True)

    # TM del día por categoría
    tm_dia_mant = sum(d['tm_mant'] for d in dia_data)
    tm_dia_oper = sum(d['tm_oper'] for d in dia_data)
    tm_dia_proc = sum(d['tm_proc'] for d in dia_data)
    tm_dia_total = tm_dia_mant + tm_dia_oper + tm_dia_proc

    # Top causas TM del día
    top_causas = tm_dia.groupby('Descripción')['Tiempo (Min)'].sum().sort_values(ascending=False).head(5)

    # Proyección mensual
    meta_total = sum(METAS[t] for t in TEAMS if t in METAS)
    proy_total = acum_total + prom_dia_total * DR
    pct_total = proy_total / meta_total * 100 if meta_total > 0 else 0
    brecha_total = proy_total - meta_total

    # Tendencia (últimos 3 días) — usar Manual si hay
    ultimos_dias = sorted(prod_mes['Dia'].unique())[-3:]
    tendencia = []
    for d in ultimos_dias:
        d_int = int(d)
        if vol_oficial_diario is not None:
            v = sum(val for (team, dia), val in vol_oficial_diario.items() if dia == d_int)
        else:
            v = prod_mes[prod_mes['Dia'] == d]['Vol'].sum()
        tendencia.append((d, v))

    # Alertas
    equipos_baja_disp = [d for d in dia_data if d['disp'] < 80 and d['turno_min'] > 0]
    equipos_baja_disp.sort(key=lambda x: x['disp'])

    equipos_bajo_rend = [d for d in dia_data if d['rend'] < 15 and d['hrs'] > 0]
    equipos_bajo_rend.sort(key=lambda x: x['rend'])

    # Proyección por equipo
    proy_equipo = []
    for t in TEAMS:
        acum = vol_mes_total.get(t, 0)
        meta = METAS[t]
        prom = acum / DD if DD > 0 else 0
        proy = acum + prom * DR
        pct = proy / meta * 100 if meta > 0 else 0
        proy_equipo.append({'team': t, 'acum': acum, 'meta': meta, 'proy': proy, 'pct': pct})

    proy_equipo.sort(key=lambda x: x['pct'])

    # ══════════════════════════════════════════════════════
    # CONSTRUIR MENSAJE
    # ══════════════════════════════════════════════════════
    L = []

    # ── ENCABEZADO ──
    L.append(f"━━━━━━━━━━━━━━━━━━━━━")
    L.append(f"*{TITULO}*")
    L.append(f"_{SUBTITULO}_")
    L.append(f"📅 {dow} {fecha_str}")
    L.append(f"━━━━━━━━━━━━━━━━━━━━━")
    L.append("")

    # ── PRODUCCIÓN DEL DÍA ──
    delta_sign = '+' if delta >= 0 else ''
    delta_emoji = '📈' if delta >= 0 else '📉'
    L.append(f"*📦 PRODUCCIÓN DEL DÍA*")
    L.append(f"Total: *{fmt(vol_dia_total)} m³* {delta_emoji} {delta_sign}{delta:.0f}% vs promedio")
    L.append(f"Promedio mes: {fmt(prom_dia_total)} m³/día")
    L.append("")

    # Ranking por equipo
    L.append(f"*Ranking:*")
    for i, d in enumerate(dia_data):
        if d['vol'] == 0 and d['hrs'] == 0:
            continue
        medal = ['🥇','🥈','🥉'][i] if i < 3 else '▪️'
        rend_txt = f" ({d['rend']:.1f} m³/hr)" if d['hrs'] > 0 else ""
        L.append(f"{medal} {d['team']}: {fmt(d['vol'])} m³{rend_txt}")
    L.append("")

    # ── TIEMPOS PERDIDOS ──
    L.append(f"*⏱️ TIEMPOS PERDIDOS: {tm_dia_total/60:.1f}h*")
    if tm_dia_total > 0:
        pct_mant = tm_dia_mant / tm_dia_total * 100
        pct_oper = tm_dia_oper / tm_dia_total * 100
        pct_proc = tm_dia_proc / tm_dia_total * 100
        L.append(f"🔧 Mantención: {tm_dia_mant/60:.1f}h ({pct_mant:.0f}%)")
        L.append(f"⚙️ Operacional: {tm_dia_oper/60:.1f}h ({pct_oper:.0f}%)")
        L.append(f"🔄 Proceso: {tm_dia_proc/60:.1f}h ({pct_proc:.0f}%)")
        L.append("")

        L.append(f"*Top causas:*")
        for i, (causa, mins) in enumerate(top_causas.items()):
            if i >= 4: break
            pct = mins / tm_dia_total * 100
            nombre = causa.replace('Reparacion/','Rep. ').replace('Falla ','F. ')
            if len(nombre) > 40:
                nombre = nombre[:38] + '…'
            L.append(f"  {i+1}. {nombre}: {mins/60:.1f}h ({pct:.0f}%)")
    L.append("")

    # ── DISPONIBILIDAD ──
    L.append(f"*🔋 DISPONIBILIDAD*")
    for d in dia_data:
        if d['turno_min'] == 0:
            continue
        if d['disp'] < 70:
            emoji = '🔴'
        elif d['disp'] < 80:
            emoji = '🟡'
        else:
            emoji = '🟢'
        tm_txt = f" (TM: {d['tm_mant']/60:.1f}h)" if d['tm_mant'] > 0 else ""
        L.append(f"{emoji} {d['team']}: {d['disp']:.0f}%{tm_txt}")
    L.append("")

    # ── RENDIMIENTO ──
    L.append(f"*⚡ RENDIMIENTO (m³/hr)*")
    rend_sorted = sorted([d for d in dia_data if d['hrs'] > 0], key=lambda x: x['rend'], reverse=True)
    for d in rend_sorted:
        if d['rend'] < 15:
            emoji = '🔴'
        elif d['rend'] < 20:
            emoji = '🟡'
        else:
            emoji = '🟢'
        L.append(f"{emoji} {d['team']}: {d['rend']:.1f} m³/hr ({d['hrs']:.1f}h efectivas)")
    L.append("")

    # ── PROYECCIÓN MENSUAL ──
    L.append(f"━━━━━━━━━━━━━━━━━━━━━")
    L.append(f"*📊 PROYECCIÓN {MESES[MES].upper()} {ANIO}*")
    L.append(f"Día {DD} de {DM} | Quedan {DR} días")
    L.append("")
    L.append(f"Acumulado: *{fmt(acum_total)} m³* de {fmt(meta_total)}")
    L.append(f"Proyección: *{fmt(proy_total)} m³* ({pct_total:.0f}%)")
    brecha_emoji = '✅' if brecha_total >= 0 else '⚠️'
    L.append(f"Brecha: {brecha_emoji} *{brecha_total:+,.0f} m³*".replace(',','.'))
    L.append("")

    # Por equipo (de peor a mejor)
    L.append(f"*Por faena:*")
    for p in proy_equipo:
        if p['pct'] >= 80:
            emoji = '🟢'
        elif p['pct'] >= 60:
            emoji = '🟡'
        else:
            emoji = '🔴'
        L.append(f"{emoji} {p['team']}: {fmt(p['proy'])} m³ ({p['pct']:.0f}%) meta {fmt(p['meta'])}")
    L.append("")

    # ── TENDENCIA ──
    if len(tendencia) >= 2:
        L.append(f"*📈 TENDENCIA (últimos 3 días)*")
        for d, v in tendencia:
            bar = '█' * max(1, int(v / prom_dia_total * 10)) if prom_dia_total > 0 else '█'
            L.append(f"  Día {d}: {fmt(v)} m³ {bar}")

        # Dirección de la tendencia
        if tendencia[-1][1] < tendencia[-2][1]:
            L.append(f"  ↘️ Tendencia a la baja")
        elif tendencia[-1][1] > tendencia[-2][1]:
            L.append(f"  ↗️ Tendencia al alza")
        else:
            L.append(f"  ➡️ Estable")
        L.append("")

    # ── ALERTAS Y FOCO OPERATIVO ──
    alertas = []
    # Equipos con baja disponibilidad
    for d in equipos_baja_disp:
        alertas.append(f"🔧 {d['team']}: disponibilidad {d['disp']:.0f}% — revisar mantención")
    # Equipos con bajo rendimiento
    for d in equipos_bajo_rend:
        if not any(d['team'] in a for a in alertas):
            alertas.append(f"⚡ {d['team']}: rendimiento {d['rend']:.1f} m³/hr — bajo estándar")
    # Proyección crítica
    for p in proy_equipo:
        if p['pct'] < 60:
            alertas.append(f"📉 {p['team']}: proyección {p['pct']:.0f}% — requiere plan de recuperación")
    # TM excesivo
    if tm_dia_total / 60 > 30:
        alertas.append(f"⏱️ TM total {tm_dia_total/60:.0f}h — sobre umbral aceptable")
    # Causa dominante
    if len(top_causas) > 0 and top_causas.iloc[0] > 120:
        causa_top = top_causas.index[0].replace('Reparacion/','Rep. ').replace('Falla ','F. ')
        alertas.append(f"🎯 Causa principal TM: {causa_top} ({top_causas.iloc[0]/60:.1f}h)")

    if alertas:
        L.append(f"*🚨 ALERTAS Y FOCO OPERATIVO*")
        for a in alertas:
            L.append(a)
        L.append("")

    # ── FOOTER ──
    L.append(f"━━━━━━━━━━━━━━━━━━━━━")
    L.append(f"🔗 produccion.millalemu.com")
    hora = datetime.now().strftime('%H:%M')
    L.append(f"_Generado: {datetime.now().strftime('%d/%m/%Y')} {hora}_")

    resumen = "\n".join(L)

    # Guardar
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write(resumen)

    print(f"✅ Resumen diario generado: {OUTPUT}")
    print(f"\n{'='*50}")
    print(resumen)
    print(f"{'='*50}")

    return str(OUTPUT)

if __name__ == '__main__':
    generate()
