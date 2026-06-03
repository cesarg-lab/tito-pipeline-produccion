"""
GENERAR_HTML.py — v1.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Lee Base2NOC.csv y TiemposPerdidos.csv
y genera Dashboard_Cosecha.html — un tablero interactivo
con los mismos KPIs del Excel, listo para abrir en cualquier
navegador o compartir por WhatsApp/correo.

Doble clic sobre este archivo → HTML generado.
Requisitos: Python 3 con pandas
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
import pandas as pd
import numpy as np
import json
import calendar
import os, sys
from datetime import date, datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)
from normalizar_produccion import normalizar  # noqa: E402

# ── Argumento opcional: --snapshot YYYY-MM ──────────────────────
# Si se invoca con --snapshot 2026-04, lee desde _snapshots/2026-04/
# y genera Dashboard_Cosecha_2026-04.html (no toca el del mes en curso).
SNAPSHOT_MODE = None
if '--snapshot' in sys.argv:
    idx = sys.argv.index('--snapshot')
    if idx + 1 < len(sys.argv):
        SNAPSHOT_MODE = sys.argv[idx + 1]  # ej: '2026-04'

if SNAPSHOT_MODE:
    SNAP_DIR = os.path.join(BASE_DIR, '_snapshots', SNAPSHOT_MODE)
    if not os.path.isdir(SNAP_DIR):
        print(f"❌ No se encontró snapshot: {SNAP_DIR}"); sys.exit(1)
    CSV_PROD = os.path.join(SNAP_DIR, "data_diario.csv")
    CSV_TM   = os.path.join(SNAP_DIR, "data_tm.csv")
    META_JSON = os.path.join(SNAP_DIR, "meta.json")
    OUTPUT   = os.path.join(BASE_DIR, f"Dashboard_Cosecha_{SNAPSHOT_MODE}.html")
    print(f"📦 Modo SNAPSHOT: leyendo {SNAP_DIR}")
else:
    CSV_PROD = os.path.join(BASE_DIR, "Base2NOC.csv")
    CSV_TM   = os.path.join(BASE_DIR, "TiemposPerdidos.csv")
    OUTPUT   = os.path.join(BASE_DIR, "Dashboard_Cosecha.html")

for f in [CSV_PROD, CSV_TM]:
    if not os.path.exists(f):
        print(f"❌ No se encontró: {f}"); sys.exit(1)

print("📂 Generando Dashboard HTML...")

# ── Configuración ────────────────────────────────────────────
TEAM_MAP = {
    'S123':'Millalemu 1.1','S58':'Millalemu 1.2','S223':'Millalemu 1.3',
    'S246':'Millalemu 1.4','MG5':'Millalemu 5','TEA02':'Millalemu 7',
    'TEA08':'Millalemu 9','T125':'Millalemu 11','TEA30':'Millalemu 1.3',
}
TEAMS = ['Millalemu 1.1','Millalemu 1.2','Millalemu 1.3','Millalemu 1.4',
         'Millalemu 5','Millalemu 7','Millalemu 9','Millalemu 11']
CLASIF = {
    # Mantención — falla / reparación / mantención de equipos
    1:'Mantención',2:'Mantención',3:'Mantención',4:'Mantención',5:'Mantención',
    6:'Mantención',7:'Mantención',8:'Mantención',10:'Mantención',12:'Mantención',
    58:'Mantención',69:'Mantención',
    # Operacional — logística / personal / abastecimiento / externo
    13:'Operacional',14:'Operacional',15:'Operacional',20:'Operacional',21:'Operacional',
    22:'Operacional',31:'Operacional',32:'Operacional',33:'Operacional',38:'Operacional',
    41:'Operacional',
    # Proceso — flujo interno de la faena
    16:'Proceso',17:'Proceso',18:'Proceso',25:'Proceso',26:'Proceso',61:'Proceso',
    65:'Proceso',66:'Proceso',68:'Proceso',
    # Programado — NO es pérdida (se muestra aparte, no suma al tiempo perdido)
    42:'Programado',43:'Programado',
}
# Leer metas desde Excel si existe, sino usar defaults
EXCEL = os.path.join(BASE_DIR, "Dashboard_CosechaForestal.xlsx")
METAS_DEFAULT = {
    'Millalemu 1.1': 7000.0, 'Millalemu 1.2': 7000.0, 'Millalemu 1.3': 7000.0,
    'Millalemu 1.4': 7000.0, 'Millalemu 5': 4500.0, 'Millalemu 7': 7000.0,
    'Millalemu 9': 7000.0, 'Millalemu 11': 6000.0
}
METAS = dict(METAS_DEFAULT)
if os.path.exists(EXCEL):
    try:
        from openpyxl import load_workbook
        wb_cfg = load_workbook(EXCEL, data_only=True)
        if "CONFIGURACIÓN" in wb_cfg.sheetnames:
            cfg_ws = wb_cfg["CONFIGURACIÓN"]
            for i, t in enumerate(TEAMS):
                v = cfg_ws.cell(24 + i, 5).value
                if v: METAS[t] = float(v)
            print(f"📋 Metas leídas desde Excel: {METAS}")
        wb_cfg.close()
    except Exception as e:
        print(f"⚠️  No se pudieron leer metas del Excel: {e}. Usando defaults.")

# ── Cargar histórico de cierres mensuales (alimenta pestaña Comparativo) ────
HIST_CSV = os.path.join(BASE_DIR, "historico_cierres_mensuales.csv")
historico_list = []
if os.path.exists(HIST_CSV):
    try:
        hist_df = pd.read_csv(HIST_CSV, sep=';', encoding='utf-8-sig')
        # Normalizar y serializar
        for _, h in hist_df.iterrows():
            try:
                pct_v = h['cumplimiento_pct']
                pct_num = float(pct_v) if pd.notna(pct_v) and str(pct_v).strip() not in ('','nan') else None
            except Exception:
                pct_num = None
            historico_list.append({
                'mes': int(h['mes']),
                'anio': int(h['anio']),
                'mesNombre': str(h['mes_nombre']),
                'equipo': str(h['equipo']),
                'meta': float(h['meta_mensual']) if pd.notna(h['meta_mensual']) else 0,
                'vol': float(h['vol_total']) if pd.notna(h['vol_total']) else 0,
                'cumpl': pct_num,
                'promDia': float(h['prom_diario']) if pd.notna(h['prom_diario']) else 0,
                'diasTrab': int(h['dias_trabajados']) if pd.notna(h['dias_trabajados']) else 0,
            })
        print(f"📈 Histórico cargado: {len(historico_list)} registros ({hist_df['mes_nombre'].nunique()} meses)")
    except Exception as e:
        print(f"⚠️  No se pudo cargar histórico: {e}")

# ── Histórico de tiempos perdidos por mes ──
TM_HIST_CSV = os.path.join(BASE_DIR, "historico_tm_mensual.csv")
tm_mensual_list = []
if os.path.exists(TM_HIST_CSV):
    try:
        _tmh = pd.read_csv(TM_HIST_CSV, sep=';', encoding='utf-8-sig')
        for _, r in _tmh.iterrows():
            tm_mensual_list.append({
                'mes': int(r['mes']), 'anio': int(r['anio']), 'mesNombre': str(r['mes_nombre']),
                'mant': float(r['mantencion_h']), 'oper': float(r['operacional_h']),
                'proc': float(r['proceso_h']), 'prog': float(r['programado_h']),
                'perdido': float(r['total_perdido_h']), 'topCausas': str(r.get('top_causas', '')),
            })
        print(f"📉 TM mensual cargado: {len(tm_mensual_list)} meses")
    except Exception as e:
        print(f"⚠️  No se pudo cargar TM mensual: {e}")

# ── Leer y procesar datos ────────────────────────────────────
if SNAPSHOT_MODE:
    # Modo snapshot: data_diario.csv y data_tm.csv ya están cocinados
    import json as _json
    with open(META_JSON, 'r', encoding='utf-8') as f:
        SNAP_META = _json.load(f)
    SNAP_MES = SNAP_META['mes']; SNAP_ANIO = SNAP_META['anio']
    SNAP_METAS = SNAP_META.get('metas', {})

    dd_csv = pd.read_csv(CSV_PROD, sep=';', encoding='utf-8-sig')
    # Construir prod con las columnas que espera el resto del código
    sigla_por_team = {}
    for sigla, team in TEAM_MAP.items():
        if team not in sigla_por_team:
            sigla_por_team[team] = sigla

    prod = pd.DataFrame()
    prod['Dia'] = dd_csv['Dia'].astype(int)
    prod['Equipo'] = dd_csv['Team'].map(sigla_por_team)
    prod['Team'] = dd_csv['Team']
    prod['Vol'] = pd.to_numeric(dd_csv['Vol_m3'], errors='coerce').fillna(0)
    prod['HrsEf'] = pd.to_numeric(dd_csv['Hrs_Ef'], errors='coerce').fillna(0)
    prod['Volumen SSC PU'] = prod['Vol']
    prod['Volumen SSC AS'] = 0.0
    prod['Tiempo Efectivo'] = (prod['HrsEf'] * 3600).astype(float)
    prod['Turno_seg'] = (prod['HrsEf'] * 3600 + 3600).astype(float)
    prod['Arb'] = pd.to_numeric(dd_csv.get('Arboles', 0), errors='coerce').fillna(0)
    prod['Ciclos'] = pd.to_numeric(dd_csv.get('Ciclos', 0), errors='coerce').fillna(0)
    prod['Desc Especie'] = ''
    prod['Especie'] = ''
    prod['Origen'] = ''
    prod['Código Predio'] = ''
    prod['Predio'] = ''
    prod['Código Equipo'] = prod['Equipo']
    prod['Hora Inicio'] = 0
    prod['Hora Fin'] = 0
    prod['Fecha_dt'] = pd.to_datetime([f"{SNAP_ANIO}-{SNAP_MES:02d}-{int(d):02d}" for d in prod['Dia']])
    prod['Fecha'] = prod['Fecha_dt'].dt.strftime('%d-%m-%Y')

    tm_csv = pd.read_csv(CSV_TM, sep=';', encoding='utf-8-sig')
    tm = pd.DataFrame()
    tm['Dia'] = tm_csv['Dia'].astype(int)
    tm['Team'] = tm_csv['Team']
    tm['Código Equipo'] = tm['Team'].map(sigla_por_team)
    tm['Código Tiempo Perdido'] = pd.to_numeric(tm_csv['Codigo'], errors='coerce').fillna(0).astype(int)
    tm['Descripción'] = tm_csv['Descripcion']
    tm['Clasif'] = tm_csv['Clasificacion']
    tm['Tiempo (Min)'] = pd.to_numeric(tm_csv['Minutos'], errors='coerce').fillna(0)
    tm['Observación'] = tm_csv.get('Observacion', '')
    tm['Fecha_dt'] = pd.to_datetime([f"{SNAP_ANIO}-{SNAP_MES:02d}-{int(d):02d}" for d in tm['Dia']])
    tm['Fecha'] = tm['Fecha_dt'].dt.strftime('%d-%m-%Y')

    METAS = {t: float(v) for t, v in SNAP_METAS.items()}
    print(f"📦 Snapshot {SNAPSHOT_MODE} cargado: {len(prod)} filas prod, {len(tm)} filas TM")
else:
    prod = pd.read_csv(CSV_PROD, sep=';', encoding='utf-8-sig')
    prod = normalizar(prod)
    tm = pd.read_csv(CSV_TM, sep=';', encoding='utf-8-sig')

# ── Manual.csv DESACTIVADO (Cesar 2026-05-02) ──────────────────────────────
# Selenium es fuente única desde 2026-04-18. El Manual.csv ya no se usa ni se
# auto-importa desde ~/Downloads. Toda la lógica posterior queda inerte porque
# MANUAL_CSV apunta a una ruta inexistente.
USAR_MANUAL = False
MANUAL_CSV = "/__manual_desactivado__"  # path imposible → os.path.exists() siempre False

vol_oficial_diario = None  # {(team, dia): vol}
vol_oficial_total_eq = None  # {team: total mensual oficial}
manual_disponible = False
if os.path.exists(MANUAL_CSV):
    # Validar que Manual no sea de un mes anterior al del CSV detallado
    import datetime as _dt
    mtime_manual = _dt.datetime.fromtimestamp(os.path.getmtime(MANUAL_CSV))
    # Determinar mes del detallado
    try:
        # Base 2 NOC trae 'FECHA'; PG legacy traía 'Fecha NOC' — leer ambas posibles
        try:
            _prod_chk = pd.read_csv(CSV_PROD, sep=';', encoding='utf-8-sig', usecols=['FECHA'])
            _prod_chk = _prod_chk.rename(columns={'FECHA': 'Fecha NOC'})
        except (ValueError, KeyError):
            _prod_chk = pd.read_csv(CSV_PROD, sep=';', encoding='utf-8-sig', usecols=['Fecha NOC'])
        _prod_chk['_dt'] = pd.to_datetime(_prod_chk['Fecha NOC'], dayfirst=True, errors='coerce')
        mes_actual = int(_prod_chk['_dt'].dt.month.mode()[0])
        anio_actual = int(_prod_chk['_dt'].dt.year.mode()[0])
        # Si el Manual fue modificado antes del 1 del mes actual → es del mes anterior, descartar
        inicio_mes = _dt.datetime(anio_actual, mes_actual, 1)
        if mtime_manual < inicio_mes:
            print(f"⚠️  Manual.csv es del mes anterior ({mtime_manual.strftime('%d-%b')}). Descargá uno nuevo desde GeoNOC para el mes actual.")
            print(f"   Por ahora se generará el dashboard SIN validar contra oficial.")
            os.rename(MANUAL_CSV, MANUAL_CSV + '.mes_anterior.bak')  # lo renombra para evitar que se use
            vol_oficial_diario = None
        else:
            raise StopIteration  # señalar que el Manual es válido y seguir
    except StopIteration:
        pass
    except Exception as e:
        print(f"⚠️  Validación mes Manual falló: {e}. Se intentará usar de todos modos.")

if os.path.exists(MANUAL_CSV):
    try:
        mdf = pd.read_csv(MANUAL_CSV, sep=';', encoding='utf-8-sig', decimal=',', thousands='.')
        vol_oficial_diario = {}
        vol_oficial_total_eq = {}
        for _, r in mdf.iterrows():
            eq_full = str(r.get('Equipo','')).strip()
            if '-' not in eq_full:
                continue
            sigla = eq_full.split('-')[0]
            team = TEAM_MAP.get(sigla)
            if not team:
                continue
            if pd.notna(r.get('Total')):
                vol_oficial_total_eq[team] = float(r['Total'])
            for d in range(1, 32):
                v = r.get(str(d))
                if pd.notna(v) and v != 0:
                    vol_oficial_diario[(team, d)] = float(v)
        manual_disponible = True
        print(f"📋 Manual.csv cargado: {len(vol_oficial_diario)} registros día/equipo (fuente de verdad para m³)")
    except Exception as e:
        print(f"⚠️  No se pudo leer Manual.csv: {e}. Usando suma PU+AS del detallado.")
        vol_oficial_diario = None
else:
    print("ℹ️  'ProduccionEquipo Manual.csv' no encontrado — usando CSV detallado tal cual (vía Selenium ya viene depurado).")

if not SNAPSHOT_MODE:
    for c in ['Volumen SSC PU','Volumen SSC AS']:
        prod[c] = pd.to_numeric(prod[c].astype(str).str.replace(',','.'), errors='coerce')
    prod['Vol'] = prod['Volumen SSC PU'].fillna(0) + prod['Volumen SSC AS'].fillna(0)
    # Tiempo Efectivo: detectar unidad automáticamente
    _te_raw = pd.to_numeric(prod['Tiempo Efectivo'].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    # Unidad robusta: usar la MEDIANA de valores >0 (no el máximo, que se rompe con
    # un solo outlier). Día típico ~480 min ó ~28800 seg → umbral 1000.
    _te_pos = _te_raw[_te_raw > 0]
    _te_divisor = 3600 if (len(_te_pos) > 0 and _te_pos.median() > 1000) else 60
    prod['HrsEf'] = _te_raw / _te_divisor
    prod['Arb'] = pd.to_numeric(prod['Árboles Madereados'], errors='coerce').fillna(0)
    prod['Ciclos'] = pd.to_numeric(prod['Número Ciclos'], errors='coerce').fillna(0).astype(int)
    prod['Team'] = prod['Equipo'].map(TEAM_MAP)
    prod['Fecha_dt'] = pd.to_datetime(prod['Fecha NOC'], dayfirst=True, errors='coerce')
    prod['Dia'] = prod['Fecha_dt'].dt.day

    # Turno en segundos (Hora Fin - Hora Inicio)
    for c in ['Hora Inicio', 'Hora Fin']:
        prod[c] = pd.to_numeric(prod[c], errors='coerce').fillna(0)
    prod['Turno_seg'] = (prod['Hora Fin'] - prod['Hora Inicio']).clip(lower=0)

    tm['Clasif'] = tm['Código Tiempo Perdido'].map(CLASIF).fillna('Operacional')
    tm['Team'] = tm['Código Equipo'].map(TEAM_MAP)
    tm['Fecha_dt'] = pd.to_datetime(tm['Fecha'], dayfirst=True, errors='coerce')
    tm['Dia'] = tm['Fecha_dt'].dt.day

MES = int(prod['Fecha_dt'].dt.month.mode()[0])

# Tiempo PROGRAMADO (colación / descanso): NO cuenta como tiempo perdido.
# Se guarda el total para mostrarlo aparte y se excluye de TODAS las agregaciones
# de TM (categorías, top causas, tendencia, disponibilidad), dejándolas coherentes.
tm_prog_min = int(tm.loc[tm['Clasif'] == 'Programado', 'Tiempo (Min)'].sum()) if 'Clasif' in tm.columns else 0
tm = tm[tm['Clasif'] != 'Programado'].copy()

# Guardar copia del mes anterior para comparativa (si hay datos)
MES_PREV = MES - 1 if MES > 1 else 12
ANIO_PREV = int(prod['Fecha_dt'].dt.year.mode()[0]) if MES > 1 else int(prod['Fecha_dt'].dt.year.mode()[0]) - 1
prod_prev = prod[prod['Fecha_dt'].dt.month == MES_PREV].copy()
tm_prev = tm[tm['Fecha_dt'].dt.month == MES_PREV].copy()

# Filtrar registros que no pertenecen al mes actual (ej: 31 marzo en datos de abril)
prod = prod[prod['Fecha_dt'].dt.month == MES]
tm = tm[tm['Fecha_dt'].dt.month == MES]
ANIO = int(prod['Fecha_dt'].dt.year.mode()[0])
DM = calendar.monthrange(ANIO, MES)[1]
# Feriados irrenunciables (fechas fijas): no se trabaja, no cuentan como día trabajado.
FERIADOS_IRR = {'01-01', '05-01', '09-18', '09-19', '12-25'}
def _habil(_mes, _dia):
    return f"{_mes:02d}-{_dia:02d}" not in FERIADOS_IRR
if vol_oficial_diario is not None:
    ULTIMO_DIA = max(d for (_, d) in vol_oficial_diario.keys())
else:
    ULTIMO_DIA = int(prod['Dia'].max())
# Regla: los equipos deben producir TODOS los días; un día de falla (producción 0)
# cuenta igual. Por eso "días trabajados" = días corridos hábiles transcurridos
# (NO los días con registro), menos feriados irrenunciables.
DT = sum(1 for d in range(1, DM + 1) if _habil(MES, d))            # días hábiles del mes
DD = sum(1 for d in range(1, ULTIMO_DIA + 1) if _habil(MES, d))    # días hábiles transcurridos
DR = max(DT - DD, 0)

def dominant(s):
    m = s.dropna().mode()
    return m.iloc[0] if len(m) > 0 else ''

# FIX 2026-05-23: HrsEf y Turno_seg son del folio completo, no del
# producto. Cada folio aparece N veces en el CSV (una por producto) con
# el MISMO HrsEf. Sumar por fila los multiplica por N → bug que infla
# horas 3-6x. Solucion: dedup por folio (Numero Noc) tomando first,
# despues sumar al nivel (Dia, Team). Vol/Arb/Ciclos SI se suman por
# fila (cada producto aporta su parte). Mismo patron aplicado en
# GENERAR_RESUMEN.py commit d4b0aac.
if 'Número Noc' in prod.columns:
    prod_folio = prod.groupby(['Dia','Team','Número Noc']).agg(
        Vol=('Vol','sum'),
        HrsEf=('HrsEf','first'),
        Turno_seg=('Turno_seg','first'),
        Arb=('Arb','sum'),
        Ciclos=('Ciclos','sum'),
        Especie=('Desc Especie', dominant),
        Predio=('Origen', dominant),
    ).reset_index()
    daily = prod_folio.groupby(['Dia','Team']).agg(
        Vol=('Vol','sum'), HrsEf=('HrsEf','sum'),
        Turno_seg=('Turno_seg','sum'),
        Arb=('Arb','sum'), Ciclos=('Ciclos','sum'),
        Especie=('Especie', dominant),
        Predio=('Predio', dominant)
    ).reset_index()
else:
    # Modo snapshot historico: prod ya viene pre-agregado, no hay folio
    daily = prod.groupby(['Dia','Team']).agg(
        Vol=('Vol','sum'), HrsEf=('HrsEf','sum'),
        Turno_seg=('Turno_seg','sum'),
        Arb=('Arb','sum'), Ciclos=('Ciclos','sum'),
        Especie=('Desc Especie', dominant),
        Predio=('Origen', dominant)
    ).reset_index()

# Si hay archivo Manual oficial, RECONSTRUIR los volúmenes diarios con los valores depurados
# El Manual es la fuente de verdad: si un (Team, Día) no está en Manual, significa que ese día
# ese equipo NO produjo oficialmente (aunque el detallado tenga registros fantasma para ese día)
if vol_oficial_diario:
    _total_antes = daily['Vol'].sum()
    # Si (Team, Dia) está en Manual, usar Manual (depurado).
    # Si NO está en Manual, conservar el valor de Base2NOC (Selenium ya viene depurado).
    daily['Vol'] = daily.apply(
        lambda r: vol_oficial_diario.get((r['Team'], int(r['Dia'])), r['Vol']), axis=1
    )
    # Eliminar filas con Vol=0 Y sin datos de detalle (fueron fantasmas depurados)
    mask_quitar = (daily['Vol'] == 0) & (daily.get('HrsEf', 0) == 0) & (daily.get('Arb', 0) == 0)
    daily = daily[~mask_quitar].reset_index(drop=True)
    # Agregar filas que existen en Manual pero no en el detallado
    idx_existentes = set((r['Team'], int(r['Dia'])) for _, r in daily.iterrows())
    filas_nuevas = []
    for (t, d), v in vol_oficial_diario.items():
        if (t, d) not in idx_existentes:
            filas_nuevas.append({
                'Dia': d, 'Team': t, 'Vol': v, 'HrsEf': 0, 'Turno_seg': 0,
                'Arb': 0, 'Ciclos': 0, 'Especie': '', 'Predio': ''
            })
    if filas_nuevas:
        daily = pd.concat([daily, pd.DataFrame(filas_nuevas)], ignore_index=True)
    _total_despues = daily['Vol'].sum()
    print(f"   Volúmenes ajustados: {_total_antes:.1f} → {_total_despues:.1f} m³ (fantasmas depurados)")

piv = tm.groupby(['Dia','Team','Clasif'])['Tiempo (Min)'].sum().unstack(fill_value=0).reset_index()
piv.columns.name = None
for c in ['Mantención','Operacional','Proceso']:
    if c not in piv.columns: piv[c] = 0
piv = piv.rename(columns={'Mantención':'TM_Mant','Operacional':'TM_Oper','Proceso':'TM_PP'})
daily = daily.merge(piv[['Dia','Team','TM_Mant','TM_Oper','TM_PP']], on=['Dia','Team'], how='left')
for c in ['TM_Mant','TM_Oper','TM_PP']:
    daily[c] = daily[c].fillna(0)
daily['TM_Total'] = daily['TM_Mant'] + daily['TM_Oper'] + daily['TM_PP']

# TM por equipo desde los registros CRUDOS (no la tabla cruzada con producción, que
# pierde fallas largas sin fila de producción que calce → disponibilidad inflada).
# Mismo fix que las categorías globales.
_tm_team = tm.groupby(['Team', 'Clasif'])['Tiempo (Min)'].sum().unstack(fill_value=0)
def _tmcat(_team, _cat):
    return int(_tm_team.loc[_team, _cat]) if (_team in _tm_team.index and _cat in _tm_team.columns) else 0

# ── Generar JSON ─────────────────────────────────────────────
team_kpis = []
for t in TEAMS:
    td = daily[daily['Team'] == t]
    acum = round(td['Vol'].sum(), 1)
    meta = METAS[t]
    dias_team = DD  # regla: todos deben producir los días hábiles (faltar no premia)
    prom = round(acum / dias_team, 1) if dias_team > 0 else 0
    proy = round(acum + prom * DR, 1)
    hrs = round(td['HrsEf'].sum(), 1)
    turno_min = round(td['Turno_seg'].sum() / 60, 1) if len(td) > 0 else 0
    team_kpis.append({
        't': t, 'a': acum, 'm': meta,
        'c': round((acum/meta)*100,1) if meta else 0,
        'p': prom, 'pr': proy, 'b': round(proy-meta,1),
        'ci': round((proy/meta)*100,1) if meta else 0,
        'r': round(acum/hrs,2) if hrs else 0,
        'h': hrs, 'tm': _tmcat(t, 'Mantención'),
        'tt': _tmcat(t, 'Mantención') + _tmcat(t, 'Operacional') + _tmcat(t, 'Proceso'),
        'turno': turno_min,
        'e': td['Especie'].mode().iloc[0] if len(td)>0 and len(td['Especie'].mode())>0 else '',
        'pr2': td['Predio'].mode().iloc[0] if len(td)>0 and len(td['Predio'].mode())>0 else '',
        'd': dias_team
    })

grid = {}
for _, r in daily.iterrows():
    d = int(r['Dia']); t = r['Team']
    if d not in grid: grid[d] = {}
    grid[d][t] = round(float(r['Vol']),1)

# Grid de TM Mantención y HrsEf por dia/equipo
gridTM = {}
gridHrs = {}
for _, r in daily.iterrows():
    d = int(r['Dia']); t = r['Team']
    if d not in gridTM: gridTM[d] = {}
    if d not in gridHrs: gridHrs[d] = {}
    gridTM[d][t] = round(float(r['TM_Total']),0)  # TM total (Mant+Oper+Proceso) para consistencia con heatmap
    gridHrs[d][t] = round(float(r['HrsEf']),1)

trend = daily.groupby('Dia')['Vol'].sum().reset_index().sort_values('Dia')
trend_list = [{'d': int(r['Dia']), 'v': round(float(r['Vol']),1)} for _, r in trend.iterrows()]

tm_causes = tm.groupby('Descripción')['Tiempo (Min)'].sum().sort_values(ascending=False).head(10)
tm_list = [{'n': str(k), 'm': int(v)} for k,v in tm_causes.items()]

esp = daily.groupby('Especie')['Vol'].sum()
esp_list = [{'n': str(k), 'v': round(float(v),1)} for k,v in esp.items()]

# TM por categoría (donut)
# Sumar de los registros crudos de TM (misma fuente que el top causas), no de la
# tabla cruzada con producción — así no se pierden TM de equipos sin producción
# ese día ni de códigos de equipo sin mapear. Mantiene coherencia con tm_causes.
tm_by_cat = daily.groupby('Team')[['TM_Mant','TM_Oper','TM_PP']].sum()
_tm_cat = tm.groupby('Clasif')['Tiempo (Min)'].sum()
tm_cat_total = {'Mantención': int(_tm_cat.get('Mantención', 0)),
                'Operacional': int(_tm_cat.get('Operacional', 0)),
                'Proceso': int(_tm_cat.get('Proceso', 0))}
tm_cat_list = [{'n': k, 'v': v} for k, v in tm_cat_total.items() if v > 0]

# TM tendencia diaria
tm_daily_trend = daily.groupby('Dia')[['TM_Mant','TM_Oper','TM_PP','TM_Total']].sum().reset_index().sort_values('Dia')
tm_trend_list = [{'d': int(r['Dia']), 'mant': int(r['TM_Mant']), 'oper': int(r['TM_Oper']),
                  'proc': int(r['TM_PP']), 'total': int(r['TM_Total'])} for _, r in tm_daily_trend.iterrows()]

# TM por equipo desglosado por categoría (stacked bar)
tm_team_cat = []
for t in TEAMS:
    td = daily[daily['Team'] == t]
    tm_team_cat.append({
        't': t.replace('Millalemu ', 'M'),
        'tf': t,
        'mant': _tmcat(t, 'Mantención'),
        'oper': _tmcat(t, 'Operacional'),
        'proc': _tmcat(t, 'Proceso'),
        'total': _tmcat(t, 'Mantención') + _tmcat(t, 'Operacional') + _tmcat(t, 'Proceso'),
        'hrs': round(td['HrsEf'].sum(), 1),
        'disp': round((1 - _tmcat(t, 'Mantención') / (td['Turno_seg'].sum()/60)) * 100, 1) if td['Turno_seg'].sum() > 0 else 100
    })

# ── DATOS NUEVOS ─────────────────────────────────────────

# 1. Avance acumulado diario (real vs plan)
avance_diario = []
acum_real = 0
meta_total = sum(METAS.values())
for d in sorted(daily['Dia'].unique().astype(int)):
    vol_dia = daily[daily['Dia'] == d]['Vol'].sum()
    acum_real += vol_dia
    plan_dia = meta_total / DT * d
    avance_diario.append({'d': int(d), 'real': round(acum_real, 1), 'plan': round(plan_dia, 1)})

# 2. Rendimiento por especie (m³/hr por especie y global)
rend_especie = []
for esp_name, grp in daily.groupby('Especie'):
    if pd.isna(esp_name) or str(esp_name).strip() == '': continue
    vol_e = grp['Vol'].sum()
    hrs_e = grp['HrsEf'].sum()
    rend_especie.append({
        'n': str(esp_name), 'vol': round(vol_e, 1),
        'hrs': round(hrs_e, 1), 'rend': round(vol_e/hrs_e, 2) if hrs_e > 0 else 0
    })
# Por especie y faena
rend_esp_team = []
for (esp_name, team), grp in daily.groupby(['Especie', 'Team']):
    if pd.isna(esp_name) or str(esp_name).strip() == '': continue
    vol_e = grp['Vol'].sum()
    hrs_e = grp['HrsEf'].sum()
    if hrs_e > 0:
        rend_esp_team.append({
            'e': str(esp_name), 't': team,
            'rend': round(vol_e/hrs_e, 2), 'vol': round(vol_e, 1)
        })

# 3. Causas TM por faena (todas, para permitir análisis Pareto 80/20)
tm_team_causes = {}
for t in TEAMS:
    tm_t = tm[tm['Team'] == t]
    all_causas = tm_t.groupby('Descripción')['Tiempo (Min)'].sum().sort_values(ascending=False).head(20)
    tm_team_causes[t] = [{'n': str(k), 'm': int(v)} for k, v in all_causas.items()]

# 3b. Pareto global de FAENA + cruce causa×equipo + recurrencia + recomendaciones
tm_pareto_global = []
tm_heatmap_causa_equipo = []
tm_recomendaciones = []
try:
    tm_total_min = float(tm['Tiempo (Min)'].sum())
    tm_total_h = tm_total_min / 60.0

    # Agregación global por causa
    g = tm.groupby('Descripción').agg(
        min_total=('Tiempo (Min)', 'sum'),
        eventos=('Tiempo (Min)', 'count'),
        dias=('Fecha_dt', lambda s: s.dt.day.nunique()),
    ).reset_index().sort_values('min_total', ascending=False)
    g['pct'] = g['min_total'] / tm_total_min * 100
    g['pct_acum'] = g['pct'].cumsum()
    g['horas'] = g['min_total'] / 60.0

    for _, r in g.iterrows():
        tm_pareto_global.append({
            'n': str(r['Descripción']),
            'h': round(float(r['horas']), 1),
            'ev': int(r['eventos']),
            'd': int(r['dias']),
            'pct': round(float(r['pct']), 1),
            'pctAcum': round(float(r['pct_acum']), 1),
            'hPorEv': round(float(r['horas']) / max(int(r['eventos']), 1), 2),
        })

    # Heatmap: top 10 causas (las que cubren ~80%) × equipos
    n_vital = max(int((g['pct_acum'] <= 80).sum()) + 1, 5)
    causas_vitales = g.head(n_vital)['Descripción'].tolist()
    cruce = tm[tm['Descripción'].isin(causas_vitales)].groupby(['Descripción', 'Team'])['Tiempo (Min)'].sum().unstack(fill_value=0) / 60.0
    for causa in causas_vitales:
        if causa in cruce.index:
            row = cruce.loc[causa]
            equipos_causa = {str(team): round(float(row.get(team, 0)), 1) for team in TEAMS}
            total_causa = sum(equipos_causa.values())
            top_team = max(equipos_causa.items(), key=lambda x: x[1]) if total_causa > 0 else (None, 0)
            concentracion = (top_team[1] / total_causa * 100) if total_causa > 0 else 0
            tm_heatmap_causa_equipo.append({
                'causa': str(causa),
                'totalH': round(total_causa, 1),
                'porEquipo': equipos_causa,
                'topEquipo': top_team[0] if top_team[1] > 0 else '',
                'topPct': round(concentracion, 1),
            })

    # Recomendaciones automáticas para mayo
    # Lógica: top causas crónicas (≥10 días) ordenadas por horas; si una está concentrada en 1 equipo, recomendación específica
    proy_m3hr_default = 50  # heurística conservadora: cada hora recuperada ≈ 50 m³ de faena
    rec_id = 1
    for h in tm_heatmap_causa_equipo[:5]:
        causa_data = next((c for c in tm_pareto_global if c['n'] == h['causa']), None)
        if not causa_data:
            continue
        # ¿Crónica o evento puntual?
        es_cronica = causa_data['d'] >= 10
        if not es_cronica:
            continue
        # Estimar ahorro al reducir 30% del tiempo perdido en mayo
        ahorro_h_30 = round(causa_data['h'] * 0.30, 1)
        ahorro_m3_30 = int(ahorro_h_30 * proy_m3hr_default)
        # Foco específico
        if h['topPct'] >= 50 and h['topEquipo']:
            foco = f"foco en {h['topEquipo']} ({h['porEquipo'][h['topEquipo']]} h, {h['topPct']:.0f}% de la causa)"
        else:
            foco = "transversal a la faena"
        tm_recomendaciones.append({
            'id': rec_id,
            'causa': h['causa'],
            'horasMes': causa_data['h'],
            'pctMes': causa_data['pct'],
            'foco': foco,
            'recurrencia': causa_data['d'],
            'ahorroPotH': ahorro_h_30,
            'ahorroPotM3': ahorro_m3_30,
            'accion': (
                "Plan mantención preventiva semanal en cabezal de trozado" if 'Trozado' in h['causa']
                else "Diagnóstico mecánico equipo de volteo + revisión hidráulica" if 'volteo' in h['causa']
                else "Revisión sistema madereo y relevo de operadores" if 'madereo' in h['causa']
                else "Auditoría operacional y coordinación logística"
            ),
        })
        rec_id += 1
except Exception as _e:
    print(f"⚠️  Error en análisis 80/20: {_e}")
    tm_pareto_global = []
    tm_heatmap_causa_equipo = []
    tm_recomendaciones = []

# 4. Tendencia semanal (últimos 7 días vs 7 anteriores)
dias_ordenados = sorted(daily['Dia'].unique().astype(int))
tendencia_team = {}
for t in TEAMS:
    td = daily[daily['Team'] == t].set_index('Dia')
    if len(dias_ordenados) >= 7:
        sem_actual = dias_ordenados[-7:]
        sem_anterior = dias_ordenados[-14:-7] if len(dias_ordenados) >= 14 else dias_ordenados[:len(dias_ordenados)-7]
    else:
        sem_actual = dias_ordenados
        sem_anterior = []
    vol_actual = td.loc[td.index.isin(sem_actual), 'Vol'].sum()
    vol_anterior = td.loc[td.index.isin(sem_anterior), 'Vol'].sum() if sem_anterior else 0
    dias_act = len([d for d in sem_actual if d in td.index])
    dias_ant = len([d for d in sem_anterior if d in td.index])
    prom_act = vol_actual / dias_act if dias_act > 0 else 0
    prom_ant = vol_anterior / dias_ant if dias_ant > 0 else 0
    if prom_ant > 0:
        cambio = round((prom_act - prom_ant) / prom_ant * 100, 1)
    else:
        cambio = 0
    tendencia_team[t] = {'pa': round(prom_act, 1), 'pp': round(prom_ant, 1), 'c': cambio}

# 5. Días sin producción por equipo
# Usar el rango completo del mes hasta el último día con datos (operan 7 días/semana)
dias_sin_prod = []
all_days = set(range(1, ULTIMO_DIA + 1))
for t in TEAMS:
    td = daily[daily['Team'] == t]
    dias_con = set(td['Dia'].unique().astype(int))
    dias_sin = sorted(all_days - dias_con)
    for d in dias_sin:
        # Buscar causa principal en TM ese día (con categoría)
        tm_dia = tm[(tm['Team'] == t) & (tm['Dia'] == d)]
        causa = ''
        categoria = ''
        mins_total = 0
        if len(tm_dia) > 0:
            mins_total = int(tm_dia['Tiempo (Min)'].sum())
            top = tm_dia.groupby('Descripción')['Tiempo (Min)'].sum().sort_values(ascending=False)
            causa = str(top.index[0]) if len(top) > 0 else 'Sin detalle'
            mins = int(top.iloc[0])
            # Categoría dominante del día (por minutos)
            cat_top = tm_dia.groupby('Clasif')['Tiempo (Min)'].sum().sort_values(ascending=False)
            categoria = str(cat_top.index[0]) if len(cat_top) > 0 else ''
        else:
            causa = 'Sin registro'
            mins = 0
            categoria = '—'
        dias_sin_prod.append({
            't': t, 'd': int(d),
            'causa': causa, 'mins': mins,
            'minsTot': mins_total, 'cat': categoria
        })

# 6. Análisis por Predio
# FIX: HrsEf es por folio (mismo turno repetido por producto). Deduplicar por
# (predio, folio) tomando 'first' antes de sumar, igual que la vista por equipo;
# si no, las horas se inflan N veces y el m³/hr sale deflactado.
predio_stats = []
if 'Número Noc' in prod.columns:
    _nom = 'Origen' if 'Origen' in prod.columns else 'Código Predio'
    _pf = prod.groupby(['Código Predio', 'Número Noc']).agg(
        vol=('Vol', 'sum'), hrs=('HrsEf', 'first'),
        arb=('Arb', 'sum'), ciclos=('Ciclos', 'sum'),
        nombre=(_nom, 'first'),
        equipos=('Team', lambda s: tuple(sorted(set(s.dropna())))),
        especies=('Desc Especie', lambda s: tuple(sorted(set(s.dropna())))),
    ).reset_index()
    pred_grp = _pf.groupby('Código Predio').agg(
        vol=('vol', 'sum'), hrs=('hrs', 'sum'),
        arb=('arb', 'sum'), ciclos=('ciclos', 'sum'),
        nombre=('nombre', 'first'),
        equipos=('equipos', lambda s: sorted(set(e for tup in s for e in tup))),
        especies=('especies', lambda s: sorted(set(e for tup in s for e in tup))),
    ).reset_index()
else:
    _nom = 'Origen' if 'Origen' in prod.columns else 'Código Predio'
    pred_grp = prod.groupby('Código Predio').agg(
        vol=('Vol','sum'), hrs=('HrsEf','sum'),
        arb=('Arb','sum'), ciclos=('Ciclos','sum'),
        nombre=(_nom, 'first'),
        equipos=('Team', lambda s: sorted(set(s.dropna()))),
        especies=('Desc Especie', lambda s: sorted(set(s.dropna())))
    ).reset_index()
for _, r in pred_grp.iterrows():
    predio_stats.append({
        'pr': str(r['Código Predio']),
        'nombre': str(r['nombre']) if pd.notna(r['nombre']) else '',
        'vol': round(r['vol'], 1),
        'hrs': round(r['hrs'], 1),
        'rend': round(r['vol']/r['hrs'], 2) if r['hrs'] > 0 else 0,
        'arb': int(r['arb']),
        'eq': [e.replace('Millalemu ','M') for e in r['equipos']],
        'esp': list(r['especies'])
    })
predio_stats.sort(key=lambda x: -x['vol'])

# 7. Ranking de eficiencia (m³/hr por equipo)
ranking_ef = []
for t in TEAMS:
    td = daily[daily['Team'] == t]
    vol = td['Vol'].sum()
    hrs = td['HrsEf'].sum()
    ciclos = td['Ciclos'].sum()
    arb = td['Arb'].sum()
    ranking_ef.append({
        't': t,
        'vol': round(vol, 1),
        'hrs': round(hrs, 1),
        'rend': round(vol/hrs, 2) if hrs > 0 else 0,
        'mArb': round(vol/arb, 3) if arb > 0 else 0,
        'mCic': round(vol/ciclos, 2) if ciclos > 0 else 0
    })
ranking_ef.sort(key=lambda x: -x['rend'])

# 8. Mix de producción por especie (% del total)
esp_mix = []
esp_total = daily['Vol'].sum()
esp_grp = prod.groupby('Desc Especie').agg(vol=('Vol','sum'), hrs=('HrsEf','sum')).reset_index()
ESP_NAMES = {'PIRA': 'Pino Radiata', 'EUGL': 'E. Globulus', 'EUNI': 'E. Nitens'}
for _, r in esp_grp.iterrows():
    esp_mix.append({
        'cod': str(r['Desc Especie']),
        'nom': ESP_NAMES.get(str(r['Desc Especie']), str(r['Desc Especie'])),
        'vol': round(r['vol'], 1),
        'pct': round(r['vol']/esp_total*100, 1) if esp_total > 0 else 0,
        'rend': round(r['vol']/r['hrs'], 2) if r['hrs'] > 0 else 0
    })
esp_mix.sort(key=lambda x: -x['vol'])

# 9. Heatmap TM por día/faena (minutos totales)
tm_heatmap = {}
for (d, t), g in tm.groupby(['Dia', 'Team']):
    mins = int(g['Tiempo (Min)'].sum())
    d_int = int(d)
    if d_int not in tm_heatmap: tm_heatmap[d_int] = {}
    tm_heatmap[d_int][t] = mins

# 10. MTBF / MTTR por equipo
mtbf_mttr = []
for t in TEAMS:
    td = daily[daily['Team'] == t]
    tm_eq = tm[tm['Team'] == t]
    # Solo fallas de Mantención
    fallas = tm_eq[tm_eq['Clasif'] == 'Mantención']
    n_fallas = len(fallas)
    hrs_prod = td['HrsEf'].sum()
    min_reparacion = int(fallas['Tiempo (Min)'].sum())
    mtbf_hrs = round(hrs_prod / n_fallas, 1) if n_fallas > 0 else None
    mttr_min = round(min_reparacion / n_fallas, 1) if n_fallas > 0 else None
    mtbf_mttr.append({
        't': t,
        'fallas': n_fallas,
        'hrsProd': round(hrs_prod, 1),
        'minReparacion': min_reparacion,
        'mtbf': mtbf_hrs,
        'mttr': mttr_min
    })

# 11. Comparativa vs mes anterior
comp_mes = None
if len(prod_prev) > 0:
    for c in ['Volumen SSC PU','Volumen SSC AS']:
        prod_prev[c] = pd.to_numeric(prod_prev[c].astype(str).str.replace(',','.'), errors='coerce')
    prod_prev['Vol'] = prod_prev['Volumen SSC PU'].fillna(0) + prod_prev['Volumen SSC AS'].fillna(0)
    _te_raw_p = pd.to_numeric(prod_prev['Tiempo Efectivo'].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    prod_prev['HrsEf'] = _te_raw_p / (3600 if _te_raw_p.max() > 1000 else 60)
    prod_prev['Team'] = prod_prev['Equipo'].map(TEAM_MAP)

    vol_prev = prod_prev['Vol'].sum()
    hrs_prev = prod_prev['HrsEf'].sum()
    dias_prev = prod_prev['Fecha_dt'].dt.day.nunique()
    rend_prev = vol_prev / hrs_prev if hrs_prev > 0 else 0
    prom_prev = vol_prev / dias_prev if dias_prev > 0 else 0

    vol_act = daily['Vol'].sum()
    hrs_act = daily['HrsEf'].sum()
    rend_act = vol_act / hrs_act if hrs_act > 0 else 0
    prom_act = vol_act / ULTIMO_DIA if ULTIMO_DIA > 0 else 0

    comp_mes = {
        'mesPrev': MES_PREV,
        'diasPrev': int(dias_prev),
        'volPrev': round(vol_prev, 1),
        'promPrev': round(prom_prev, 1),
        'rendPrev': round(rend_prev, 2),
        'volAct': round(vol_act, 1),
        'promAct': round(prom_act, 1),
        'rendAct': round(rend_act, 2),
        'deltaVol': round(((vol_act - vol_prev) / vol_prev * 100) if vol_prev > 0 else 0, 1),
        'deltaProm': round(((prom_act - prom_prev) / prom_prev * 100) if prom_prev > 0 else 0, 1),
        'deltaRend': round(((rend_act - rend_prev) / rend_prev * 100) if rend_prev > 0 else 0, 1),
        'suficiente': dias_prev >= 5  # comparativa válida si hay al menos 5 días
    }

# 12. Alertas para Resumen Ejecutivo
# Misma fórmula que el KPI "Proyección" del header: acum + promDia_global × días_restantes
_acum_total = daily['Vol'].sum()
_prom_dia_global = _acum_total / DD if DD > 0 else 0
_proy_total = _acum_total + _prom_dia_global * DR
_meta_total = sum(METAS.values())
proy_total_pct = round(_proy_total / _meta_total * 100, 1) if _meta_total > 0 else 0

# 14. Análisis de Observaciones/Comentarios de Tiempos Perdidos
import re as _re
CATEGORIAS_OBS = [
    ('Fallas técnicas',   '#DC2626', ['falla', 'falta', 'reparac', 'rep\\.', 'fuga', 'cambio', 'pasador', 'cilindro', 'cadena', 'flexible', 'correa', 'espada', 'lexan', 'motor', 'hidr', 'cabezal', 'volteo', 'trozado', 'madereo', 'shovel', 'skidder', 'carro', 'brazo']),
    ('Factores externos', '#F59E0B', ['traslado', 'vehicul', 'tendido', 'carretera', 'camion', 'sodi', 'transport', 'horilla', 'orolla', 'electric']),
    ('Pausas / descanso', '#94A3B8', ['pausa', 'colacion', 'descanso', 'capacitacion', 'charla', 'almuerz']),
    ('Traslados',         '#8B5CF6', ['traslado', 'predio', 'cancha', 'sector', 'cama']),
    ('Cargas / logística','#3B82F6', ['carguio', 'stock', 'distancia', 'apoyo']),
]
def clasificar_obs(txt):
    t = (txt or '').lower()
    for nombre, color, keywords in CATEGORIAS_OBS:
        if any(k in t for k in keywords):
            return nombre, color
    return 'Otros', '#64748B'

obs_analisis = {
    'totalComentarios': 0,
    'porCategoria': [],      # [{'cat': 'Fallas técnicas', 'color': '#DC2626', 'n': 42, 'minutos': 3240}]
    'topComentarios': [],    # [{'txt': '...', 'n': 10, 'minutos': 420, 'cat': 'Pausas'}]
    'porEquipo': {},         # {'Millalemu 1.1': [{'cat':..., 'n':..., 'minutos':...}]}
    'palabrasClave': [],     # [{'palabra': 'falla', 'n': 28}]
}
tm_con_obs = tm.dropna(subset=['Observación']).copy()
tm_con_obs = tm_con_obs[tm_con_obs['Observación'].astype(str).str.strip() != '']
obs_analisis['totalComentarios'] = len(tm_con_obs)

# Clasificación por categoría
cat_agg = {}
for _, r in tm_con_obs.iterrows():
    cat, color = clasificar_obs(r['Observación'])
    if cat not in cat_agg: cat_agg[cat] = {'color': color, 'n': 0, 'minutos': 0}
    cat_agg[cat]['n'] += 1
    cat_agg[cat]['minutos'] += int(r['Tiempo (Min)'])
obs_analisis['porCategoria'] = sorted(
    [{'cat': k, **v} for k, v in cat_agg.items()],
    key=lambda x: -x['minutos']
)

# Top comentarios por frecuencia (con tiempo acumulado)
obs_group = tm_con_obs.groupby('Observación').agg(
    n=('Observación', 'size'),
    minutos=('Tiempo (Min)', 'sum')
).reset_index().sort_values('minutos', ascending=False).head(15)
for _, r in obs_group.iterrows():
    cat, color = clasificar_obs(r['Observación'])
    obs_analisis['topComentarios'].append({
        'txt': str(r['Observación']).strip()[:120],
        'n': int(r['n']),
        'minutos': int(r['minutos']),
        'cat': cat,
        'color': color
    })

# Palabras clave más frecuentes (excluyendo stopwords)
stopwords = {'de', 'del', 'la', 'el', 'y', 'o', 'en', 'por', 'a', 'al', 'un', 'una', 'se', 'que', 'lo', 'los', 'las', 'con', 'para', 'es', 'su', 'como'}
palabras = {}
for txt in tm_con_obs['Observación'].dropna():
    for w in _re.findall(r'\b[a-záéíóúñ]{4,}\b', str(txt).lower()):
        if w not in stopwords:
            palabras[w] = palabras.get(w, 0) + 1
obs_analisis['palabrasClave'] = sorted(
    [{'palabra': k, 'n': v} for k, v in palabras.items() if v >= 2],
    key=lambda x: -x['n']
)[:20]

# Por equipo (categorización)
for t in TEAMS:
    tm_t = tm_con_obs[tm_con_obs['Team'] == t]
    if len(tm_t) == 0: continue
    cats_team = {}
    for _, r in tm_t.iterrows():
        cat, color = clasificar_obs(r['Observación'])
        if cat not in cats_team: cats_team[cat] = {'color': color, 'n': 0, 'minutos': 0}
        cats_team[cat]['n'] += 1
        cats_team[cat]['minutos'] += int(r['Tiempo (Min)'])
    obs_analisis['porEquipo'][t] = sorted(
        [{'cat': k, **v} for k, v in cats_team.items()],
        key=lambda x: -x['minutos']
    )
proy_total_m3 = round(_proy_total, 1)
peor_faena = min(team_kpis, key=lambda k: k['ci']) if team_kpis else None
mejor_faena = max(team_kpis, key=lambda k: k['ci']) if team_kpis else None
total_sin_reporte = sum(1 for dsp in dias_sin_prod)

# 13a. Validación cruzada Selenium vs API (el PG actual es el limpio del Selenium)
validacion_api = {'disponible': False, 'filas': [], 'difTotal': 0, 'fantasmas': 0}
API_REF_CSV = os.path.join(BASE_DIR, ".pg_api_reference.csv")
if os.path.exists(API_REF_CSV):
    try:
        _api = pd.read_csv(API_REF_CSV, sep=';', encoding='utf-8-sig')
        for _c in ['Volumen SSC PU','Volumen SSC AS']:
            _api[_c] = pd.to_numeric(_api[_c].astype(str).str.replace(',','.'), errors='coerce')
        _api['Vol'] = _api['Volumen SSC PU'].fillna(0) + _api['Volumen SSC AS'].fillna(0)
        _api['Team'] = _api['Equipo'].map(TEAM_MAP)
        _api['Fecha_dt'] = pd.to_datetime(_api['Fecha NOC'], dayfirst=True, errors='coerce')
        _api = _api[_api['Fecha_dt'].dt.month == MES]
        _api_por_team = _api.groupby('Team')['Vol'].sum().to_dict()

        total_dif = 0
        for k in team_kpis:
            t = k['t']
            sel_vol = round(k['a'], 2)
            api_vol = round(_api_por_team.get(t, 0), 2)
            dif = round(api_vol - sel_vol, 2)  # API - Selenium = fantasmas eliminados
            total_dif += dif
            validacion_api['filas'].append({
                't': t, 'sel': sel_vol, 'api': api_vol, 'dif': dif
            })
        validacion_api['disponible'] = True
        validacion_api['difTotal'] = round(total_dif, 2)
        validacion_api['fantasmas'] = len(_api) - len(prod)
    except Exception as e:
        print(f"⚠️  No se pudo comparar con API reference: {e}")

# 13. Validación cruzada contra Manual.csv (legacy - solo si existe el archivo)
validacion_manual = {
    'disponible': manual_disponible,
    'filas': []
}
if manual_disponible and vol_oficial_total_eq:
    for k in team_kpis:
        t = k['t']
        dash = round(k['a'], 2)
        oficial = round(vol_oficial_total_eq.get(t, 0), 2)
        dif = round(dash - oficial, 2)
        pct_dif = round((dif / oficial * 100) if oficial > 0 else 0, 2)
        validacion_manual['filas'].append({
            't': t, 'dash': dash, 'of': oficial, 'dif': dif, 'pct': pct_dif
        })

MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# Agregar el mes EN CURSO al histórico de TM (en progreso) y ordenar
if not any(m['mes'] == MES and m['anio'] == ANIO for m in tm_mensual_list):
    _mant_h = round(tm_cat_total.get('Mantención', 0) / 60, 1)
    _oper_h = round(tm_cat_total.get('Operacional', 0) / 60, 1)
    _proc_h = round(tm_cat_total.get('Proceso', 0) / 60, 1)
    tm_mensual_list.append({
        'mes': MES, 'anio': ANIO, 'mesNombre': MESES[MES] + ' (en curso)',
        'mant': _mant_h, 'oper': _oper_h, 'proc': _proc_h,
        'prog': round(tm_prog_min / 60, 1),
        'perdido': round(_mant_h + _oper_h + _proc_h, 1), 'topCausas': '',
    })
tm_mensual_list = sorted(tm_mensual_list, key=lambda m: (m['anio'], m['mes']))

data_json = json.dumps({
    'cfg': {'mes':MES,'anio':ANIO,'dm':DM,'dd':DD,'dt':DT,'dr':DR,
            'ta': round(daily['Vol'].sum(),1), 'tm': sum(METAS.values()),
            'mesNombre': MESES[MES]},
    'kpis': team_kpis,
    'grid': {str(k):v for k,v in grid.items()},
    'gridTM': {str(k):v for k,v in gridTM.items()},
    'gridHrs': {str(k):v for k,v in gridHrs.items()},
    'trend': trend_list,
    'tmTop': tm_list,
    'tmCat': tm_cat_list,
    'tmTrend': tm_trend_list,
    'tmTeamCat': tm_team_cat,
    'esp': esp_list,
    'avance': avance_diario,
    'rendEsp': rend_especie,
    'rendEspTeam': rend_esp_team,
    'tmTeamCauses': tm_team_causes,
    'tmParetoGlobal': tm_pareto_global,
    'tmHeatmapCE': tm_heatmap_causa_equipo,
    'tmRecomendaciones': tm_recomendaciones,
    'tendencia': tendencia_team,
    'diasSinProd': dias_sin_prod,
    'predios': predio_stats,
    'rankingEf': ranking_ef,
    'espMix': esp_mix,
    'tmHeatmap': {str(k):v for k,v in tm_heatmap.items()},
    'mtbfMttr': mtbf_mttr,
    'compMes': comp_mes,
    'historico': historico_list,
    'tmMensual': tm_mensual_list,
    'snapshotsDisponibles': sorted([
        d for d in os.listdir(os.path.join(BASE_DIR, '_snapshots'))
        if os.path.isdir(os.path.join(BASE_DIR, '_snapshots', d))
        and os.path.exists(os.path.join(BASE_DIR, '_snapshots', d, 'data_diario.csv'))
    ]) if os.path.isdir(os.path.join(BASE_DIR, '_snapshots')) else [],
    'obsAnalisis': obs_analisis,
    'resumenEj': {
        'proyPct': proy_total_pct,
        'proyM3': proy_total_m3,
        'peor': peor_faena,
        'mejor': mejor_faena,
        'sinReporte': total_sin_reporte
    },
    'validacion': validacion_manual,
    'validacionApi': validacion_api,
    'teams': TEAMS,
    'generado': datetime.now().strftime('%d/%m/%Y %H:%M')
}, ensure_ascii=False)

# ── Modal de gráfico ampliado (clic en un gráfico → se despliega grande) ──
# Definido como strings Python normales para no doblar llaves dentro del f-string.
MODAL_CSS = '''
/* === Modal de grafico ampliado === */
.chart-card:has(canvas) { cursor: pointer; position: relative; transition: box-shadow .15s ease, transform .15s ease; }
.chart-card:has(canvas):hover { box-shadow: 0 6px 20px rgba(0,0,0,0.13); transform: translateY(-1px); }
.chart-card:has(canvas)::after { content: "\\2922"; position: absolute; top: 12px; right: 14px; font-size: 15px; color: #94A3B8; opacity: .5; transition: opacity .15s, color .15s; pointer-events: none; }
.chart-card:has(canvas):hover::after { opacity: 1; color: #1A5276; }
#chartModalOverlay { position: fixed; inset: 0; background: rgba(15,23,42,0.62); display: flex; align-items: center; justify-content: center; padding: 24px; z-index: 1000; opacity: 0; pointer-events: none; transition: opacity .22s ease; }
#chartModalOverlay.open { opacity: 1; pointer-events: auto; }
#chartModalPanel { background: #fff; border-radius: 16px; width: min(920px, 96vw); max-height: 92vh; display: flex; flex-direction: column; overflow: hidden; box-shadow: 0 24px 60px rgba(0,0,0,0.35); transform: translateY(14px) scale(.97); opacity: 0; transition: transform .22s cubic-bezier(.2,.8,.2,1), opacity .22s ease; }
#chartModalOverlay.open #chartModalPanel { transform: translateY(0) scale(1); opacity: 1; }
#chartModalHead { display: flex; align-items: center; justify-content: space-between; gap: 12px; padding: 15px 20px; border-bottom: 1px solid #E2E8F0; }
#chartModalTitle { font-size: 16px; font-weight: 800; color: #1A5276; margin: 0; }
#chartModalClose { border: none; background: #F1F5F9; color: #475569; width: 34px; height: 34px; border-radius: 9px; font-size: 20px; line-height: 1; cursor: pointer; flex: none; transition: background .15s; }
#chartModalClose:hover { background: #E2E8F0; color: #0F172A; }
#chartModalBody { padding: 18px 20px 22px; overflow-y: auto; }
#chartModalCanvasWrap { position: relative; height: min(46vh, 420px); }
#chartModalTable { width: 100%; border-collapse: collapse; margin-top: 16px; font-size: 13px; }
#chartModalTable th, #chartModalTable td { padding: 7px 10px; text-align: right; border-bottom: 1px solid #EEF2F6; }
#chartModalTable th:first-child, #chartModalTable td:first-child { text-align: left; }
#chartModalTable thead th { color: #64748B; font-weight: 700; font-size: 11px; text-transform: uppercase; letter-spacing: .4px; border-bottom: 2px solid #E2E8F0; }
#chartModalTable tbody tr:hover { background: #F8FAFC; }
@media (max-width: 640px) {
  #chartModalOverlay { padding: 0; }
  #chartModalPanel { width: 100vw; height: 100%; max-height: 100vh; border-radius: 0; }
  #chartModalCanvasWrap { height: 42vh; }
}
'''

MODAL_HTML = '''
<div id="chartModalOverlay" role="dialog" aria-modal="true" aria-label="Grafico ampliado">
  <div id="chartModalPanel">
    <div id="chartModalHead">
      <h3 id="chartModalTitle">Grafico</h3>
      <button id="chartModalClose" aria-label="Cerrar">&times;</button>
    </div>
    <div id="chartModalBody">
      <div id="chartModalCanvasWrap"><canvas id="chartModalCanvas"></canvas></div>
      <table id="chartModalTable"></table>
    </div>
  </div>
</div>
'''

MODAL_JS = r"""
(function(){
  function setup(){
  var overlay = document.getElementById('chartModalOverlay');
  if (!overlay) return;
  var titleEl = document.getElementById('chartModalTitle');
  var tableEl = document.getElementById('chartModalTable');
  var canvas  = document.getElementById('chartModalCanvas');
  var modalChart = null;

  function fmtNum(v){
    if (v === null || v === undefined || v === '') return '';
    if (typeof v === 'number') return v.toLocaleString('es-CL', {maximumFractionDigits:1});
    if (typeof v === 'object') return (v && v.y != null) ? v.y : '';
    return String(v);
  }

  function buildTable(data){
    var ds = data.datasets || [];
    var labels = data.labels || [];
    if (!labels.length && ds.length && ds[0].data) labels = ds[0].data.map(function(_, i){ return 'Item ' + (i + 1); });
    var head = '<thead><tr><th>Categoria</th>' + ds.map(function(d){ return '<th>' + (d.label || 'Serie') + '</th>'; }).join('') + '</tr></thead>';
    var body = '<tbody>' + labels.map(function(lab, i){
      return '<tr><td>' + lab + '</td>' + ds.map(function(d){ return '<td>' + fmtNum(d.data ? d.data[i] : '') + '</td>'; }).join('') + '</tr>';
    }).join('') + '</tbody>';
    tableEl.innerHTML = head + body;
  }

  function openModal(srcChart, title){
    titleEl.textContent = title || 'Grafico';
    var data;
    try { data = JSON.parse(JSON.stringify(srcChart.config.data)); }
    catch (err) { data = srcChart.config.data; }
    var opts = Object.assign({}, srcChart.config.options, {responsive:true, maintainAspectRatio:false, animation:{duration:300}});
    if (modalChart) { modalChart.destroy(); modalChart = null; }
    modalChart = new Chart(canvas.getContext('2d'), {type: srcChart.config.type, data: data, options: opts});
    buildTable(data);
    overlay.classList.add('open');
    document.body.style.overflow = 'hidden';
  }

  function closeModal(){
    overlay.classList.remove('open');
    document.body.style.overflow = '';
    setTimeout(function(){ if (modalChart) { modalChart.destroy(); modalChart = null; } }, 240);
  }

  document.addEventListener('click', function(e){
    if (overlay.contains(e.target)) return;
    var card = e.target.closest ? e.target.closest('.chart-card') : null;
    if (!card) return;
    var cv = card.querySelector('canvas');
    if (!cv) return;
    var ch = (typeof Chart !== 'undefined' && Chart.getChart) ? Chart.getChart(cv) : null;
    if (!ch) return;
    var h3 = card.querySelector('h3');
    openModal(ch, h3 ? h3.textContent.trim() : 'Grafico');
  });

  document.getElementById('chartModalClose').addEventListener('click', closeModal);
  overlay.addEventListener('click', function(e){ if (e.target === overlay) closeModal(); });
  document.addEventListener('keydown', function(e){ if (e.key === 'Escape' && overlay.classList.contains('open')) closeModal(); });
  }
  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', setup);
  else setup();
})();
"""

# ── HTML Template ────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<title>Dashboard Cosecha - Millalemu {MESES[MES]} {ANIO}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
  :root {{
    --primary: #1A5276; --primary-dark: #154360; --accent: #1E8449;
    --warning: #E67E22; --danger: #C0392B; --bg: #F8FAFC;
    --card: #FFFFFF; --subtle: #F1F5F9; --border: #E2E8F0;
    --text: #1E293B; --text-light: #64748B; --text-muted: #94A3B8;
  }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Segoe UI', -apple-system, sans-serif; background: var(--bg); color: var(--text); padding: 16px; }}

  .header {{
    background: linear-gradient(135deg, var(--primary), var(--primary-dark));
    border-radius: 16px; padding: 20px 28px; margin-bottom: 20px; color: white;
    display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 12px;
  }}
  .header h1 {{ font-size: 22px; font-weight: 800; letter-spacing: -0.5px; }}
  .header .subtitle {{ font-size: 13px; opacity: 0.8; margin-top: 4px; }}
  .header .kpi-row {{ display: flex; gap: 28px; flex-wrap: wrap; }}
  .header .kpi {{ text-align: center; }}
  .header .kpi-label {{ font-size: 11px; opacity: 0.7; }}
  .header .kpi-value {{ font-size: 24px; font-weight: 800; }}
  .header .kpi-unit {{ font-size: 10px; opacity: 0.6; }}

  .nav {{ display: flex; gap: 8px; margin-bottom: 16px; flex-wrap: wrap; }}
  .nav button {{
    padding: 8px 18px; border-radius: 8px; border: none; cursor: pointer;
    font-size: 13px; font-weight: 600; background: var(--card); color: var(--text-light);
    box-shadow: 0 1px 2px rgba(0,0,0,0.05); transition: all 0.2s;
  }}
  .nav button.active {{ background: var(--primary); color: white; box-shadow: 0 2px 8px rgba(26,82,118,0.3); }}

  .cards {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(260px, 1fr)); gap: 12px; margin-bottom: 24px; }}
  .team-card {{
    background: var(--card); border-radius: 12px; padding: 14px 16px;
    border: 2px solid transparent; cursor: pointer; box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    transition: all 0.2s; position: relative;
  }}
  .team-card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.12); }}
  .team-card .team-name {{ font-size: 13px; font-weight: 700; }}
  .team-card .team-detail {{ font-size: 10px; color: var(--text-muted); }}
  .team-card .kpi-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 6px; font-size: 11px; margin-top: 8px; }}
  .team-card .kpi-grid .label {{ color: var(--text-muted); }}
  .team-card .kpi-grid .value {{ font-weight: 700; font-size: 14px; }}

  .gauge-container {{ display: flex; align-items: flex-start; justify-content: space-between; margin-bottom: 4px; }}

  .charts-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px; }}
  .chart-card {{ background: var(--card); border-radius: 12px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
  .chart-card h3 {{ margin-bottom: 12px; font-size: 14px; color: var(--primary); }}

  .grid-container {{ background: var(--card); border-radius: 12px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  /* gridTable: usar separate para que las líneas de Total/Promedio sean continuas */
  #gridTable {{ border-collapse: separate; border-spacing: 0; }}
  th {{ padding: 8px 10px; text-align: right; }}
  th:first-child {{ text-align: left; }}
  thead tr {{ background: var(--primary); color: white; }}
  thead th:first-child {{ border-radius: 8px 0 0 0; }}
  thead th:last-child {{ border-radius: 0 8px 0 0; }}
  td {{ padding: 6px 10px; text-align: right; }}
  td:first-child {{ text-align: left; font-weight: 600; color: var(--primary); }}
  tr:nth-child(even) {{ background: var(--subtle); }}
  .total-row {{ background: var(--primary) !important; color: white; font-weight: 700; }}
  .total-row td {{ color: white; }}

  .cell-high {{ background: #D1FAE5; }}
  .cell-med {{ background: #FEF9C3; }}
  .cell-low {{ background: #FED7AA; }}
  .cell-zero {{ background: #FEE2E2; }}
  .cell-high-soft {{ background: #ECFDF5; }}
  .cell-med-soft {{ background: #FEFCE8; }}
  .cell-low-soft {{ background: #FFF7ED; }}

  .tm-section {{ display: grid; grid-template-columns: 2fr 1fr; gap: 16px; }}
  .bar-item {{ margin-bottom: 8px; }}
  .bar-item .bar-header {{ display: flex; justify-content: space-between; font-size: 11px; margin-bottom: 2px; }}
  .bar-item .bar-track {{ height: 6px; background: var(--subtle); border-radius: 3px; }}
  .bar-item .bar-fill {{ height: 6px; border-radius: 3px; transition: width 0.5s; }}

  .team-detail {{
    background: var(--card); border-radius: 16px; padding: 20px; margin-bottom: 20px;
    box-shadow: 0 8px 24px rgba(0,0,0,0.08);
  }}
  .team-detail .detail-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px; }}
  .team-detail .kpi-cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px; margin-bottom: 20px; }}
  .team-detail .kpi-card {{ background: #F8FAFC; border-radius: 8px; padding: 8px 12px; }}
  .team-detail .kpi-card .kpi-card-label {{ font-size: 10px; color: var(--text-muted); }}
  .team-detail .kpi-card .kpi-card-value {{ font-size: 16px; font-weight: 700; }}
  .team-detail .detail-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; overflow: hidden; }}
  .team-detail table {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 16px; }}
  .disp-bar {{ margin-top: 8px; border-radius: 6px; padding: 4px 8px; display: flex; justify-content: space-between; align-items: center; }}
  .disp-track {{ width: 50px; height: 5px; border-radius: 3px; background: #E5E7EB; overflow: hidden; display: inline-block; }}
  .disp-fill {{ height: 5px; border-radius: 3px; }}
  .rank-badge {{ position: absolute; top: 6px; right: 8px; font-size: 16px; }}
  .rank-badge-bottom {{ position: absolute; top: 6px; right: 8px; font-size: 9px; background: #FEE2E2; color: #991B1B; border-radius: 4px; padding: 2px 5px; font-weight: 600; }}
  .rank-badge-mid {{ position: absolute; top: 6px; right: 8px; font-size: 9px; background: #F1F5F9; color: #64748B; border-radius: 4px; padding: 2px 5px; font-weight: 600; }}
  .semaforo {{ display: flex; gap: 12px; margin-bottom: 16px; flex-wrap: wrap; }}
  .semaforo-pill {{ display: flex; align-items: center; gap: 6px; border-radius: 8px; padding: 6px 14px; }}
  .alertas-panel {{ display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-bottom: 16px; }}
  .alerta {{ display: flex; align-items: flex-start; gap: 10px; border-radius: 10px; padding: 10px 16px; font-size: 13px; line-height: 1.4; }}
  .alerta-critica {{ background: #FEF2F2; border-left: 4px solid #DC2626; }}
  .alerta-advertencia {{ background: #FFF7ED; border-left: 4px solid #F59E0B; }}
  .alerta-icono {{ font-size: 16px; flex-shrink: 0; margin-top: 1px; }}
  .alerta-texto {{ flex: 1; }}
  .alerta-titulo {{ font-weight: 700; color: #1E293B; }}
  .alerta-detalle {{ color: #64748B; font-size: 12px; margin-top: 2px; }}
  .semaforo-dot {{ width: 10px; height: 10px; border-radius: 50%; }}
  .dbl-hint {{ font-size: 9px; color: #CBD5E1; text-align: center; margin-top: 6px; }}
  .close-btn {{ background: #F1F5F9; border: none; border-radius: 8px; padding: 6px 14px; cursor: pointer; font-size: 12px; font-weight: 600; color: #64748B; }}
  .close-btn:hover {{ background: #E2E8F0; }}
  .tm-box {{ border-radius: 8px; padding: 8px 10px; text-align: center; }}
  .footer {{ text-align: center; margin-top: 24px; padding: 12px; font-size: 11px; color: var(--text-muted); }}
  .section {{ display: none; }}
  .section.active {{ display: block; }}

  @media (max-width: 768px) {{
    body {{ padding: 8px; }}
    .charts-row, .tm-section, .alertas-panel {{ grid-template-columns: 1fr; }}
    .header {{ flex-direction: column; text-align: center; padding: 14px 16px; }}
    .header h1 {{ font-size: 18px; }}
    .header .kpi-row {{ justify-content: center; gap: 16px; }}
    .cards {{ grid-template-columns: 1fr 1fr; gap: 8px; }}
    .team-card {{ padding: 10px 12px; }}
    .team-card .kpi-grid {{ grid-template-columns: 1fr 1fr 1fr; gap: 4px; font-size: 10px; }}
    .team-card .kpi-grid .value {{ font-size: 12px; }}
    .team-detail {{ padding: 14px; border-radius: 12px; }}
    .team-detail .kpi-cards {{ grid-template-columns: repeat(2, 1fr); gap: 6px; }}
    .team-detail .kpi-card .kpi-card-value {{ font-size: 14px; }}
    .team-detail .detail-grid {{ grid-template-columns: 1fr !important; }}
    .team-detail .detail-grid > div {{ width: 100% !important; max-width: 100% !important; overflow: hidden; }}
    .team-detail .detail-grid canvas {{ max-width: 100%; height: auto !important; }}
    .team-detail .detail-header {{ flex-direction: column; gap: 8px; text-align: center; }}
    .nav {{ gap: 4px; }}
    .nav button {{ padding: 6px 12px; font-size: 11px; }}
    .grid-container {{ padding: 10px; }}
    table {{ font-size: 10px; }}
    th, td {{ padding: 4px 6px; }}
    .semaforo {{ gap: 6px; }}
    .semaforo-pill {{ padding: 4px 8px; font-size: 11px; }}
    .gauge-container {{ flex-direction: column; align-items: center; gap: 8px; }}
  }}
  @media (max-width: 420px) {{
    .cards {{ grid-template-columns: 1fr; }}
    .team-detail .kpi-cards {{ grid-template-columns: 1fr 1fr; }}
    .team-detail .kpi-card .kpi-card-value {{ font-size: 13px; word-break: break-word; }}
    .tm-box {{ padding: 6px 4px !important; }}
    .tm-box div:last-child {{ font-size: 14px !important; }}
  }}
{MODAL_CSS}
</style>
</head>
<body>

<div class="header">
  <div>
    <h1>Control de Cosecha Forestal</h1>
    <p class="subtitle">Forestal Millalemu | {MESES[MES]} {ANIO} | {DD} de {DT} días trabajados</p>
    {'<p class="subtitle" style="color:#FBBF24;font-weight:600;margin:3px 0 0">⚠️ Datos preliminares: proyección con pocos días, puede variar</p>' if DD <= 5 else ''}
    <p style="margin:2px 0 0;font-size:11px;opacity:0.5">Datos al: {ULTIMO_DIA}-{MESES[MES][:3]}-{ANIO}</p>
  </div>
  <div class="kpi-row" id="headerKpis"></div>
</div>

<div class="nav" id="nav"></div>

<!-- Segmentador global de MES -->
<div id="filtroBar" style="display:flex;align-items:center;gap:14px;background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;padding:10px 14px;margin-bottom:14px;flex-wrap:wrap;font-size:13px">
  <div style="display:flex;align-items:center;gap:6px">
    <span style="color:#64748B">📅 Mes:</span>
    <select id="selectMes" style="padding:4px 8px;border-radius:6px;border:1px solid #CBD5E1;background:white;font-size:13px;font-weight:600;color:#1A5276">
      <option value="__current__" selected>Mes en curso</option>
    </select>
  </div>
  <span id="filtroMesPill" style="display:none;background:#DBEAFE;color:#1E40AF;padding:3px 10px;border-radius:99px;font-size:11px;font-weight:600">Mes seleccionado</span>
  <button id="btnLimpiarFiltro" style="display:none;margin-left:auto;padding:4px 10px;border-radius:6px;border:1px solid #CBD5E1;background:white;cursor:pointer;font-size:11px;color:#475569">
    ✕ Limpiar filtro
  </button>
</div>

<div id="avisoMesHistorico" style="display:none;background:#EFF6FF;border-left:4px solid #3B82F6;padding:10px 14px;border-radius:8px;margin-bottom:14px;font-size:12px;color:#1E40AF">
  📋 <strong>Vista mes histórico:</strong> los datos crudos (Resumen/Tabla de Producciones/Tiempos Perdidos) solo están disponibles para el mes en curso.
  Para meses anteriores se muestra el resumen agregado en la pestaña <strong>Comparativo Mensual</strong>.
</div>

<div id="resumenEjecutivo"></div>

<!-- Panel de alertas (CRÍTICAS/ADVERTENCIAS) movido a la pestaña "Análisis Operacional" -->

<div class="section active" id="sec-resumen">
  <div id="teamDetail"></div>
  <div class="cards" id="teamCards"></div>
  <div class="chart-card"><h3>Producción Diaria vs Meta Diaria Requerida (m³ SSC)</h3><canvas id="chartTrend" style="min-height:280px"></canvas></div>
  <div class="charts-row" style="margin-top:16px">
    <div class="chart-card"><h3>Avance Acumulado vs Plan Mensual</h3><canvas id="chartAvance"></canvas></div>
    <div class="chart-card"><h3>Comparativo por Equipo: Acumulado vs Proyección vs Meta</h3><canvas id="chartTeams"></canvas></div>
  </div>
</div>

<div class="section" id="sec-grid">
  <div class="grid-container"><h3 style="margin-bottom:12px;font-size:14px;color:var(--primary)">Control Mensual por Equipo</h3><table id="gridTable"></table></div>
</div>

<div class="section" id="sec-tiempos">
  <div id="tmSummaryCards" style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px"></div>
  <div class="tm-section" style="margin-bottom:16px">
    <div class="chart-card"><h3>Top 10 Causas de Tiempos Perdidos</h3><canvas id="chartTM"></canvas></div>
    <div>
      <div class="chart-card" style="margin-bottom:16px"><h3>Distribución por Categoría</h3><canvas id="chartTMCat" style="max-height:160px"></canvas></div>
      <div class="chart-card"><h3>Producción por Especie</h3><canvas id="chartEsp" style="max-height:140px"></canvas></div>
    </div>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
    <div class="chart-card"><h3>Tendencia Diaria TM (horas)</h3><canvas id="chartTMTrend"></canvas></div>
    <div class="chart-card"><h3>TM por Equipo Desglosado</h3><canvas id="chartTMTeam"></canvas></div>
  </div>
  <div class="chart-card"><h3>Resumen Disponibilidad y Tiempos por Equipo</h3><div id="tmDispTable"></div></div>
  <div class="chart-card" style="margin-top:16px"><h3>Top 5 Causas de Pérdida por Faena</h3><div id="tmTeamCausesTable"></div></div>

  <!-- ═══ ANÁLISIS 80/20 ENRIQUECIDO PARA PLANIFICAR MAYO ═══ -->
  <div class="chart-card" style="margin-top:24px;border-top:4px solid #D97706">
    <h3 style="display:flex;align-items:center;gap:8px">⚡ Análisis Pareto 80/20 — Faena Completa</h3>
    <div id="tmParetoBox"></div>
    <canvas id="chartParetoGlobal" style="max-height:340px;margin-top:8px"></canvas>
  </div>

  <div class="chart-card" style="margin-top:16px">
    <h3>🔥 Focos Crónicos: Causas vs Equipos</h3>
    <div style="font-size:11px;color:#64748B;margin-bottom:8px">Las celdas muestran horas perdidas. Color rojo = concentración del problema en ese equipo.</div>
    <div id="tmHeatmapCETable" style="overflow-x:auto"></div>
  </div>

  <div class="chart-card" style="margin-top:16px;background:linear-gradient(135deg,#F0FDF4 0%,#ECFCCB 100%);border-left:5px solid #16A34A">
    <h3 style="color:#166534">🎯 Recomendaciones Accionables para Mayo</h3>
    <div style="font-size:11px;color:#475569;margin-bottom:12px">Basado en patrones del mes cerrado. Ahorro potencial = reducir 30% del tiempo perdido por la causa.</div>
    <div id="tmRecBox"></div>
  </div>
</div>

<div class="section" id="sec-analisis">
  <!-- Panel de alertas y semáforo movido aquí desde arriba -->
  <div class="semaforo" id="semaforo" style="margin-bottom:12px"></div>
  <div class="alertas-panel" id="alertasPanel" style="margin-bottom:16px"></div>
  <div class="chart-card"><h3>Cumplimiento Diario por Faena</h3><div id="diasSinProdTable"></div></div>
  <div class="charts-row" style="margin-top:16px">
    <div class="chart-card"><h3>Ranking de Eficiencia por Faena (m³/hr)</h3><div id="rankingEfTable"></div></div>
    <div class="chart-card"><h3>Mix de Producción por Especie</h3><div id="espMixBox"></div></div>
  </div>
  <div class="chart-card" style="margin-top:16px"><h3>Análisis por Predio</h3><div id="prediosTable"></div></div>
  <div class="charts-row" style="margin-top:16px">
    <div class="chart-card"><h3>Rendimiento por Especie (m³/hr)</h3><canvas id="chartRendEsp"></canvas></div>
    <div class="chart-card"><h3>Rendimiento por Especie y Faena</h3><canvas id="chartRendEspTeam"></canvas></div>
  </div>
  <div class="chart-card" style="margin-top:16px"><h3>Heatmap de Tiempos Perdidos (horas por día/faena)</h3><div id="tmHeatmapBox"></div></div>
  <div class="charts-row" style="margin-top:16px">
    <div class="chart-card"><h3>Disponibilidad Mecánica (MTBF / MTTR)</h3><div id="mtbfMttrTable"></div></div>
    <div class="chart-card"><h3>Comparativa vs Mes Anterior</h3><div id="compMesBox"></div></div>
  </div>
  <div class="chart-card" style="margin-top:16px"><h3>Análisis de Observaciones (Tiempos Perdidos)</h3><div id="obsAnalisisBox"></div></div>
</div>

<div class="section" id="sec-comparativo">
  <div class="chart-card" style="margin-bottom:16px">
    <h3>Resumen Histórico — Total Faena por Mes</h3>
    <div id="histKpis" style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:12px"></div>
    <canvas id="chartHistTotal" style="max-height:280px"></canvas>
  </div>
  <div class="chart-card" style="margin-bottom:16px">
    <h3>Producción por Equipo y Mes (m³ SSC)</h3>
    <div id="histTablaVol" style="overflow-x:auto"></div>
  </div>
  <div class="chart-card" style="margin-bottom:16px">
    <h3>Cumplimiento % por Equipo y Mes</h3>
    <div id="histTablaCumpl" style="overflow-x:auto"></div>
  </div>
  <div class="chart-card">
    <h3>Tendencia por Equipo (m³ mensuales)</h3>
    <canvas id="chartHistEquipos" style="max-height:340px"></canvas>
  </div>
  <div class="chart-card" style="margin-top:16px">
    <h3>Tiempos Perdidos por Mes (horas)</h3>
    <canvas id="chartTmMensual" style="max-height:300px"></canvas>
    <div id="tmMensualTabla" style="overflow-x:auto;margin-top:14px"></div>
  </div>
</div>

<div class="footer">Dashboard Control Cosecha | Forestal Millalemu | Generado: <span id="genDate"></span></div>

<script>
const D = {data_json};
const TEAMS = D.teams;
const COLORS = ['#1A5276','#2980B9','#1ABC9C','#27AE60','#8E44AD','#E67E22','#C0392B','#2C3E50'];
const fmt = n => n != null ? n.toLocaleString('es-CL', {{maximumFractionDigits:1}}) : '—';
const NEUTRAL = '#334155';
const BAD = '#C0392B';

// Header KPIs
const cfg = D.cfg;
const promDia = cfg.ta / cfg.dd;
// Mes cerrado = no quedan días por trabajar (DR<=1 cuando dd>=dt o dt-dd<=0)
const mesCerrado = cfg.dd >= cfg.dt;
const proyTotal = mesCerrado ? cfg.ta : (cfg.ta + promDia * cfg.dr);
const cumplTotal = (proyTotal / cfg.tm * 100).toFixed(1);
const cierreColor = parseFloat(cumplTotal)>=80?'#A3E635':parseFloat(cumplTotal)>=60?'#FDE68A':'#FCA5A5';
const brechaVal = proyTotal - cfg.tm;
const brechaColor = brechaVal >= 0 ? '#A3E635' : '#FCA5A5';
const tituloKpi = mesCerrado ? 'Cierre Real' : 'Cierre Proy.';
const tituloProy = mesCerrado ? 'Total Mes' : 'Proyección';
const sufijoBrecha = brechaVal >= 0 ? 'Sobre meta' : (mesCerrado ? 'Brecha real: ' : 'Brecha: ') + fmt(Math.round(brechaVal));
document.getElementById('headerKpis').innerHTML = `
  <div class="kpi" style="min-width:100px"><div class="kpi-label">Acumulado</div><div class="kpi-value" style="font-size:22px">${{fmt(Math.round(cfg.ta))}}</div><div class="kpi-unit">m³ SSC</div></div>
  <div class="kpi" style="min-width:100px"><div class="kpi-label">Meta</div><div class="kpi-value" style="font-size:22px">${{fmt(cfg.tm)}}</div><div class="kpi-unit">m³ SSC</div></div>
  <div class="kpi" style="min-width:170px"><div class="kpi-label" style="font-weight:700">${{tituloKpi}}</div><div class="kpi-value" style="color:${{cierreColor}};font-size:52px;line-height:1;font-weight:900">${{cumplTotal}}%</div></div>
  <div class="kpi" style="min-width:150px"><div class="kpi-label" style="font-weight:700">${{tituloProy}}</div><div class="kpi-value" style="font-size:38px;line-height:1;font-weight:800">${{fmt(Math.round(proyTotal))}}</div><div class="kpi-unit" style="color:${{brechaColor}};font-weight:700">${{sufijoBrecha}}</div></div>
`;

// Navigation
const sections = ['resumen','grid','tiempos','analisis','comparativo'];
const labels = ['Resumen por Equipo','Tabla de Producciones','Tiempos Perdidos','Análisis Operacional','Comparativo Mensual'];
const nav = document.getElementById('nav');
sections.forEach((s,i) => {{
  const btn = document.createElement('button');
  btn.textContent = labels[i];
  btn.className = i === 0 ? 'active' : '';
  btn.onclick = () => {{
    document.querySelectorAll('.section').forEach(el => el.classList.remove('active'));
    document.getElementById('sec-'+s).classList.add('active');
    document.querySelectorAll('.nav button').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
  }};
  nav.appendChild(btn);
}});

// Semáforo summary
const enMeta = D.kpis.filter(k => (k.pr/k.m*100) >= 90).length;
const conBrecha = D.kpis.length - enMeta;
const avgDisp = D.kpis.reduce((s,k) => s + (k.turno>0 ? (1-k.tm/k.turno)*100 : 100), 0) / D.kpis.length;
const semaforoEl = document.getElementById('semaforo');
semaforoEl.innerHTML = `
  <div class="semaforo-pill" style="background:#F0FDF4"><div class="semaforo-dot" style="background:#1E8449"></div><span style="font-size:13px;font-weight:600;color:#166534">${{enMeta}} ${{enMeta===1?'equipo':'equipos'}} en meta</span></div>
  <div class="semaforo-pill" style="background:#FEF2F2"><div class="semaforo-dot" style="background:#C0392B"></div><span style="font-size:13px;font-weight:600;color:#991B1B">${{conBrecha}} con brecha</span></div>
  <div class="semaforo-pill" style="background:${{avgDisp>=80?'#F0FDF4':avgDisp>=60?'#FFF7ED':'#FEF2F2'}}"><span style="font-size:13px;font-weight:600;color:${{avgDisp>=80?'#166534':avgDisp>=60?'#92400E':'#C0392B'}}">Disp. promedio: ${{avgDisp.toFixed(1)}}%</span></div>
`;

// ── Alertas Ejecutivas ──────────────────────────────────
const alertas = [];
const expectedPctDay = (cfg.dd / cfg.dt * 100);

// 1. Disponibilidad < 80% → CRÍTICA
D.kpis.forEach(k => {{
  const disp = k.turno > 0 ? (1 - k.tm / k.turno) * 100 : 100;
  if (disp < 60) {{
    alertas.push({{ tipo: 'critica', icono: '🔴', valor: disp,
      titulo: `${{k.t}} — Disponibilidad crítica: ${{disp.toFixed(1)}}%`,
      detalle: `TM Mantención: ${{(k.tm/60).toFixed(0)}} hrs. Intervención inmediata.`
    }});
  }} else if (disp < 80) {{
    alertas.push({{ tipo: 'advertencia', icono: '🟡', valor: disp,
      titulo: `${{k.t}} — Disponibilidad bajo meta: ${{disp.toFixed(1)}}%`,
      detalle: `TM Mantención: ${{(k.tm/60).toFixed(0)}} hrs. Revisar mantenimiento preventivo.`
    }});
  }}
}});

// 2. Proyección < 80% de meta → riesgo de incumplimiento
D.kpis.forEach(k => {{
  const pctProy = k.pr / k.m * 100;
  if (pctProy < 70) {{
    alertas.push({{ tipo: 'critica', icono: '📉', valor: pctProy,
      titulo: `${{k.t}} — Riesgo alto: proyección al ${{pctProy.toFixed(0)}}%`,
      detalle: `Proyecta ${{fmt(k.pr)}} vs meta ${{fmt(k.m)}} m³. Brecha: ${{fmt(Math.abs(k.b))}} m³.`
    }});
  }} else if (pctProy < 90) {{
    alertas.push({{ tipo: 'advertencia', icono: '⚠️', valor: pctProy,
      titulo: `${{k.t}} — Proyección ajustada: ${{pctProy.toFixed(0)}}% de meta`,
      detalle: `Necesita ${{fmt(Math.round((k.m - k.a) / Math.max(cfg.dr, 1)))}} m³/día. Actual: ${{fmt(k.p)}} m³/día.`
    }});
  }}
}});

// 3. Caída de producción (último día vs promedio)
const lastDay = Math.max(...Object.keys(D.grid).map(Number));
D.kpis.forEach(k => {{
  const volLastDay = D.grid[lastDay] && D.grid[lastDay][k.t] ? D.grid[lastDay][k.t] : 0;
  if (k.p > 0 && volLastDay < k.p * 0.5 && volLastDay > 0) {{
    alertas.push({{ tipo: 'advertencia', icono: '📊', valor: volLastDay/k.p*100,
      titulo: `${{k.t}} — Caída de producción día ${{lastDay}}`,
      detalle: `Produjo ${{fmt(volLastDay)}} m³ vs promedio ${{fmt(k.p)}} m³/día (${{(volLastDay/k.p*100).toFixed(0)}}%).`
    }});
  }}
}});

// Render alertas
const alertasEl = document.getElementById('alertasPanel');
if (alertas.length > 0) {{
  const criticas = alertas.filter(a => a.tipo === 'critica').sort((a,b) => (a.valor||50) - (b.valor||50));
  const advertencias = alertas.filter(a => a.tipo === 'advertencia').sort((a,b) => (a.valor||50) - (b.valor||50));
  let html = '';
  if (criticas.length > 0) {{
    html += `<div style="grid-column:1/-1;font-size:11px;font-weight:700;color:#991B1B;text-transform:uppercase;letter-spacing:0.5px">🔴 Críticas</div>`;
    html += criticas.map(a => `<div class="alerta alerta-critica"><span class="alerta-icono">${{a.icono}}</span><div class="alerta-texto"><div class="alerta-titulo">${{a.titulo}}</div><div class="alerta-detalle">${{a.detalle}}</div></div></div>`).join('');
  }}
  if (advertencias.length > 0) {{
    html += `<div style="grid-column:1/-1;font-size:11px;font-weight:700;color:#92400E;text-transform:uppercase;letter-spacing:0.5px;${{criticas.length>0?'margin-top:4px':''}}">⚠️ Advertencias</div>`;
    html += advertencias.map(a => `<div class="alerta alerta-advertencia"><span class="alerta-icono">${{a.icono}}</span><div class="alerta-texto"><div class="alerta-titulo">${{a.titulo}}</div><div class="alerta-detalle">${{a.detalle}}</div></div></div>`).join('');
  }}
  alertasEl.innerHTML = html;
}} else {{
  alertasEl.innerHTML = '<div class="alerta" style="background:#F0FDF4;border-left:4px solid #16A34A"><span class="alerta-icono">✅</span><div class="alerta-texto"><div class="alerta-titulo" style="color:#166534">Sin alertas — Todas las faenas operando dentro de parámetros</div></div></div>';
}}

// Ranking (don't reorder, just tag)
const rankSorted = [...D.kpis].sort((a,b) => b.ci - a.ci);
const rankMap = {{}};
rankSorted.forEach((k,i) => {{ rankMap[k.t] = i + 1; }});
const medals = {{1:'\\u{{1F947}}',2:'\\u{{1F948}}',3:'\\u{{1F949}}'}};
const totalTeams = D.kpis.length;

// Team Cards with double-click expand
const cardsEl = document.getElementById('teamCards');
const detailEl = document.getElementById('teamDetail');
const expectedPct = cfg.dd / cfg.dm * 100;

function showTeamDetail(teamName) {{
  const k = D.kpis.find(x => x.t === teamName);
  const i = D.kpis.indexOf(k);
  const color = COLORS[i % COLORS.length];
  // "m³/día requerido" = ritmo necesario para cerrar la meta del mes
  const planDia = Math.round((k.m - k.a) / Math.max(cfg.dr, 1));  // ritmo real necesario
  const planTeorico = Math.round(k.m / cfg.dt * 10) / 10;         // referencia teórica
  const avancePlan = Math.round(planTeorico * cfg.dd);
  const difPlan = Math.round(k.a - avancePlan);
  const days = Object.keys(D.grid).map(Number).sort((a,b) => a-b);

  const kpis = [
    {{ l:'Meta Mensual', v:fmt(k.m), c:NEUTRAL }},
    {{ l:'m³/día requerido', v:fmt(planDia), c:planDia>k.p?BAD:NEUTRAL, sub:'para cerrar meta' }},
    {{ l:'Total Acumulado', v:fmt(k.a), c:NEUTRAL, bold:true }},
    {{ l:'Prom. diario actual', v:fmt(k.p), c:k.p<planTeorico?BAD:NEUTRAL }},
    {{ l:'Dif. vs Plan', v:fmt(difPlan), c:difPlan<0?BAD:NEUTRAL }},
    {{ l:'Cumplimiento %', v:k.c.toFixed(1)+'%', c:k.c<expectedPct?BAD:NEUTRAL }},
    {{ l:'Proyección Mes', v:fmt(k.pr), c:k.pr<k.m?BAD:NEUTRAL }},
    {{ l:'% Proyectado', v:(k.pr/k.m*100).toFixed(1)+'%', c:k.pr<k.m?BAD:NEUTRAL }},
  ];

  let kpiHtml = kpis.map(kp => `
    <div class="kpi-card" style="border-left:3px solid ${{kp.c===BAD?BAD:color}}">
      <div class="kpi-card-label">${{kp.l}}</div>
      <div class="kpi-card-value" style="color:${{kp.c}};${{kp.bold?'font-weight:800':''}}">${{kp.v}}</div>
    </div>`).join('');

  const disp = (k.turno > 0 ? (1 - k.tm / k.turno) * 100 : 100).toFixed(1);

  // Plan día DINÁMICO: se recalcula al inicio de cada día
  //   plan_día_d = (meta - acumulado_hasta_d-1) / días_restantes_desde_d
  // Ejemplo: si al final del día 5 llevas 1000 m³ de meta 7000 y quedan 25 días → plan día 6 = 240 m³/día
  let gridRows = '';
  let acumR = 0;
  days.forEach((d, ri) => {{
    const acumPrev = acumR;  // acumulado al inicio del día (antes de este día)
    const diasRest = cfg.dm - d + 1;  // días restantes incluyendo este día
    const metaRest = k.m - acumPrev;
    const planDiaDinamico = diasRest > 0 ? Math.round(metaRest / diasRest) : 0;
    const v = D.grid[d] ? (D.grid[d][teamName] || 0) : 0;
    acumR += v;
    const dif = v - planDiaDinamico;
    const bg = ri%2===0 ? '#F8FAFC' : 'white';
    // Color del NÚMERO según volumen (celda blanca, igual que la grilla principal)
    const cellTxt = v===0?'#DC2626':v>=200?'#059669':v>=100?'#CA8A04':'#EA580C';
    gridRows += `<tr style="background:${{bg}}">
      <td style="padding:5px 10px;font-weight:700;color:#1A5276;border-right:1px solid #E2E8F0">${{d}}</td>
      <td style="padding:5px 10px;text-align:right;font-weight:700;font-size:13px;color:${{cellTxt}};border-right:1px solid #E2E8F0">${{v>0?fmt(v):'—'}}</td>
      <td style="padding:5px 10px;text-align:right;color:#94A3B8;border-right:1px solid #E2E8F0">${{fmt(planDiaDinamico)}}</td>
      <td style="padding:5px 10px;text-align:right;font-weight:600;color:${{v===0?'#94A3B8':dif<0?BAD:NEUTRAL}};border-right:1px solid #E2E8F0">${{v>0?fmt(Math.round(dif)):'—'}}</td>
      <td style="padding:5px 10px;text-align:right;font-weight:700;color:#1A5276;box-shadow:inset 3px 0 0 #1A5276">${{fmt(Math.round(acumR))}}</td>
    </tr>`;
  }});

  detailEl.innerHTML = `
    <div class="team-detail" style="border:2px solid ${{color}};box-shadow:0 8px 24px ${{color}}18">
      <div class="detail-header">
        <div style="display:flex;align-items:center;gap:16px">
          <div>
            <h2 style="margin:0;font-size:20px;font-weight:800;color:${{color}}">${{teamName}}</h2>
            <p style="margin:2px 0 0;font-size:12px;color:#94A3B8">${{k.e}} | ${{k.pr2}} | ${{k.d}} días con datos</p>
          </div>
        </div>
        <button class="close-btn" onclick="detailEl.innerHTML=''">✕ Cerrar</button>
      </div>
      <div class="kpi-cards">${{kpiHtml}}</div>
      <div class="detail-grid">
        <div>
          <h4 style="margin:0 0 8px;font-size:13px;color:#1A5276">Producción Diaria (m³)</h4>
          <canvas id="chartTeamDetail" height="180"></canvas>
        </div>
        <div>
          <h4 style="margin:0 0 8px;font-size:13px;color:#1A5276">Tiempos Perdidos</h4>
          <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:12px">
            <div class="tm-box" style="background:#FEF2F2"><div style="font-size:9px;color:#94A3B8">TM Mantenc.</div><div style="font-size:16px;font-weight:700;color:#C0392B">${{(k.tm/60).toFixed(1)}}h</div></div>
            <div class="tm-box" style="background:#FFF7ED"><div style="font-size:9px;color:#94A3B8">TM Total</div><div style="font-size:16px;font-weight:700;color:#E67E22">${{(k.tt/60).toFixed(1)}}h</div></div>
            <div class="tm-box" style="background:#F0FDF4"><div style="font-size:9px;color:#94A3B8">Hrs Efectivas</div><div style="font-size:16px;font-weight:700;color:#1E8449">${{k.h}}h</div></div>
          </div>
          <div style="font-size:11px;color:#64748B">
            <div style="display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid #F1F5F9">
              <span>Disponibilidad estimada</span><b style="color:${{parseFloat(disp)<80?BAD:NEUTRAL}}">${{disp}}%</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid #F1F5F9">
              <span>m³/hora efectiva</span><b>${{k.r}}</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0">
              <span>Promedio diario</span><b>${{fmt(k.p)}} m³</b>
            </div>
          </div>
        </div>
      </div>
      <div style="margin-top:16px">
        <h4 style="margin:0 0 8px;font-size:13px;color:#1A5276">Análisis 80/20 — Pareto de Tiempos Perdidos</h4>
        ${{(() => {{
          const causas = D.tmTeamCauses[teamName] || [];
          if (causas.length === 0) return '<div style="font-size:11px;color:#94A3B8;padding:8px">Sin tiempos perdidos registrados.</div>';
          const total = causas.reduce((s, c) => s + c.m, 0);
          let acum = 0; let causasVital = []; let causasTrivial = [];
          causas.forEach(c => {{ acum += c.m; const pctAcum = acum/total*100;
            (pctAcum <= 80 || causasVital.length === 0) ? causasVital.push({{...c, pctAcum}}) : causasTrivial.push(c);
          }});
          let h = `<div style="background:#FEF3C7;border-left:3px solid #D97706;padding:8px 12px;border-radius:6px;margin-bottom:8px;font-size:11px">
            <strong style="color:#92400E">Hallazgo 80/20:</strong> <span style="color:#78350F">${{causasVital.length}} causa${{causasVital.length!==1?'s':''}} concentran el ${{causasVital[causasVital.length-1].pctAcum.toFixed(0)}}% del tiempo perdido total (${{(total/60).toFixed(1)}} hrs).</span>
          </div>`;
          h += '<table style="width:100%;border-collapse:collapse;font-size:11px">';
          h += '<thead><tr style="background:#F1F5F9"><th style="padding:4px 8px;text-align:left">Causa</th><th style="padding:4px 8px;text-align:right">Tiempo</th><th style="padding:4px 8px;text-align:right">%</th><th style="padding:4px 8px;text-align:right">% Acum</th></tr></thead><tbody>';
          causasVital.forEach(c => {{
            const pct = (c.m/total*100).toFixed(1);
            h += `<tr style="background:#FEF9C3">
              <td style="padding:4px 8px;font-weight:600">⚡ ${{c.n}}</td>
              <td style="padding:4px 8px;text-align:right">${{(c.m/60).toFixed(1)}}h</td>
              <td style="padding:4px 8px;text-align:right">${{pct}}%</td>
              <td style="padding:4px 8px;text-align:right;font-weight:700;color:#D97706">${{c.pctAcum.toFixed(0)}}%</td>
            </tr>`;
          }});
          causasTrivial.forEach(c => {{
            const pct = (c.m/total*100).toFixed(1);
            h += `<tr style="background:white">
              <td style="padding:4px 8px;color:#64748B">${{c.n}}</td>
              <td style="padding:4px 8px;text-align:right;color:#94A3B8">${{(c.m/60).toFixed(1)}}h</td>
              <td style="padding:4px 8px;text-align:right;color:#94A3B8">${{pct}}%</td>
              <td style="padding:4px 8px;text-align:right;color:#94A3B8">—</td>
            </tr>`;
          }});
          h += '</tbody></table>';
          h += `<div style="font-size:10px;color:#64748B;margin-top:6px">💡 <strong>Atacar las causas en amarillo</strong> reduciría el ${{causasVital[causasVital.length-1].pctAcum.toFixed(0)}}% de TM.</div>`;
          return h;
        }})()}}
      </div>
      <div style="margin-top:16px;overflow-x:auto">
        <h4 style="margin:0 0 8px;font-size:13px;color:#1A5276">Tabla de Producciones</h4>
        <table>
          <thead><tr style="background:${{color}};color:white">
            <th style="padding:6px 10px;text-align:left;border-radius:6px 0 0 0">Día</th>
            <th style="padding:6px 10px;text-align:right">m³ SSC</th>
            <th style="padding:6px 10px;text-align:right">Plan día</th>
            <th style="padding:6px 10px;text-align:right">Dif.</th>
            <th style="padding:6px 10px;text-align:right;border-radius:0 6px 0 0">Acum.</th>
          </tr></thead>
          <tbody>${{gridRows}}</tbody>
        </table>
      </div>
    </div>`;

  // Render team chart con línea de proyección
  const ctx = document.getElementById('chartTeamDetail').getContext('2d');
  // Línea horizontal del Plan día (teórico)
  const planLine = days.map(() => planTeorico);
  // Línea DINÁMICA del m³/día requerido: se recalcula cada día en base a lo que faltaba ese día
  let _acum = 0;
  const reqLine = days.map(d => {{
    const diasRest = cfg.dm - d + 1;
    const metaRest = k.m - _acum;
    const req = diasRest > 0 ? Math.max(0, Math.round(metaRest / diasRest)) : 0;
    _acum += (D.grid[d]?.[teamName] || 0);
    return req;
  }});
  // Línea de tendencia: promedio móvil últimos 3 días proyectado linealmente
  const realData = days.map(d => D.grid[d]?.[teamName] || 0);
  const tendData = [];
  for (let i = 0; i < realData.length; i++) {{
    if (realData[i] > 0) {{
      const ini = Math.max(0, i - 2);
      const ventana = realData.slice(ini, i + 1).filter(v => v > 0);
      tendData.push(ventana.length > 0 ? ventana.reduce((a, b) => a + b, 0) / ventana.length : null);
    }} else tendData.push(null);
  }}
  new Chart(ctx, {{
    type: 'bar',
    data: {{
      labels: days,
      datasets: [
        {{ label: 'm³ real', data: realData, backgroundColor: color + 'CC', borderRadius: 3, order: 3 }},
        {{ label: 'Tendencia (prom 3d)', data: tendData, type: 'line', borderColor: '#059669', borderWidth: 2,
          borderDash: [0], fill: false, pointRadius: 0, tension: 0.3, order: 1 }},
        {{ label: 'Plan teórico (' + fmt(planTeorico) + ')', data: planLine, type: 'line',
          borderColor: '#94A3B8', borderWidth: 1.5, borderDash: [5, 5], fill: false, pointRadius: 0, order: 2 }},
        {{ label: 'Req. p/ cerrar (' + fmt(planDia) + ')', data: reqLine, type: 'line',
          borderColor: '#DC2626', borderWidth: 2, borderDash: [2, 3], fill: false, pointRadius: 0, order: 2 }},
      ]
    }},
    options: {{
      responsive: true,
      plugins: {{ legend: {{ display: true, position: 'bottom', labels: {{ font: {{ size: 10 }}, boxWidth: 12 }} }} }},
      scales: {{ y: {{ beginAtZero: true }} }}
    }}
  }});

  detailEl.scrollIntoView({{ behavior: 'smooth' }});
}}

// Build all cards HTML first, then set innerHTML once
let cardsHtml = '';
D.kpis.forEach((k, i) => {{
  const color = COLORS[i % COLORS.length];
  const pct = Math.min(k.ci/100, 1);
  const gaugeColor = pct >= 0.9 ? '#1E8449' : pct >= 0.7 ? '#E67E22' : '#C0392B';
  const cumplColor = k.c < expectedPct ? BAD : NEUTRAL;
  const brechaColor = k.b < 0 ? BAD : NEUTRAL;
  const rank = rankMap[k.t];
  const isBottom3 = rank > totalTeams - 3;
  const badgeHtml = rank <= 3 ? `<div class="rank-badge">${{medals[rank]}}</div>` : `<div class="rank-badge-mid">#${{rank}}</div>`;
  const disp = k.turno > 0 ? (1 - k.tm / k.turno) * 100 : 100;
  const dispColor = disp >= 80 ? '#166534' : disp >= 60 ? '#92400E' : '#C0392B';
  const dispBg = disp >= 80 ? '#F0FDF4' : disp >= 60 ? '#FFF7ED' : '#FEF2F2';
  cardsHtml += `
    <div class="team-card" data-team="${{k.t}}" style="border-color:${{color}}22">
      ${{badgeHtml}}
      <div class="gauge-container">
        <div>
          <div class="team-name" style="color:${{color}}">${{k.t}}</div>
          <div style="font-size:10px;color:#94A3B8">${{k.e}} | ${{k.pr2}}</div>
        </div>
        <svg width="64" height="40" viewBox="0 0 64 40">
          <path d="M 6 36 A 26 26 0 0 1 58 36" fill="none" stroke="#E5E7EB" stroke-width="6" stroke-linecap="round"/>
          <path d="M 6 36 A 26 26 0 0 1 58 36" fill="none" stroke="${{gaugeColor}}" stroke-width="6" stroke-linecap="round"
            stroke-dasharray="${{pct * 82}} 82"/>
          <text x="32" y="30" text-anchor="middle" font-size="12" font-weight="700" fill="${{gaugeColor}}">${{k.ci.toFixed(1)}}%</text>
        </svg>
      </div>
      <div class="kpi-grid">
        <div><span class="label">Acumulado</span><br><span class="value">${{fmt(k.a)}}</span></div>
        <div><span class="label">Meta</span><br><span class="value">${{fmt(k.m)}}</span></div>
        <div><span class="label">Brecha</span><br><span class="value" style="color:${{brechaColor}}">${{fmt(k.b)}}</span></div>
        <div><span class="label">Proy.</span><br><span class="value">${{fmt(k.pr)}}</span></div>
        <div><span class="label">m³/hr</span><br><span class="value">${{k.r}}</span></div>
        <div><span class="label">Ritmo Cierre</span><br><span class="value" style="color:${{(() => {{ const ritmo=Math.round((k.m-k.a)/Math.max(cfg.dr,1)); return ritmo>k.p?BAD:NEUTRAL; }})()}}">${{(() => {{ const ritmo=Math.round((k.m-k.a)/Math.max(cfg.dr,1)); return fmt(ritmo); }})()}} <span style="font-size:9px">m³/d</span></span></div>
      </div>
      <div class="disp-bar" style="background:${{dispBg}}">
        <span style="font-size:10px;color:#64748B">Disp. Mecánica</span>
        <span style="display:flex;align-items:center;gap:6px">
          <span class="disp-track"><span class="disp-fill" style="width:${{Math.min(disp,100)}}%;background:${{dispColor}}"></span></span>
          <b style="font-size:11px;color:${{dispColor}}">${{disp.toFixed(1)}}%</b>
        </span>
      </div>
      <div class="dbl-hint">doble click para detalle</div>
    </div>`;
}});

// ── TOTAL Card ──
(() => {{
  const totalBrecha = Math.round(proyTotal - cfg.tm);
  const totalCiPct = Math.min(proyTotal / cfg.tm * 100, 999);
  const avgRate = (D.kpis.reduce((s,k) => s + k.r, 0) / D.kpis.length).toFixed(1);
  const ritmoTotal = Math.round((cfg.tm - cfg.ta) / Math.max(cfg.dr, 1));
  const avancePlan = Math.round(cfg.tm / cfg.dt * cfg.dd);
  const difVsPlan = Math.round(cfg.ta - avancePlan);
  const pctG = Math.min(totalCiPct / 100, 1);
  const gColor = pctG >= 0.9 ? '#1E8449' : pctG >= 0.7 ? '#E67E22' : '#C0392B';
  const brechaColor = totalBrecha < 0 ? '#FF6B6B' : '#7DCEA0';
  const difColor = difVsPlan < 0 ? '#FF6B6B' : '#7DCEA0';
  const ritmoColor = ritmoTotal > promDia ? '#FF6B6B' : '#7DCEA0';
  cardsHtml += `
    <div class="team-card" style="border-color:#1A527644;background:linear-gradient(135deg,#1A5276 0%,#154360 100%);color:white;max-width:340px;box-shadow:0 4px 12px rgba(26,82,118,0.25)">
      <div class="gauge-container">
        <div>
          <div class="team-name" style="color:white;font-size:14px">TOTAL</div>
          <div style="font-size:10px;opacity:0.7">${{D.kpis.length}} equipos | ${{cfg.mesNombre}} ${{cfg.anio}}</div>
        </div>
        <svg width="64" height="40" viewBox="0 0 64 40">
          <path d="M 6 36 A 26 26 0 0 1 58 36" fill="none" stroke="rgba(255,255,255,0.2)" stroke-width="6" stroke-linecap="round"/>
          <path d="M 6 36 A 26 26 0 0 1 58 36" fill="none" stroke="${{gColor}}" stroke-width="6" stroke-linecap="round"
            stroke-dasharray="${{pctG * 82}} 82"/>
          <text x="32" y="30" text-anchor="middle" font-size="12" font-weight="700" fill="${{gColor}}">${{totalCiPct.toFixed(1)}}%</text>
        </svg>
      </div>
      <div class="kpi-grid" style="color:white;grid-template-columns:1fr 1fr 1fr 1fr">
        <div><span class="label" style="color:rgba(255,255,255,0.6)">Acumulado</span><br><span class="value" style="color:white;font-size:14px">${{fmt(cfg.ta)}}</span></div>
        <div><span class="label" style="color:rgba(255,255,255,0.6)">Meta</span><br><span class="value" style="color:white">${{fmt(cfg.tm)}}</span></div>
        <div><span class="label" style="color:rgba(255,255,255,0.6)">Dif. vs Plan</span><br><span class="value" style="color:${{difColor}}">${{fmt(difVsPlan)}}</span></div>
        <div><span class="label" style="color:rgba(255,255,255,0.6)">Brecha Proy.</span><br><span class="value" style="color:${{brechaColor}}">${{fmt(totalBrecha)}}</span></div>
        <div><span class="label" style="color:rgba(255,255,255,0.6)">Proy.</span><br><span class="value" style="color:white">${{fmt(Math.round(proyTotal))}}</span></div>
        <div><span class="label" style="color:rgba(255,255,255,0.6)">m³/hr</span><br><span class="value" style="color:white">${{avgRate}}</span></div>
        <div><span class="label" style="color:rgba(255,255,255,0.6)">Ritmo Cierre</span><br><span class="value" style="color:${{ritmoColor}}">${{fmt(ritmoTotal)}} <span style="font-size:9px">m³/d</span></span></div>
      </div>
    </div>`;
}})();
cardsEl.innerHTML = cardsHtml;

// Attach dblclick listeners after DOM is ready
cardsEl.querySelectorAll('.team-card').forEach(card => {{
  card.addEventListener('dblclick', () => {{
    const teamName = card.getAttribute('data-team');
    showTeamDetail(teamName);
  }});
}});

// Charts — Producción Diaria con Meta Diaria Móvil
const trendCtx = document.getElementById('chartTrend').getContext('2d');
// Meta diaria móvil: (Meta total - Acumulado hasta ayer) / Días restantes
const metaDiariaMovil = [];
let acumHastaAyer = 0;
D.trend.forEach((t, i) => {{
  const diasRestantes = cfg.dm - t.d + 1;
  const metaRestante = cfg.tm - acumHastaAyer;
  metaDiariaMovil.push(Math.round(Math.max(metaRestante / diasRestantes, 0)));
  acumHastaAyer += t.v;
}});
// Color de barras: verde si supera meta móvil, rojo si no
const barColors = D.trend.map((t, i) => t.v >= metaDiariaMovil[i] ? '#27AE60CC' : '#E74C3CCC');
new Chart(trendCtx, {{
  type: 'bar',
  data: {{
    labels: D.trend.map(t => 'Día ' + t.d),
    datasets: [
      {{ label: 'Producción Real', data: D.trend.map(t => t.v),
        backgroundColor: barColors, borderRadius: 4, maxBarThickness: 48, order: 2 }},
      {{ label: 'Meta Diaria Requerida', data: metaDiariaMovil,
        type: 'line', borderColor: '#E67E22', borderWidth: 2.5,
        borderDash: [6,3], pointRadius: 3, pointBackgroundColor: '#E67E22',
        fill: false, tension: 0.3, order: 1 }}
    ]
  }},
  options: {{
    responsive: true,
    plugins: {{
      legend: {{ position: 'top', labels: {{ font: {{ size: 11 }} }} }},
      tooltip: {{ callbacks: {{
        label: ctx => {{
          if (ctx.datasetIndex === 0) return 'Real: ' + fmt(ctx.raw) + ' m³';
          return 'Meta requerida: ' + fmt(ctx.raw) + ' m³/día';
        }}
      }} }}
    }},
    scales: {{ y: {{ beginAtZero: true, title: {{ display: true, text: 'm³ SSC' }} }} }}
  }}
}});

const teamsCtx = document.getElementById('chartTeams').getContext('2d');
new Chart(teamsCtx, {{
  type: 'bar',
  data: {{
    labels: D.kpis.map(k => k.t.replace('Millalemu ','M')),
    datasets: [
      {{ label: 'Acumulado', data: D.kpis.map(k => k.a), backgroundColor: '#1A5276CC', borderRadius: 3 }},
      {{ label: 'Proyección', data: D.kpis.map(k => k.pr), backgroundColor: '#1E844988', borderRadius: 3 }},
      {{ label: 'Meta', data: D.kpis.map(k => k.m), backgroundColor: '#C0392B44', borderRadius: 3 }}
    ]
  }},
  options: {{ responsive: true, scales: {{ y: {{ beginAtZero: true }} }} }}
}});

// Grid Table with KPI rows
const gridEl = document.getElementById('gridTable');
const days = Object.keys(D.grid).map(Number).sort((a,b) => a-b);
const DIAS_SEMANA = ['Dom','Lun','Mar','Mié','Jue','Vie','Sáb'];
const getDow = d => DIAS_SEMANA[new Date(cfg.anio, cfg.mes-1, d).getDay()];
const isWeekend = d => {{ const dow = new Date(cfg.anio, cfg.mes-1, d).getDay(); return dow===0||dow===6; }};

const totalHrs = D.kpis.reduce((s,k) => s+k.h, 0);
const totalTMmant = D.kpis.reduce((s,k) => s+k.tm, 0);
const totalTurno = D.kpis.reduce((s,k) => s+k.turno, 0);
const avgDispTotal = totalTurno > 0 ? (1 - totalTMmant/totalTurno)*100 : 100;

let thead = '<thead><tr><th style="text-align:left">KPI / Día</th>';
TEAMS.forEach(t => thead += `<th>${{t.replace('Millalemu ','M')}}</th>`);
thead += '<th style="border-left:3px solid #FFFFFF;font-weight:800">Total</th></tr></thead>';

const pctColor = pct => pct >= 80 ? '#059669' : pct >= 60 ? '#D97706' : BAD;
const kpiDefs = [
  // % Proy al inicio — lo más accionable
  {{ label: '% Proyección', fn: k => {{ const p=(k.pr/k.m*100); return p.toFixed(0)+'%'; }},
    color: k => pctColor(k.pr/k.m*100), total: cumplTotal+'%', bg: '#F1F5F9', bold: true, big: true,
    totalColor: pctColor(parseFloat(cumplTotal)) }},
  {{ label: 'Meta Mensual', fn: k => fmt(k.m), total: fmt(cfg.tm), bg: '#F8FAFC', bold: true }},
  {{ label: 'Total Acumulado', fn: k => fmt(k.a), total: fmt(cfg.ta), bg: '#F1F5F9', bold: true, big: true,
    color: k => {{ const ap=k.m/cfg.dt*cfg.dd; return ap>0 && k.a/ap>=1.0 ? '#059669' : ap>0 && k.a/ap>=0.85 ? '#D97706' : BAD; }},
    totalColor: (() => {{ const ap=cfg.tm/cfg.dt*cfg.dd; return ap>0 && cfg.ta/ap>=1.0 ? '#059669' : ap>0 && cfg.ta/ap>=0.85 ? '#D97706' : BAD; }})() }},
  {{ label: 'Avance Plan', fn: k => fmt(Math.round(k.m/cfg.dt*cfg.dd)), total: fmt(Math.round(cfg.tm/cfg.dt*cfg.dd)), bg: '#F8FAFC' }},
  {{ label: 'Dif. vs Plan', fn: k => {{ const d=Math.round(k.a-(k.m/cfg.dt*cfg.dd)); return fmt(d); }},
    color: k => (k.a-(k.m/cfg.dt*cfg.dd))<0 ? BAD : NEUTRAL, total: fmt(Math.round(cfg.ta-cfg.tm/cfg.dt*cfg.dd)), bg: '#F8FAFC' }},
  {{ label: 'Ritmo Cierre', fn: k => {{ const needed=(k.m-k.a)/Math.max(cfg.dr,1); return fmt(Math.round(needed)); }},
    color: k => {{ const needed=(k.m-k.a)/Math.max(cfg.dr,1); return needed>k.p ? BAD : NEUTRAL; }},
    total: fmt(Math.round((cfg.tm-cfg.ta)/Math.max(cfg.dr,1))), bg: '#F8FAFC' }},
  {{ label: 'm³/hr Efectiva', fn: k => k.r.toFixed(1), total: totalHrs>0?(cfg.ta/totalHrs).toFixed(1):'—', bg: '#EFF6FF', bold: true }},
  {{ label: 'TM Mant. (hrs)', fn: k => (k.tm/60).toFixed(1),
    color: k => k.tm>2000 ? BAD : NEUTRAL, total: (totalTMmant/60).toFixed(1), bg: '#FEF2F2' }},
  // Proyección (m³) al final — destacado
  {{ label: 'Proyección Mes', fn: k => fmt(k.pr),
    color: k => pctColor(k.pr/k.m*100), total: fmt(Math.round(proyTotal)), bg: '#F1F5F9', bold: true, big: true,
    totalColor: pctColor(parseFloat(cumplTotal)) }},
];

let tbody = '<tbody>';
kpiDefs.forEach((kpi, ki) => {{
  const borderBot = ki === kpiDefs.length-1 ? 'border-bottom:3px solid #1A5276;' : '';
  const fs = kpi.big ? '14px' : '11px';
  const fsLabel = kpi.big ? '12px' : '11px';
  let row = `<tr style="background:${{kpi.bg}}"><td style="padding:${{kpi.big?'7px 12px':'5px 12px'}};font-weight:700;font-size:${{fsLabel}};color:#1A5276;${{borderBot}}">${{kpi.label}}</td>`;
  TEAMS.forEach(t => {{
    const k = D.kpis.find(x => x.t === t);
    const c = kpi.color ? `color:${{kpi.color(k)}};` : `color:${{NEUTRAL}};`;
    const fw = (kpi.bold || kpi.big) ? 'font-weight:700;' : 'font-weight:500;';
    row += `<td style="padding:${{kpi.big?'7px 10px':'5px 10px'}};text-align:right;font-size:${{fs}};${{fw}}${{c}}border-right:1px solid #E2E8F0;${{borderBot ? borderBot : 'border-bottom:1px solid #E2E8F0;'}}">${{kpi.fn(k)}}</td>`;
  }});
  const totalC = kpi.totalColor ? `color:${{kpi.totalColor}};` : 'color:#1A5276;';
  row += `<td style="padding:${{kpi.big?'7px 10px':'5px 10px'}};text-align:right;font-weight:800;font-size:${{fs}};border-left:3px solid #1A5276;${{totalC}}${{borderBot}}">${{kpi.total}}</td></tr>`;
  tbody += row;
}});

// Daily data rows: zebra simple, color SOLO en el número (no en celda)
const planDiaFaena = cfg.tm / cfg.dt;  // m³/día total que la faena debe hacer para cumplir meta
days.forEach((d, ri) => {{
  let total = 0;
  const dow = getDow(d);
  const bgRow = ri%2===0 ? 'var(--subtle)' : 'white';
  let row = `<tr style="background:${{bgRow}}"><td style="font-weight:700;font-size:13px;border-right:1px solid #E2E8F0;white-space:nowrap;color:#1A5276">${{d}} <span style="font-size:10px;font-weight:400;color:#94A3B8">${{dow}}</span></td>`;
  TEAMS.forEach(t => {{
    const v = D.grid[d] ? (D.grid[d][t] || 0) : 0;
    const tmVal = D.gridTM && D.gridTM[d] ? (D.gridTM[d][t] || 0) : 0;
    total += v;
    // Color del NÚMERO según nivel (celda queda blanca)
    const txtColor = v === 0 ? '#DC2626' : v >= 200 ? '#059669' : v >= 100 ? '#CA8A04' : '#EA580C';
    const tmHint = tmVal > 0 ? `<div style="font-size:8px;color:#C0392B;font-weight:400;margin-top:1px">${{(tmVal/60).toFixed(1)}}h TM</div>` : '';
    row += `<td style="font-weight:700;font-size:13px;border-right:1px solid #E2E8F0;color:${{txtColor}}">${{v > 0 ? fmt(v) : '—'}}${{tmHint}}</td>`;
  }});
  // Color del Total del día: comparar vs plan/día faena (~1750 m³/día con meta 52500/30)
  const pctPlan = planDiaFaena > 0 ? total/planDiaFaena : 0;
  const totalColor = total === 0 ? '#DC2626' : pctPlan >= 1.0 ? '#059669' : pctPlan >= 0.80 ? '#CA8A04' : '#EA580C';
  row += `<td style="font-weight:800;font-size:13px;background:#F1F5F9;border-left:3px solid #1A5276;color:${{totalColor}}">${{fmt(total)}}</td></tr>`;
  tbody += row;
}});

// Fila TOTAL removida — ya está el "Total Acumulado" en el bloque KPI superior

// Promedio row — línea superior continua + color por nivel vs plan/día equipo
tbody += '<tr style="font-weight:700"><td style="padding:8px 12px;font-size:11px;color:#1A5276;border-top:3px solid #1A5276">PROMEDIO/DÍA</td>';
TEAMS.forEach(t => {{
  const k = D.kpis.find(x => x.t === t);
  const avg = k && k.d > 0 ? k.a / k.d : 0;
  // Plan/día por equipo = meta_equipo / días_a_trabajar
  const planEq = k ? k.m / cfg.dt : 0;
  const pct = planEq > 0 ? avg/planEq : 0;
  const avgColor = avg === 0 ? '#DC2626' : pct >= 1.0 ? '#059669' : pct >= 0.80 ? '#CA8A04' : '#EA580C';
  tbody += `<td style="padding:8px 10px;text-align:right;font-size:12px;color:${{avgColor}};border-top:3px solid #1A5276">${{fmt(Math.round(avg))}}</td>`;
}});
// Promedio Total faena vs plan/día faena
const promFaena = cfg.dd > 0 ? cfg.ta/cfg.dd : 0;
const pctPF = planDiaFaena > 0 ? promFaena/planDiaFaena : 0;
const promColor = promFaena === 0 ? '#DC2626' : pctPF >= 1.0 ? '#059669' : pctPF >= 0.80 ? '#CA8A04' : '#EA580C';
tbody += `<td style="padding:8px 10px;text-align:right;font-size:12px;color:${{promColor}};font-weight:800;background:#F1F5F9;border-top:3px solid #1A5276;border-left:3px solid #1A5276">${{fmt(Math.round(promFaena))}}</td></tr></tbody>`;
gridEl.innerHTML = thead + tbody;

// TM Summary Cards
const tmCat = D.tmCat || [];
const tmTrend = D.tmTrend || [];
const tmTeamCat = D.tmTeamCat || [];
const totalTM = tmCat.reduce((s,c) => s + c.v, 0);
const CAT_COLORS_MAP = {{'Mantención':'#C0392B','Operacional':'#E67E22','Proceso':'#3498DB'}};
const tmSummaryEl = document.getElementById('tmSummaryCards');
tmSummaryEl.innerHTML = `
  <div style="background:white;border-radius:10px;padding:12px 16px;box-shadow:0 1px 3px rgba(0,0,0,0.08)">
    <div style="font-size:10px;color:#94A3B8">TM Total Flota</div>
    <div style="font-size:22px;font-weight:800;color:#1A5276">${{fmt(totalTM)}} <span style="font-size:11px;font-weight:400">min</span></div>
    <div style="font-size:11px;color:#64748B">${{(totalTM/60).toFixed(1)}} horas</div>
  </div>
  ${{tmCat.map(c => `
  <div style="background:white;border-radius:10px;padding:12px 16px;box-shadow:0 1px 3px rgba(0,0,0,0.08);border-left:3px solid ${{CAT_COLORS_MAP[c.n]||'#94A3B8'}}">
    <div style="font-size:10px;color:#94A3B8">${{c.n}}</div>
    <div style="font-size:22px;font-weight:800;color:${{CAT_COLORS_MAP[c.n]||'#334155'}}">${{fmt(c.v)}} <span style="font-size:11px;font-weight:400">min</span></div>
    <div style="font-size:11px;color:#64748B">${{totalTM > 0 ? (c.v/totalTM*100).toFixed(1) : 0}}% del total</div>
  </div>`).join('')}}
`;

// TM Chart (Pareto)
const tmCtx = document.getElementById('chartTM').getContext('2d');
new Chart(tmCtx, {{
  type: 'bar',
  data: {{
    labels: D.tmTop.map(t => t.n.length > 30 ? t.n.substring(0,30)+'...' : t.n),
    datasets: [{{ label: 'Horas', data: D.tmTop.map(t => Math.round(t.m/60*10)/10),
      backgroundColor: D.tmTop.map((_,i) => i < 2 ? '#C0392BCC' : i < 5 ? '#E67E22CC' : '#94A3B8CC'),
      borderRadius: 4 }}]
  }},
  options: {{ indexAxis: 'y', responsive: true, plugins: {{ legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: c => c.raw + ' h' }} }} }},
    scales: {{ x: {{ beginAtZero: true, title: {{ display: true, text: 'horas' }} }} }} }}
}});

// TM Category Donut
const tmCatCtx = document.getElementById('chartTMCat').getContext('2d');
new Chart(tmCatCtx, {{
  type: 'doughnut',
  data: {{
    labels: tmCat.map(c => c.n),
    datasets: [{{ data: tmCat.map(c => Math.round(c.v/60*10)/10), backgroundColor: ['#C0392B','#E67E22','#3498DB'] }}]
  }},
  options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }} }} }},
      tooltip: {{ callbacks: {{ label: c => c.label + ': ' + c.raw + ' h' }} }} }} }}
}});

// Species Pie
const espCtx = document.getElementById('chartEsp').getContext('2d');
new Chart(espCtx, {{
  type: 'doughnut',
  data: {{
    labels: D.esp.map(e => e.n),
    datasets: [{{ data: D.esp.map(e => e.v), backgroundColor: ['#27AE60','#2980B9','#E67E22','#8E44AD'] }}]
  }},
  options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }} }} }} }} }}
}});

// TM Daily Trend (stacked)
const tmTrendCtx = document.getElementById('chartTMTrend').getContext('2d');
new Chart(tmTrendCtx, {{
  type: 'bar',
  data: {{
    labels: tmTrend.map(t => t.d),
    datasets: [
      {{ label: 'Mantención', data: tmTrend.map(t => Math.round(t.mant/60*10)/10), backgroundColor: '#C0392BCC', stack: 'a' }},
      {{ label: 'Operacional', data: tmTrend.map(t => Math.round(t.oper/60*10)/10), backgroundColor: '#E67E22CC', stack: 'a' }},
      {{ label: 'Proceso', data: tmTrend.map(t => Math.round(t.proc/60*10)/10), backgroundColor: '#3498DBCC', stack: 'a', borderRadius: 3 }}
    ]
  }},
  options: {{ responsive: true, plugins: {{ legend: {{ position: 'top', labels: {{ font: {{ size: 10 }} }} }},
      tooltip: {{ callbacks: {{ label: c => c.dataset.label + ': ' + c.raw + ' h' }} }} }},
    scales: {{ x: {{ stacked: true }}, y: {{ stacked: true, beginAtZero: true, title: {{ display: true, text: 'horas' }} }} }} }}
}});

// TM by Team (stacked horizontal)
const tmTeamCtx = document.getElementById('chartTMTeam').getContext('2d');
new Chart(tmTeamCtx, {{
  type: 'bar',
  data: {{
    labels: tmTeamCat.map(t => t.t),
    datasets: [
      {{ label: 'Mantención', data: tmTeamCat.map(t => Math.round(t.mant/60*10)/10), backgroundColor: '#C0392BCC', stack: 'a' }},
      {{ label: 'Operacional', data: tmTeamCat.map(t => Math.round(t.oper/60*10)/10), backgroundColor: '#E67E22CC', stack: 'a' }},
      {{ label: 'Proceso', data: tmTeamCat.map(t => Math.round(t.proc/60*10)/10), backgroundColor: '#3498DBCC', stack: 'a', borderRadius: 3 }}
    ]
  }},
  options: {{ indexAxis: 'y', responsive: true, plugins: {{ legend: {{ position: 'top', labels: {{ font: {{ size: 10 }} }} }},
      tooltip: {{ callbacks: {{ label: c => c.dataset.label + ': ' + c.raw + ' h' }} }} }},
    scales: {{ x: {{ stacked: true, beginAtZero: true, title: {{ display: true, text: 'horas' }} }}, y: {{ stacked: true }} }} }}
}});

// Availability Table
const dispTableEl = document.getElementById('tmDispTable');
let dispHtml = `<table style="width:100%;border-collapse:collapse;font-size:12px">
  <thead><tr style="background:#1A5276;color:white">
    <th style="padding:8px 10px;text-align:left;border-radius:8px 0 0 0">Equipo</th>
    <th style="padding:8px 10px;text-align:right">Hrs Efect.</th>
    <th style="padding:8px 10px;text-align:right">TM Mant.</th>
    <th style="padding:8px 10px;text-align:right">TM Oper.</th>
    <th style="padding:8px 10px;text-align:right">TM Proc.</th>
    <th style="padding:8px 10px;text-align:right">TM Total</th>
    <th style="padding:8px 10px;text-align:right;border-radius:0 8px 0 0">Disp. %</th>
  </tr></thead><tbody>`;
tmTeamCat.forEach((t, i) => {{
  const dispColor = t.disp >= 80 ? '#166534' : t.disp >= 60 ? '#92400E' : '#C0392B';
  const mantColor = t.mant > 1000 ? '#C0392B' : '#334155';
  dispHtml += `<tr style="background:${{i%2===0?'#F8FAFC':'white'}}">
    <td style="padding:6px 10px;font-weight:600;color:#1A5276">${{t.tf}}</td>
    <td style="padding:6px 10px;text-align:right">${{t.hrs}}h</td>
    <td style="padding:6px 10px;text-align:right;color:${{mantColor}}">${{(t.mant/60).toFixed(1)}}h</td>
    <td style="padding:6px 10px;text-align:right">${{(t.oper/60).toFixed(1)}}h</td>
    <td style="padding:6px 10px;text-align:right">${{(t.proc/60).toFixed(1)}}h</td>
    <td style="padding:6px 10px;text-align:right;font-weight:600">${{(t.total/60).toFixed(1)}}h</td>
    <td style="padding:6px 10px;text-align:right;font-weight:700;color:${{dispColor}}">${{t.disp}}%</td>
  </tr>`;
}});
dispHtml += '</tbody></table>';
dispTableEl.innerHTML = dispHtml;

// ══════════════════════════════════════════════════════════
// NUEVAS VISUALIZACIONES
// ══════════════════════════════════════════════════════════

// ── 1. Curva Avance Acumulado vs Plan ────────────────────
const avanceCtx = document.getElementById('chartAvance').getContext('2d');
const avLabels = D.avance.map(a => 'Día ' + a.d);
new Chart(avanceCtx, {{
  type: 'line',
  data: {{
    labels: avLabels,
    datasets: [
      {{ label: 'Real Acumulado', data: D.avance.map(a => a.real),
         borderColor: '#1A5276', backgroundColor: 'rgba(26,82,118,0.08)',
         fill: true, tension: 0.3, pointRadius: 3, borderWidth: 2 }},
      {{ label: 'Plan Lineal', data: D.avance.map(a => a.plan),
         borderColor: '#E67E22', borderDash: [6,3],
         fill: false, tension: 0, pointRadius: 0, borderWidth: 2 }}
    ]
  }},
  options: {{
    responsive: true,
    plugins: {{
      legend: {{ position: 'top', labels: {{ font: {{ size: 11 }} }} }},
      tooltip: {{ callbacks: {{ label: ctx => ctx.dataset.label + ': ' + fmt(ctx.raw) + ' m³' }} }}
    }},
    scales: {{
      y: {{ beginAtZero: true, title: {{ display: true, text: 'm³ SSC' }} }}
    }}
  }}
}});

// ── 2. Flechas de Tendencia en Tarjetas ──────────────────
document.querySelectorAll('.team-card').forEach(card => {{
  const teamName = card.dataset.team;
  const tend = D.tendencia[teamName];
  if (!tend) return;
  const arrow = tend.c > 5 ? '↑' : tend.c < -5 ? '↓' : '→';
  const aColor = tend.c > 5 ? '#059669' : tend.c < -5 ? '#DC2626' : '#64748B';
  const trendEl = document.createElement('div');
  trendEl.style.cssText = 'margin-top:6px;padding:4px 8px;border-radius:6px;background:#F8FAFC;display:flex;justify-content:space-between;align-items:center;font-size:10px;color:#64748B';
  trendEl.innerHTML = `<span>Tendencia semanal</span><span style="font-weight:700;font-size:13px;color:${{aColor}}">${{arrow}} ${{tend.c > 0 ? '+' : ''}}${{tend.c}}%</span>`;
  card.appendChild(trendEl);
}});

// ── 3. Top 5 Causas TM por Faena (tabla) ─────────────────
const tmCausesEl = document.getElementById('tmTeamCausesTable');
let tcHtml = '<table style="width:100%;border-collapse:collapse;font-size:11px">';
tcHtml += '<thead><tr style="background:#1A5276;color:white"><th style="padding:8px 10px;text-align:left;border-radius:8px 0 0 0">Faena</th>';
for (let i = 1; i <= 5; i++) tcHtml += `<th style="padding:8px 6px;text-align:left">#${{i}} Causa</th>`;
tcHtml += '</tr></thead><tbody>';
TEAMS.forEach((t, ti) => {{
  const causes = D.tmTeamCauses[t] || [];
  tcHtml += `<tr style="background:${{ti%2===0?'#F8FAFC':'white'}}">`;
  tcHtml += `<td style="padding:6px 10px;font-weight:600;color:#1A5276;white-space:nowrap">${{t.replace('Millalemu ','M')}}</td>`;
  for (let i = 0; i < 5; i++) {{
    if (causes[i]) {{
      const hrs = (causes[i].m/60).toFixed(1);
      const bgI = causes[i].m > 1200 ? '#FEF2F2' : causes[i].m > 600 ? '#FFF7ED' : '';
      tcHtml += `<td style="padding:6px;${{bgI?'background:'+bgI+';':''}}"><span style="font-weight:600">${{hrs}}h</span><br><span style="color:#64748B;font-size:10px">${{causes[i].n.substring(0,35)}}</span></td>`;
    }} else {{
      tcHtml += '<td style="padding:6px;color:#CBD5E1">—</td>';
    }}
  }}
  tcHtml += '</tr>';
}});
tcHtml += '</tbody></table>';
tmCausesEl.innerHTML = tcHtml;

// ── 4. Rendimiento por Especie ───────────────────────────
const rendEspCtx = document.getElementById('chartRendEsp').getContext('2d');
const espColors = {{'PIRA':'#2E86C1','EUGL':'#27AE60','EUNI':'#8E44AD','EUNG':'#E67E22'}};
new Chart(rendEspCtx, {{
  type: 'bar',
  data: {{
    labels: D.rendEsp.map(r => r.n),
    datasets: [{{
      label: 'm³/hr',
      data: D.rendEsp.map(r => r.rend),
      backgroundColor: D.rendEsp.map(r => espColors[r.n] || '#95A5A6'),
      borderRadius: 6
    }}]
  }},
  options: {{
    responsive: true,
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => ctx.raw.toFixed(1) + ' m³/hr (' + fmt(D.rendEsp[ctx.dataIndex].vol) + ' m³ total)' }} }}
    }},
    scales: {{ y: {{ beginAtZero: true, title: {{ display: true, text: 'm³/hr' }} }} }}
  }}
}});

// Rendimiento por especie Y faena (grouped bar)
const rendETCtx = document.getElementById('chartRendEspTeam').getContext('2d');
const especies = [...new Set(D.rendEspTeam.map(r => r.e))];
const rendETDatasets = especies.map(esp => ({{
  label: esp,
  data: TEAMS.map(t => {{
    const match = D.rendEspTeam.find(r => r.e === esp && r.t === t);
    return match ? match.rend : 0;
  }}),
  backgroundColor: espColors[esp] || '#95A5A6',
  borderRadius: 4
}}));
new Chart(rendETCtx, {{
  type: 'bar',
  data: {{
    labels: TEAMS.map(t => t.replace('Millalemu ','M')),
    datasets: rendETDatasets
  }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ position: 'top', labels: {{ font: {{ size: 10 }} }} }} }},
    scales: {{ y: {{ beginAtZero: true, title: {{ display: true, text: 'm³/hr' }} }} }}
  }}
}});

// ── 5. Días Sin Producción (heatmap + cumplimiento m³) ────
const dspEl = document.getElementById('diasSinProdTable');
(() => {{
  // Rango de días: 1 hasta el último día con datos
  const lastDayG = Math.max(...Object.keys(D.grid).map(Number));
  const days = Array.from({{length: lastDayG}}, (_, i) => i + 1);

  // Meta diaria por faena (meta mensual / días del mes)
  const metaDiaria = {{}};
  const metaMensual = {{}};
  TEAMS.forEach(t => {{
    const k = D.kpis.find(x => x.t === t);
    metaMensual[t] = k ? k.m : 0;
    metaDiaria[t] = k ? (k.m / cfg.dm) : 0;
  }});
  // Meta prorrateada al día transcurrido (lo que deberían llevar ahora)
  const metaProrrateo = t => metaDiaria[t] * days.length;

  // Escala semáforo: verde / amarillo / rojo / plomo (sin registro)
  const cellColor = (vol, meta) => {{
    if (vol <= 0) return {{bg:'#94A3B8', label:'sinReg'}};   // plomo — sin registro
    const r = meta > 0 ? vol / meta : 1;
    if (r >= 0.80) return {{bg:'#16A34A', label:'ok'}};       // verde — ≥80% meta
    if (r >= 0.50) return {{bg:'#EAB308', label:'med'}};      // amarillo — 50-80% meta
    return {{bg:'#DC2626', label:'bajo'}};                    // rojo — <50% meta
  }};

  // Alertas: huecos, rachas, bajos
  const huecos = [], bajos = [];
  TEAMS.forEach(t => {{
    days.forEach(d => {{
      const v = D.grid[d]?.[t] || 0;
      if (v <= 0) huecos.push({{t, d}});
      else if (metaDiaria[t] > 0 && v / metaDiaria[t] < 0.4) bajos.push({{t, d, v}});
    }});
  }});
  const rachas = [];
  TEAMS.forEach(t => {{
    let inicio = null, len = 0;
    days.forEach((d, i) => {{
      const v = D.grid[d]?.[t] || 0;
      if (v <= 0) {{ if (inicio === null) inicio = d; len++; }}
      else {{ if (len >= 2) rachas.push({{t, inicio, fin: days[i-1], len}}); inicio = null; len = 0; }}
    }});
    if (len >= 2) rachas.push({{t, inicio, fin: days[days.length-1], len}});
  }});

  // Totales acumulados por equipo (cruce con cumplimiento de m³)
  const totales = TEAMS.map(t => {{
    const vol = days.reduce((s, d) => s + (D.grid[d]?.[t] || 0), 0);
    const metaAcum = metaProrrateo(t);
    const pctMeta = metaAcum > 0 ? (vol / metaAcum * 100) : 0;
    const pctMensual = metaMensual[t] > 0 ? (vol / metaMensual[t] * 100) : 0;
    let reportados = 0;
    days.forEach(d => {{ if ((D.grid[d]?.[t] || 0) > 0) reportados++; }});
    const pctReporte = Math.round(reportados / days.length * 100);
    return {{ t, vol, metaAcum, pctMeta, pctMensual, reportados, pctReporte }};
  }});

  // Layout: heatmap arriba + panel alertas a la derecha
  let html = '<div style="display:grid;grid-template-columns:1fr auto;gap:16px;align-items:start">';

  // ——— HEATMAP + TABLA CUMPLIMIENTO ———
  html += '<div style="min-width:0">';
  html += `<div style="margin-bottom:10px;display:flex;gap:14px;font-size:11px;color:#64748B;flex-wrap:wrap;align-items:center">
    <span><span style="display:inline-block;width:12px;height:12px;background:#16A34A;border-radius:2px;vertical-align:middle;margin-right:4px"></span>≥80% meta</span>
    <span><span style="display:inline-block;width:12px;height:12px;background:#EAB308;border-radius:2px;vertical-align:middle;margin-right:4px"></span>50-80% meta</span>
    <span><span style="display:inline-block;width:12px;height:12px;background:#DC2626;border-radius:2px;vertical-align:middle;margin-right:4px"></span>&lt;50% meta</span>
    <span><span style="display:inline-block;width:12px;height:12px;background:#94A3B8;border-radius:2px;vertical-align:middle;margin-right:4px"></span>Sin registro</span>
  </div>`;

  html += '<div style="overflow-x:auto"><table style="border-collapse:separate;border-spacing:2px;font-size:11px"><thead><tr>';
  html += '<th style="padding:4px 8px;text-align:left;color:#64748B;font-weight:600;position:sticky;left:0;background:white">Faena</th>';
  days.forEach(d => {{
    const dow = DIAS_SEMANA[new Date(cfg.anio, cfg.mes-1, d).getDay()];
    html += `<th style="padding:2px;text-align:center;font-weight:500;color:#94A3B8;font-size:10px;min-width:24px">
      <div>${{d}}</div><div style="font-size:9px">${{dow[0]}}</div>
    </th>`;
  }});
  // Nuevas columnas: cumplimiento m³
  html += '<th style="padding:4px 10px;text-align:right;color:#64748B;font-weight:600;border-left:2px solid #E2E8F0">m³ Acum.</th>';
  html += `<th style="padding:4px 8px;text-align:right;color:#64748B;font-weight:600" title="Meta mensual prorrateada a los ${{days.length}} días transcurridos">Meta al día</th>`;
  html += '<th style="padding:4px 8px;text-align:center;color:#64748B;font-weight:600">Cumpl.</th>';
  html += '</tr></thead><tbody>';

  TEAMS.forEach((t, i) => {{
    const r = totales[i];
    const metaColor = r.pctMeta >= 100 ? '#166534' : r.pctMeta >= 80 ? '#16A34A' : r.pctMeta >= 50 ? '#EAB308' : '#DC2626';
    html += `<tr>
      <td style="padding:4px 8px;font-weight:600;color:#1A5276;white-space:nowrap;position:sticky;left:0;background:white">${{t.replace('Millalemu ','M')}}</td>`;
    days.forEach(d => {{
      const v = D.grid[d]?.[t] || 0;
      const c = cellColor(v, metaDiaria[t]);
      const tip = v > 0
        ? `${{t}} · Día ${{d}}: ${{fmt(v)}} m³ (meta diaria ${{fmt(metaDiaria[t])}}, ${{Math.round(v/metaDiaria[t]*100)}}%)`
        : `${{t}} · Día ${{d}}: SIN REGISTRO`;
      html += `<td style="width:24px;height:24px;background:${{c.bg}};border-radius:3px;cursor:help" title="${{tip}}"></td>`;
    }});
    html += `<td style="padding:4px 10px;text-align:right;font-weight:700;color:#1E293B;border-left:2px solid #E2E8F0">${{fmt(r.vol)}}</td>`;
    html += `<td style="padding:4px 8px;text-align:right;color:#64748B">${{fmt(r.metaAcum)}}</td>`;
    html += `<td style="padding:4px 8px;text-align:center;font-weight:700;color:${{metaColor}}">${{r.pctMeta.toFixed(0)}}%</td>`;
    html += '</tr>';
  }});

  // Totales generales
  const volTot = totales.reduce((s, r) => s + r.vol, 0);
  const metaTot = totales.reduce((s, r) => s + r.metaAcum, 0);
  const pctTot = metaTot > 0 ? (volTot / metaTot * 100) : 0;
  const pctTotColor = pctTot >= 100 ? '#166534' : pctTot >= 80 ? '#16A34A' : pctTot >= 50 ? '#EAB308' : '#DC2626';
  html += `<tr style="background:#F1F5F9;font-weight:700">
    <td style="padding:6px 8px;color:#1A5276;position:sticky;left:0;background:#F1F5F9">TOTAL</td>
    <td colspan="${{days.length}}" style="padding:6px 8px;text-align:center;color:#64748B;font-size:10px;font-weight:500">${{days.length}} días transcurridos</td>
    <td style="padding:6px 10px;text-align:right;color:#1E293B;border-left:2px solid #CBD5E1">${{fmt(volTot)}}</td>
    <td style="padding:6px 8px;text-align:right;color:#64748B">${{fmt(metaTot)}}</td>
    <td style="padding:6px 8px;text-align:center;color:${{pctTotColor}}">${{pctTot.toFixed(0)}}%</td>
  </tr>`;
  html += '</tbody></table></div></div>';

  // ——— PANEL DE ALERTAS ———
  html += '<div style="min-width:220px;max-width:280px;background:#F8FAFC;border-radius:8px;padding:12px">';
  html += `<div style="font-size:12px;font-weight:700;color:#1A5276;margin-bottom:8px;text-transform:uppercase;letter-spacing:0.5px">Alertas</div>`;
  if (huecos.length === 0 && bajos.length === 0 && rachas.length === 0) {{
    html += '<div style="font-size:12px;color:#166534;padding:8px 0">✓ Sin alertas. Todas las faenas reportaron dentro de rango.</div>';
  }} else {{
    if (huecos.length > 0) {{
      html += `<div style="margin-bottom:10px"><div style="font-size:11px;font-weight:600;color:#DC2626;margin-bottom:4px">Sin reporte (${{huecos.length}})</div>`;
      huecos.forEach(h => {{
        const dow = DIAS_SEMANA[new Date(cfg.anio, cfg.mes-1, h.d).getDay()];
        html += `<div style="font-size:11px;padding:2px 0;color:#334155">${{h.t.replace('Millalemu ','M')}} · <strong>${{h.d}} ${{dow}}</strong></div>`;
      }});
      html += '</div>';
    }}
    if (rachas.length > 0) {{
      html += `<div style="margin-bottom:10px"><div style="font-size:11px;font-weight:600;color:#991B1B;margin-bottom:4px">Rachas (≥2 días seguidos)</div>`;
      rachas.forEach(r => {{
        html += `<div style="font-size:11px;padding:2px 0;color:#334155">${{r.t.replace('Millalemu ','M')}} · días <strong>${{r.inicio}}${{r.inicio!==r.fin?'-'+r.fin:''}}</strong> (${{r.len}} días)</div>`;
      }});
      html += '</div>';
    }}
    if (bajos.length > 0) {{
      html += `<div><div style="font-size:11px;font-weight:600;color:#E67E22;margin-bottom:4px">Bajo meta &lt;40% (${{bajos.length}})</div>`;
      bajos.slice(0, 6).forEach(b => {{
        html += `<div style="font-size:11px;padding:2px 0;color:#334155">${{b.t.replace('Millalemu ','M')}} · día ${{b.d}} · ${{fmt(b.v)}} m³</div>`;
      }});
      if (bajos.length > 6) html += `<div style="font-size:10px;color:#94A3B8;margin-top:2px">+${{bajos.length-6}} más…</div>`;
      html += '</div>';
    }}
  }}
  html += '</div></div>';

  dspEl.innerHTML = html;
}})();

// ═══════════════════════════════════════════════════════════
// ── RESUMEN EJECUTIVO (tiles arriba) ────────────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const rEl = document.getElementById('resumenEjecutivo');
  const re = D.resumenEj;
  const proyPct = re.proyPct;
  const proyColor = proyPct >= 100 ? '#166534' : proyPct >= 90 ? '#16A34A' : proyPct >= 80 ? '#EAB308' : '#DC2626';
  const cumplAct = cfg.tm > 0 ? (cfg.ta / cfg.tm * 100) : 0;

  const tile = (label, valor, sub, color) => `
    <div style="background:white;border-radius:10px;padding:14px 16px;box-shadow:0 1px 3px rgba(0,0,0,0.08);border-left:4px solid ${{color}}">
      <div style="font-size:10px;text-transform:uppercase;letter-spacing:0.5px;color:#64748B;font-weight:600">${{label}}</div>
      <div style="font-size:22px;font-weight:800;color:${{color}};margin:4px 0 2px">${{valor}}</div>
      <div style="font-size:11px;color:#64748B">${{sub}}</div>
    </div>`;

  const alertas = [];
  if (re.peor && re.peor.ci < 90) alertas.push(`${{re.peor.t.replace('Millalemu ','M')}} proy. ${{re.peor.ci.toFixed(0)}}%`);
  if (re.sinReporte > 0) alertas.push(`${{re.sinReporte}} reportes faltantes`);
  const alertasTxt = alertas.length > 0 ? alertas.join(' | ') : 'Sin alertas críticas';
  const alertasColor = alertas.length > 0 ? '#DC2626' : '#166534';

  rEl.innerHTML = `
    <div style="display:grid;grid-template-columns:repeat(auto-fit, minmax(180px, 1fr));gap:12px;margin-bottom:16px">
      ${{tile('Cumplimiento actual', cumplAct.toFixed(0) + '%',
        `${{fmt(cfg.ta)}} de ${{fmt(cfg.tm)}} m³`,
        cumplAct >= 50 ? '#16A34A' : '#EAB308')}}
      ${{tile('Proyección cierre mes', proyPct.toFixed(0) + '%',
        proyPct >= 100 ? 'Superará meta' : 'Brecha: ' + fmt(Math.round(cfg.tm - cfg.tm * proyPct/100)) + ' m³',
        proyColor)}}
      ${{tile('Mejor faena', re.mejor ? re.mejor.t.replace('Millalemu ','M') : '—',
        re.mejor ? re.mejor.ci.toFixed(0) + '% proy.' : '',
        '#16A34A')}}
      ${{tile('Peor faena', re.peor ? re.peor.t.replace('Millalemu ','M') : '—',
        re.peor ? re.peor.ci.toFixed(0) + '% proy.' : '',
        '#DC2626')}}
      ${{tile('Alertas', alertasTxt, `Día ${{cfg.dd}}/${{cfg.dm}}`, alertasColor)}}
    </div>`;
}})();

// ═══════════════════════════════════════════════════════════
// ── RANKING DE EFICIENCIA (m³/hr) ──────────────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const el = document.getElementById('rankingEfTable');
  const maxRend = Math.max(...D.rankingEf.map(r => r.rend));
  let html = '<table style="width:100%;border-collapse:collapse;font-size:12px"><thead><tr style="background:#1A5276;color:white">';
  html += '<th style="padding:8px 10px;text-align:left;border-radius:8px 0 0 0">#</th>';
  html += '<th style="padding:8px 10px;text-align:left">Faena</th>';
  html += '<th style="padding:8px 10px;text-align:right">m³/hr</th>';
  html += '<th style="padding:8px 10px;text-align:left">vs líder</th>';
  html += '<th style="padding:8px 10px;text-align:right">m³/árbol</th>';
  html += '<th style="padding:8px 10px;text-align:right;border-radius:0 8px 0 0">m³/ciclo</th>';
  html += '</tr></thead><tbody>';
  const medalla = ['🥇','🥈','🥉'];
  D.rankingEf.forEach((r, i) => {{
    const pctBar = maxRend > 0 ? (r.rend / maxRend * 100) : 0;
    const barColor = i === 0 ? '#166534' : i <= 2 ? '#16A34A' : i >= D.rankingEf.length - 2 ? '#DC2626' : '#64748B';
    html += `<tr style="background:${{i%2===0?'#F8FAFC':'white'}}">
      <td style="padding:6px 10px;font-weight:700;color:#1A5276">${{i < 3 ? medalla[i] : (i+1)}}</td>
      <td style="padding:6px 10px;font-weight:600;color:#1A5276">${{r.t.replace('Millalemu ','M')}}</td>
      <td style="padding:6px 10px;text-align:right;font-weight:700">${{r.rend.toFixed(2)}}</td>
      <td style="padding:6px 10px">
        <div style="background:#E2E8F0;height:8px;border-radius:4px;overflow:hidden;width:80px">
          <div style="width:${{pctBar}}%;height:100%;background:${{barColor}}"></div>
        </div>
      </td>
      <td style="padding:6px 10px;text-align:right;color:#64748B">${{r.mArb > 0 ? r.mArb.toFixed(3) : '—'}}</td>
      <td style="padding:6px 10px;text-align:right;color:#64748B">${{r.mCic > 0 ? r.mCic.toFixed(2) : '—'}}</td>
    </tr>`;
  }});
  html += '</tbody></table>';
  el.innerHTML = html;
}})();

// ═══════════════════════════════════════════════════════════
// ── MIX POR ESPECIE ─────────────────────────────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const el = document.getElementById('espMixBox');
  const ESP_COLORS = {{'PIRA':'#16A34A','EUGL':'#1A5276','EUNI':'#8E44AD'}};
  let html = '<div style="display:flex;flex-direction:column;gap:10px">';
  D.espMix.forEach(e => {{
    const color = ESP_COLORS[e.cod] || '#64748B';
    html += `
      <div>
        <div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:4px">
          <span style="font-weight:600;color:#1E293B">${{e.nom}}</span>
          <span style="font-weight:700;color:${{color}}">${{e.pct}}%</span>
        </div>
        <div style="background:#E2E8F0;height:14px;border-radius:4px;overflow:hidden">
          <div style="width:${{e.pct}}%;height:100%;background:${{color}};display:flex;align-items:center;justify-content:flex-end;padding-right:6px;color:white;font-size:10px;font-weight:700">${{fmt(e.vol)}} m³</div>
        </div>
        <div style="font-size:10px;color:#64748B;margin-top:2px">Rendimiento: ${{e.rend.toFixed(2)}} m³/hr</div>
      </div>`;
  }});
  html += '</div>';
  el.innerHTML = html;
}})();

// ═══════════════════════════════════════════════════════════
// ── ANÁLISIS POR PREDIO ─────────────────────────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const el = document.getElementById('prediosTable');
  const ESP_NAMES = {{'PIRA':'Pino','EUGL':'E.Globulus','EUNI':'E.Nitens'}};
  const volTotal = D.predios.reduce((s, p) => s + p.vol, 0);
  let html = '<table style="width:100%;border-collapse:collapse;font-size:12px"><thead><tr style="background:#1A5276;color:white">';
  html += '<th style="padding:8px 10px;text-align:left;border-radius:8px 0 0 0">Predio</th>';
  html += '<th style="padding:8px 10px;text-align:right">m³ Producidos</th>';
  html += '<th style="padding:8px 10px;text-align:right">% del total</th>';
  html += '<th style="padding:8px 10px;text-align:right">m³/hr</th>';
  html += '<th style="padding:8px 10px;text-align:left">Especies</th>';
  html += '<th style="padding:8px 10px;text-align:left;border-radius:0 8px 0 0">Faenas</th>';
  html += '</tr></thead><tbody>';
  D.predios.forEach((p, i) => {{
    const pct = volTotal > 0 ? (p.vol / volTotal * 100) : 0;
    const especiesTxt = p.esp.map(e => ESP_NAMES[e] || e).join(', ');
    const faenasTxt = p.eq.join(', ');
    html += `<tr style="background:${{i%2===0?'#F8FAFC':'white'}}">
      <td style="padding:6px 10px;color:#1A5276">
        <div style="font-weight:600">${{p.nombre || p.pr}}</div>
        ${{p.nombre ? `<div style="font-size:10px;color:#94A3B8">Cód. ${{p.pr}}</div>` : ''}}
      </td>
      <td style="padding:6px 10px;text-align:right;font-weight:700">${{fmt(p.vol)}}</td>
      <td style="padding:6px 10px;text-align:right">
        <div style="display:inline-block;background:#E2E8F0;height:8px;width:60px;border-radius:4px;overflow:hidden;vertical-align:middle;margin-right:6px">
          <div style="width:${{pct}}%;height:100%;background:#16A34A"></div>
        </div>${{pct.toFixed(1)}}%
      </td>
      <td style="padding:6px 10px;text-align:right">${{p.rend.toFixed(2)}}</td>
      <td style="padding:6px 10px;font-size:11px;color:#334155">${{especiesTxt}}</td>
      <td style="padding:6px 10px;font-size:11px;color:#334155">${{faenasTxt}}</td>
    </tr>`;
  }});
  html += '</tbody></table>';
  el.innerHTML = html;
}})();

// ═══════════════════════════════════════════════════════════
// ── HEATMAP TIEMPOS PERDIDOS ────────────────────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const el = document.getElementById('tmHeatmapBox');
  const lastDayG = Math.max(...Object.keys(D.grid).map(Number));
  const days = Array.from({{length: lastDayG}}, (_, i) => i + 1);

  // Encontrar el máximo para escalar colores
  let maxMin = 0;
  TEAMS.forEach(t => days.forEach(d => {{
    const m = D.tmHeatmap[d]?.[t] || 0;
    if (m > maxMin) maxMin = m;
  }}));

  // Escala: sin datos = blanco, bajo = amarillo claro, medio = naranja, alto = rojo
  const cellColor = (min) => {{
    if (min <= 0) return '#F8FAFC';
    const r = min / maxMin;
    if (r < 0.25) return '#FEF3C7';    // amarillo muy claro
    if (r < 0.5) return '#FBBF24';     // amarillo
    if (r < 0.75) return '#F97316';    // naranja
    return '#DC2626';                  // rojo
  }};

  let html = `<div style="margin-bottom:10px;font-size:11px;color:#64748B">Escala por horas de tiempo perdido (más oscuro = más TM). Cada celda en horas. Máximo registrado: ${{(maxMin/60).toFixed(1)}} h.</div>`;
  html += '<div style="overflow-x:auto"><table style="border-collapse:separate;border-spacing:2px;font-size:11px"><thead><tr>';
  html += '<th style="padding:4px 8px;text-align:left;color:#64748B;font-weight:600;position:sticky;left:0;background:white">Faena</th>';
  days.forEach(d => {{
    const dow = DIAS_SEMANA[new Date(cfg.anio, cfg.mes-1, d).getDay()];
    html += `<th style="padding:2px;text-align:center;font-weight:500;color:#94A3B8;font-size:10px;min-width:24px">
      <div>${{d}}</div><div style="font-size:9px">${{dow[0]}}</div></th>`;
  }});
  html += '<th style="padding:4px 8px;text-align:right;color:#64748B;font-weight:600">Total (hrs)</th>';
  html += '</tr></thead><tbody>';
  TEAMS.forEach(t => {{
    let totMin = 0;
    html += `<tr><td style="padding:4px 8px;font-weight:600;color:#1A5276;white-space:nowrap;position:sticky;left:0;background:white">${{t.replace('Millalemu ','M')}}</td>`;
    days.forEach(d => {{
      const min = D.tmHeatmap[d]?.[t] || 0;
      totMin += min;
      const color = cellColor(min);
      const tip = min > 0 ? `${{t}} · Día ${{d}}: ${{min}} min (${{(min/60).toFixed(1)}} hrs)` : `${{t}} · Día ${{d}}: sin TM registrado`;
      html += `<td style="width:24px;height:24px;background:${{color}};border-radius:3px;cursor:help;text-align:center;font-size:9px;color:${{min>maxMin*0.5?'white':'#1E293B'}}" title="${{tip}}">${{min > 0 ? Math.round(min/60*10)/10 : ''}}</td>`;
    }});
    html += `<td style="padding:4px 8px;text-align:right;font-weight:700;color:#C0392B">${{(totMin/60).toFixed(1)}}</td>`;
    html += '</tr>';
  }});
  html += '</tbody></table></div>';
  el.innerHTML = html;
}})();

// ═══════════════════════════════════════════════════════════
// ── MTBF / MTTR ─────────────────────────────────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const el = document.getElementById('mtbfMttrTable');
  let html = '<table style="width:100%;border-collapse:collapse;font-size:12px"><thead><tr style="background:#1A5276;color:white">';
  html += '<th style="padding:8px 10px;text-align:left;border-radius:8px 0 0 0">Faena</th>';
  html += '<th style="padding:8px 10px;text-align:right" title="Número de fallas de Mantención">Fallas</th>';
  html += '<th style="padding:8px 10px;text-align:right" title="Mean Time Between Failures: horas de operación efectiva entre fallas">MTBF (hrs)</th>';
  html += '<th style="padding:8px 10px;text-align:right;border-radius:0 8px 0 0" title="Mean Time To Repair: horas promedio para reparar una falla">MTTR (h)</th>';
  html += '</tr></thead><tbody>';
  // Ordenar por más fallas arriba
  const sorted = [...D.mtbfMttr].sort((a, b) => b.fallas - a.fallas);
  sorted.forEach((r, i) => {{
    const mtbfColor = r.mtbf === null ? '#94A3B8' : r.mtbf >= 40 ? '#166534' : r.mtbf >= 20 ? '#EAB308' : '#DC2626';
    const mttrColor = r.mttr === null ? '#94A3B8' : r.mttr <= 60 ? '#166534' : r.mttr <= 120 ? '#EAB308' : '#DC2626';
    html += `<tr style="background:${{i%2===0?'#F8FAFC':'white'}}">
      <td style="padding:6px 10px;font-weight:600;color:#1A5276">${{r.t.replace('Millalemu ','M')}}</td>
      <td style="padding:6px 10px;text-align:right;font-weight:700">${{r.fallas}}</td>
      <td style="padding:6px 10px;text-align:right;font-weight:700;color:${{mtbfColor}}">${{r.mtbf === null ? '—' : r.mtbf}}</td>
      <td style="padding:6px 10px;text-align:right;font-weight:700;color:${{mttrColor}}">${{r.mttr === null ? '—' : (r.mttr/60).toFixed(1)+'h'}}</td>
    </tr>`;
  }});
  html += '</tbody></table>';
  html += `<div style="font-size:11px;color:#64748B;margin-top:10px;line-height:1.7;background:#F8FAFC;border-radius:8px;padding:10px 12px">
    <b style="color:#1A5276">MTBF</b> <span style="color:#94A3B8">(Mean Time Between Failures)</span> — horas promedio que la máquina <b>trabaja entre una falla y la siguiente</b>. Mientras más alto, falla menos = <b style="color:#166534">mejor</b>.<br>
    <b style="color:#1A5276">MTTR</b> <span style="color:#94A3B8">(Mean Time To Repair)</span> — horas promedio que toma <b>reparar</b> cada falla. Mientras más bajo, se arregla más rápido = <b style="color:#166534">mejor</b>.<br>
    <span style="color:#94A3B8">Solo considera fallas de Mantención.</span>
  </div>`;
  el.innerHTML = html;
}})();

// ═══════════════════════════════════════════════════════════
// ── COMPARATIVA VS MES ANTERIOR ────────────────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const el = document.getElementById('compMesBox');
  const c = D.compMes;
  const MESES_NOMBRE = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'];
  if (!c) {{
    el.innerHTML = '<p style="text-align:center;color:#64748B;padding:20px">Sin datos del mes anterior para comparar.</p>';
    return;
  }}
  if (!c.suficiente) {{
    el.innerHTML = `<p style="text-align:center;color:#64748B;padding:20px">Mes anterior (${{MESES_NOMBRE[c.mesPrev]}}) tiene solo ${{c.diasPrev}} día${{c.diasPrev!==1?'s':''}} de datos. Comparativa no confiable.</p>`;
    return;
  }}
  const row = (label, prev, act, delta, unit) => {{
    const color = delta >= 0 ? '#166534' : '#DC2626';
    const arrow = delta >= 0 ? '▲' : '▼';
    return `<tr>
      <td style="padding:6px 10px;color:#64748B">${{label}}</td>
      <td style="padding:6px 10px;text-align:right">${{fmt(prev)}} ${{unit}}</td>
      <td style="padding:6px 10px;text-align:right;font-weight:700">${{fmt(act)}} ${{unit}}</td>
      <td style="padding:6px 10px;text-align:right;font-weight:700;color:${{color}}">${{arrow}} ${{Math.abs(delta)}}%</td>
    </tr>`;
  }};
  let html = '<table style="width:100%;border-collapse:collapse;font-size:12px"><thead><tr style="background:#F1F5F9;color:#64748B">';
  html += '<th style="padding:8px 10px;text-align:left;border-radius:8px 0 0 0">Métrica</th>';
  html += `<th style="padding:8px 10px;text-align:right">${{MESES_NOMBRE[c.mesPrev]}} (${{c.diasPrev}}d)</th>`;
  html += `<th style="padding:8px 10px;text-align:right">${{cfg.mesNombre}} (${{cfg.dd}}d)</th>`;
  html += '<th style="padding:8px 10px;text-align:right;border-radius:0 8px 0 0">Δ</th>';
  html += '</tr></thead><tbody>';
  html += row('Prom. diario', c.promPrev, c.promAct, c.deltaProm, 'm³');
  html += row('Rendimiento', c.rendPrev, c.rendAct, c.deltaRend, 'm³/hr');
  html += '</tbody></table>';
  html += `<div style="font-size:10px;color:#94A3B8;margin-top:8px">Comparación en base a promedios diarios (normalizada por días de datos).</div>`;
  el.innerHTML = html;
}})();

// ═══════════════════════════════════════════════════════════
// ── ANÁLISIS DE OBSERVACIONES DE TIEMPOS PERDIDOS ──────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const el = document.getElementById('obsAnalisisBox');
  const oa = D.obsAnalisis || {{}};
  if (!oa.totalComentarios || oa.totalComentarios === 0) {{
    el.innerHTML = '<p style="text-align:center;color:#64748B;padding:20px">Sin observaciones en los registros de Tiempos Perdidos.</p>';
    return;
  }}

  const totalMin = oa.porCategoria.reduce((s, c) => s + c.minutos, 0);
  let html = `<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">`;

  // ── LADO IZQUIERDO: Categorías + Palabras clave ──
  html += '<div>';
  html += `<div style="font-size:11px;color:#64748B;margin-bottom:8px">${{oa.totalComentarios}} comentarios analizados · ${{(totalMin/60).toFixed(1)}} hrs totales</div>`;

  // Categorías con barras proporcionales
  html += '<div style="margin-bottom:14px"><div style="font-weight:700;color:#1A5276;font-size:12px;margin-bottom:6px">Por categoría</div>';
  oa.porCategoria.forEach(c => {{
    const pct = totalMin > 0 ? (c.minutos / totalMin * 100) : 0;
    html += `<div style="margin-bottom:6px">
      <div style="display:flex;justify-content:space-between;font-size:11px;margin-bottom:2px">
        <span style="color:#1E293B"><strong>${{c.cat}}</strong> · ${{c.n}} coment.</span>
        <span style="color:${{c.color}};font-weight:700">${{(c.minutos/60).toFixed(1)}}h (${{pct.toFixed(0)}}%)</span>
      </div>
      <div style="background:#F1F5F9;height:8px;border-radius:4px;overflow:hidden">
        <div style="width:${{pct}}%;height:100%;background:${{c.color}}"></div>
      </div>
    </div>`;
  }});
  html += '</div>';

  // Palabras clave como burbujas (bubble chart / circle packing)
  if (oa.palabrasClave && oa.palabrasClave.length > 0) {{
    html += '<div><div style="font-weight:700;color:#1A5276;font-size:12px;margin-bottom:6px">Palabras más frecuentes (tamaño = frecuencia)</div>';
    html += '<div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;justify-content:flex-start;padding:12px;background:linear-gradient(135deg,#F8FAFC,#EFF6FF);border-radius:8px;min-height:180px">';
    const maxN = oa.palabrasClave[0].n;
    const minN = oa.palabrasClave[oa.palabrasClave.length-1].n;
    // Paleta de colores — las más frecuentes más oscuras/rojas
    const palette = ['#DC2626','#F59E0B','#D97706','#16A34A','#2563EB','#8B5CF6','#0891B2','#64748B'];
    oa.palabrasClave.forEach((p, i) => {{
      // Tamaño del círculo: entre 40 y 100 px
      const ratio = maxN > minN ? (p.n - minN) / (maxN - minN) : 0.5;
      const size = Math.round(42 + ratio * 60);
      const fsText = Math.round(10 + ratio * 6);
      const fsNum = Math.round(8 + ratio * 3);
      const c = palette[Math.min(i, palette.length-1)];
      html += `<div style="width:${{size}}px;height:${{size}}px;border-radius:50%;background:${{c}};display:flex;flex-direction:column;align-items:center;justify-content:center;color:white;font-weight:700;box-shadow:0 2px 6px ${{c}}66;flex:0 0 auto">
        <span style="font-size:${{fsText}}px;line-height:1;text-align:center;padding:0 4px;word-break:break-word">${{p.palabra}}</span>
        <span style="font-size:${{fsNum}}px;opacity:0.8;margin-top:2px">${{p.n}}</span>
      </div>`;
    }});
    html += '</div></div>';
  }}
  html += '</div>';

  // ── LADO DERECHO: Top comentarios ──
  html += '<div>';
  html += '<div style="font-weight:700;color:#1A5276;font-size:12px;margin-bottom:6px">Top comentarios (por tiempo acumulado)</div>';
  html += '<div style="max-height:420px;overflow-y:auto">';
  oa.topComentarios.forEach((c, i) => {{
    html += `<div style="background:white;border-left:3px solid ${{c.color}};padding:6px 10px;margin-bottom:4px;border-radius:3px;font-size:11px">
      <div style="display:flex;justify-content:space-between;gap:8px">
        <span style="color:#1E293B;flex:1">${{c.txt}}</span>
        <span style="color:${{c.color}};font-weight:700;white-space:nowrap">${{(c.minutos/60).toFixed(1)}}h</span>
      </div>
      <div style="font-size:10px;color:#94A3B8;margin-top:2px">
        <span style="display:inline-block;background:${{c.color}}22;color:${{c.color}};padding:1px 6px;border-radius:2px">${{c.cat}}</span>
        · ${{c.n}} mención${{c.n!==1?'es':''}}
      </div>
    </div>`;
  }});
  html += '</div>';
  html += '</div>';

  html += '</div>';

  // ── SECCIÓN INFERIOR: Por equipo (matriz de categorías) ──
  html += '<div style="margin-top:14px;border-top:1px solid #E2E8F0;padding-top:12px">';
  html += '<div style="font-weight:700;color:#1A5276;font-size:12px;margin-bottom:8px">Categoría dominante por faena</div>';
  html += '<table style="width:100%;border-collapse:collapse;font-size:11px"><thead><tr style="background:#F1F5F9">';
  html += '<th style="padding:4px 8px;text-align:left">Faena</th>';
  html += '<th style="padding:4px 8px;text-align:left">Categoría dominante</th>';
  html += '<th style="padding:4px 8px;text-align:right">Tiempo</th>';
  html += '<th style="padding:4px 8px;text-align:right">Coment.</th>';
  html += '<th style="padding:4px 8px;text-align:left">Otras categorías</th>';
  html += '</tr></thead><tbody>';
  TEAMS.forEach(t => {{
    const cats = oa.porEquipo[t];
    if (!cats || cats.length === 0) return;
    const top = cats[0];
    const otras = cats.slice(1).map(c => `${{c.cat}} (${{(c.minutos/60).toFixed(1)}}h)`).join(' · ');
    html += `<tr>
      <td style="padding:4px 8px;font-weight:600;color:#1A5276">${{t.replace('Millalemu ','M')}}</td>
      <td style="padding:4px 8px"><span style="background:${{top.color}}22;color:${{top.color}};padding:2px 6px;border-radius:3px;font-weight:700">${{top.cat}}</span></td>
      <td style="padding:4px 8px;text-align:right;font-weight:700;color:${{top.color}}">${{(top.minutos/60).toFixed(1)}}h</td>
      <td style="padding:4px 8px;text-align:right">${{top.n}}</td>
      <td style="padding:4px 8px;font-size:10px;color:#64748B">${{otras || '—'}}</td>
    </tr>`;
  }});
  html += '</tbody></table></div>';

  el.innerHTML = html;
}})();

// ═══════════════════════════════════════════════════════════
// ── ANÁLISIS 80/20 PARETO GLOBAL (PESTAÑA TIEMPOS PERDIDOS) ─
// ═══════════════════════════════════════════════════════════
(() => {{
  const causas = D.tmParetoGlobal || [];
  if (!causas.length) {{
    document.getElementById('tmParetoBox').innerHTML = '<div style="padding:14px;color:#64748B">Sin datos de tiempos perdidos para analizar.</div>';
    return;
  }}

  const totalH = causas.reduce((s,c) => s+c.h, 0);
  const totalEv = causas.reduce((s,c) => s+c.ev, 0);
  // Causas que cubren el 80% (vitales)
  let vital = [];
  for (const c of causas) {{ vital.push(c); if (c.pctAcum >= 80) break; }}

  // Hallazgo box
  const halBox = `
    <div style="background:#FEF3C7;border-left:4px solid #D97706;padding:12px 16px;border-radius:8px;font-size:13px">
      <strong style="color:#92400E">Hallazgo 80/20:</strong>
      <span style="color:#78350F">${{vital.length}} causas (de ${{causas.length}}) concentran el ${{vital[vital.length-1].pctAcum.toFixed(0)}}% de las
      <strong>${{totalH.toFixed(1)}} hrs perdidas</strong> en el mes (${{totalEv}} eventos).</span>
      <div style="margin-top:6px;font-size:11px;color:#92400E">Atacar estas causas tiene el mayor retorno por unidad de esfuerzo.</div>
    </div>`;
  document.getElementById('tmParetoBox').innerHTML = halBox;

  // Gráfico Chart.js: barras + línea acumulada
  const top15 = causas.slice(0, 15);
  const ctx = document.getElementById('chartParetoGlobal').getContext('2d');
  new Chart(ctx, {{
    data: {{
      labels: top15.map(c => c.n.length > 38 ? c.n.slice(0,38)+'…' : c.n),
      datasets: [
        {{
          type: 'bar',
          label: 'Horas perdidas',
          data: top15.map(c => c.h),
          backgroundColor: top15.map(c => c.pctAcum <= 80 ? '#DC2626' : '#94A3B8'),
          borderRadius: 4,
          yAxisID: 'y',
        }},
        {{
          type: 'line',
          label: '% Acumulado',
          data: top15.map(c => c.pctAcum),
          borderColor: '#D97706',
          backgroundColor: 'transparent',
          borderWidth: 2.5,
          pointRadius: 4,
          pointBackgroundColor: '#D97706',
          tension: 0.2,
          yAxisID: 'y1',
        }}
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      indexAxis: 'y',
      plugins: {{
        legend: {{ position: 'top' }},
        tooltip: {{ callbacks: {{
          afterLabel: (ctx) => {{
            const c = top15[ctx.dataIndex];
            return [`Eventos: ${{c.ev}} en ${{c.d}} días`,
                    `${{c.h.toFixed(1)}} h (${{c.pct.toFixed(1)}}% del total)`];
          }}
        }}}}
      }},
      scales: {{
        x: {{ beginAtZero: true, position: 'top', title: {{ display: true, text: 'Horas / %' }} }},
        y: {{ ticks: {{ font: {{ size: 11 }} }} }},
        y1: {{ display: false }}
      }}
    }}
  }});
}})();

// ═══════════════════════════════════════════════════════════
// ── HEATMAP CAUSA × EQUIPO (focos crónicos) ────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const heat = D.tmHeatmapCE || [];
  const el = document.getElementById('tmHeatmapCETable');
  if (!heat.length) {{ el.innerHTML = '<div style="padding:14px;color:#64748B">Sin causas vitales para cruzar.</div>'; return; }}

  const equipos = TEAMS;
  // Calcular max por causa para escala de color
  const colorScale = (val, max) => {{
    if (val === 0 || max === 0) return '#F8FAFC';
    const r = val / max;
    if (r >= 0.6) return '#FCA5A5';   // rojo
    if (r >= 0.3) return '#FED7AA';   // naranja
    if (r >= 0.1) return '#FEF3C7';   // amarillo
    return '#ECFCCB';                  // verde claro
  }};

  let h = '<table style="width:100%;border-collapse:collapse;font-size:11px;min-width:780px">';
  h += '<thead><tr style="background:#1A5276;color:white"><th style="padding:8px 10px;text-align:left;min-width:240px">Causa</th>';
  equipos.forEach(eq => h += `<th style="padding:8px 6px;text-align:center;font-size:10px">${{eq.replace('Millalemu ','M')}}</th>`);
  h += '<th style="padding:8px 10px;text-align:right;background:#0F3D5E">Total h</th>';
  h += '<th style="padding:8px 10px;text-align:left;background:#0F3D5E;font-size:10px">Concentración</th>';
  h += '</tr></thead><tbody>';

  heat.forEach((row, idx) => {{
    const max = Math.max(...Object.values(row.porEquipo));
    h += `<tr style="background:${{idx%2?'#F8FAFC':'#FFFFFF'}}">`;
    h += `<td style="padding:6px 10px;font-weight:600;color:#1A5276">${{row.causa}}</td>`;
    equipos.forEach(eq => {{
      const v = row.porEquipo[eq] || 0;
      const bg = colorScale(v, max);
      const txt = v > 0 ? v.toFixed(1) : '·';
      const wt = (v === max && v > 0) ? '700' : '400';
      h += `<td style="padding:6px;text-align:center;background:${{bg}};font-weight:${{wt}}">${{txt}}</td>`;
    }});
    h += `<td style="padding:6px 10px;text-align:right;font-weight:700;background:#EBF5FB">${{row.totalH.toFixed(1)}}</td>`;
    const focoLabel = row.topPct >= 50
      ? `<span style="color:#DC2626;font-weight:600">⚠ ${{row.topEquipo.replace('Millalemu ','M')}} (${{row.topPct.toFixed(0)}}%)</span>`
      : `<span style="color:#64748B">Distribuida</span>`;
    h += `<td style="padding:6px 10px;background:#F8FAFC">${{focoLabel}}</td>`;
    h += '</tr>';
  }});
  h += '</tbody></table>';
  el.innerHTML = h;
}})();

// ═══════════════════════════════════════════════════════════
// ── RECOMENDACIONES ACCIONABLES PARA MAYO ──────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const recs = D.tmRecomendaciones || [];
  const el = document.getElementById('tmRecBox');
  if (!recs.length) {{
    el.innerHTML = '<div style="padding:14px;color:#64748B">Sin recomendaciones accionables (insuficiente recurrencia en las causas top).</div>';
    return;
  }}

  let totalAhorroH = 0, totalAhorroM3 = 0;
  recs.forEach(r => {{ totalAhorroH += r.ahorroPotH; totalAhorroM3 += r.ahorroPotM3; }});

  let h = '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:10px;margin-bottom:12px">';
  recs.forEach(r => {{
    h += `<div style="background:white;border-radius:8px;padding:12px 14px;border-left:4px solid #16A34A;box-shadow:0 1px 3px rgba(0,0,0,0.05)">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:8px">
        <div style="font-weight:700;color:#166534;font-size:13px">#${{r.id}} ${{r.causa}}</div>
        <div style="background:#16A34A;color:white;padding:2px 6px;border-radius:4px;font-size:10px;white-space:nowrap">${{r.recurrencia}} días</div>
      </div>
      <div style="font-size:11px;color:#64748B;margin-top:6px">
        <strong>${{r.horasMes.toFixed(1)}} h</strong> perdidas en abril (${{r.pctMes.toFixed(1)}}%) · <span style="color:#92400E">${{r.foco}}</span>
      </div>
      <div style="background:#F0FDF4;padding:8px 10px;border-radius:6px;margin-top:8px;font-size:11px;line-height:1.4">
        <div style="color:#166534;font-weight:600;margin-bottom:4px">▶ Acción:</div>
        <div style="color:#1F2937">${{r.accion}}</div>
      </div>
      <div style="display:flex;justify-content:space-between;margin-top:8px;font-size:11px;border-top:1px dashed #D1FAE5;padding-top:6px">
        <span style="color:#64748B">Ahorro potencial:</span>
        <span style="font-weight:700;color:#16A34A">${{r.ahorroPotH}} h ≈ ${{r.ahorroPotM3.toLocaleString('es-CL')}} m³</span>
      </div>
    </div>`;
  }});
  h += '</div>';

  // Total potencial agregado
  h += `<div style="background:white;border-radius:8px;padding:14px 18px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,0.05);border:2px dashed #16A34A">
    <div style="font-size:11px;color:#64748B;text-transform:uppercase;letter-spacing:0.5px">Si se atacan las ${{recs.length}} causas: ahorro mensual estimado</div>
    <div style="margin-top:6px;display:flex;justify-content:center;gap:30px;align-items:baseline">
      <div><span style="font-size:32px;font-weight:800;color:#16A34A">${{totalAhorroH.toFixed(0)}}</span> <span style="color:#64748B;font-size:12px">horas</span></div>
      <div style="color:#94A3B8">≈</div>
      <div><span style="font-size:32px;font-weight:800;color:#16A34A">${{totalAhorroM3.toLocaleString('es-CL')}}</span> <span style="color:#64748B;font-size:12px">m³ adicionales</span></div>
    </div>
  </div>`;
  el.innerHTML = h;
}})();

// ═══════════════════════════════════════════════════════════
// ── PESTAÑA: COMPARATIVO MENSUAL HISTÓRICO ─────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const hist = D.historico || [];
  const MESES_ABBR = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'];

  if (!hist.length) {{
    document.getElementById('histKpis').innerHTML =
      '<div style="padding:20px;color:#64748B">Sin datos históricos disponibles. Cargar historico_cierres_mensuales.csv en la carpeta del proyecto.</div>';
    return;
  }}

  // Agrupar por mes (key: anio*100+mes), ordenar cronológicamente
  const meses = {{}};
  hist.forEach(r => {{
    const k = r.anio * 100 + r.mes;
    if (!meses[k]) meses[k] = {{ anio: r.anio, mes: r.mes, label: MESES_ABBR[r.mes]+' '+String(r.anio).slice(2), equipos: {{}}, vol: 0, meta: 0 }};
    meses[k].equipos[r.equipo] = r;
    meses[k].vol += r.vol;
    meses[k].meta += r.meta;
  }});
  const mesesArr = Object.keys(meses).map(k => meses[k]).sort((a,b) => (a.anio*100+a.mes)-(b.anio*100+b.mes));
  const equiposSet = [...new Set(hist.map(r => r.equipo))].sort();

  // ── KPIs ─────────────────────────────────────────────────
  const totalAcum = mesesArr.reduce((s,m) => s+m.vol, 0);
  const mejorMes = mesesArr.reduce((a,b) => (b.vol/b.meta||0)>(a.vol/a.meta||0)?b:a);
  const peorMes  = mesesArr.reduce((a,b) => (b.vol/b.meta||1)<(a.vol/a.meta||1)?b:a);
  const promCumpl = mesesArr.reduce((s,m) => s + (m.vol/m.meta*100||0), 0)/mesesArr.length;
  const kpiHTML = [
    {{label:'Meses cargados', val:mesesArr.length, sub:mesesArr[0].label+' → '+mesesArr[mesesArr.length-1].label, color:'#1A5276'}},
    {{label:'Producción acumulada', val:Math.round(totalAcum).toLocaleString('es-CL'), sub:'m³ SSC', color:'#2980B9'}},
    {{label:'Cumplimiento promedio', val:promCumpl.toFixed(1)+'%', sub:promCumpl>=90?'En meta':'Bajo meta', color:promCumpl>=90?'#16A34A':promCumpl>=70?'#D97706':'#DC2626'}},
    {{label:'Mejor mes', val:mejorMes.label, sub:(mejorMes.vol/mejorMes.meta*100).toFixed(1)+'% · '+Math.round(mejorMes.vol).toLocaleString('es-CL')+' m³', color:'#16A34A'}},
    {{label:'Mes con brecha', val:peorMes.label, sub:(peorMes.vol/peorMes.meta*100).toFixed(1)+'% · '+Math.round(peorMes.vol).toLocaleString('es-CL')+' m³', color:'#DC2626'}},
  ];
  document.getElementById('histKpis').innerHTML = kpiHTML.map(k =>
    `<div style="flex:1;min-width:160px;background:#F8FAFC;border-left:4px solid ${{k.color}};padding:10px 14px;border-radius:6px">
       <div style="font-size:11px;color:#64748B;text-transform:uppercase;letter-spacing:.5px">${{k.label}}</div>
       <div style="font-size:22px;font-weight:800;color:${{k.color}};line-height:1.2;margin:2px 0">${{k.val}}</div>
       <div style="font-size:11px;color:#94A3B8">${{k.sub}}</div>
     </div>`
  ).join('');

  // ── Gráfico 1: Total faena por mes (barras vol vs línea meta) ─────
  const ctx1 = document.getElementById('chartHistTotal').getContext('2d');
  new Chart(ctx1, {{
    type: 'bar',
    data: {{
      labels: mesesArr.map(m => m.label),
      datasets: [
        {{
          label: 'Producción Real',
          data: mesesArr.map(m => Math.round(m.vol)),
          backgroundColor: mesesArr.map(m => (m.vol/m.meta)>=0.9?'#16A34A':(m.vol/m.meta)>=0.7?'#D97706':'#DC2626'),
          borderRadius: 6,
        }},
        {{
          label: 'Meta',
          data: mesesArr.map(m => Math.round(m.meta)),
          type: 'line',
          borderColor: '#1A5276',
          backgroundColor: 'transparent',
          borderWidth: 2.5,
          borderDash: [6,4],
          pointRadius: 5,
          pointBackgroundColor: '#1A5276',
        }}
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ position: 'top' }},
        tooltip: {{
          callbacks: {{
            afterLabel: function(ctx) {{
              const m = mesesArr[ctx.dataIndex];
              if (ctx.datasetIndex === 0) {{
                const pct = (m.vol/m.meta*100).toFixed(1);
                return 'Cumplimiento: ' + pct + '%';
              }}
              return null;
            }}
          }}
        }}
      }},
      scales: {{
        y: {{ beginAtZero: true, ticks: {{ callback: v => v.toLocaleString('es-CL')+' m³' }} }}
      }}
    }}
  }});

  // ── Tabla pivot Volumen ──────────────────────────────────
  const buildPivot = (valFn, fmtFn, colorFn) => {{
    let h = '<table style="width:100%;border-collapse:collapse;font-size:12px">';
    h += '<thead><tr style="background:#1A5276;color:white">';
    h += '<th style="padding:8px;text-align:left">Equipo</th>';
    mesesArr.forEach(m => h += `<th style="padding:8px;text-align:right">${{m.label}}</th>`);
    h += '<th style="padding:8px;text-align:right;background:#0F3D5E">Total</th>';
    h += '</tr></thead><tbody>';
    equiposSet.forEach((eq, i) => {{
      h += `<tr style="background:${{i%2?'#F8FAFC':'#FFFFFF'}}">`;
      h += `<td style="padding:6px 10px;font-weight:600;color:#1A5276">${{eq}}</td>`;
      let total = 0, n = 0;
      mesesArr.forEach(m => {{
        const r = m.equipos[eq];
        const v = r ? valFn(r) : null;
        const cell_color = (v != null && colorFn) ? colorFn(r) : 'transparent';
        const display = v != null ? fmtFn(v) : '—';
        h += `<td style="padding:6px 10px;text-align:right;background:${{cell_color}}">${{display}}</td>`;
        if (r && r.vol) {{ total += r.vol; n++; }}
      }});
      h += `<td style="padding:6px 10px;text-align:right;font-weight:700;background:#EBF5FB">${{Math.round(total).toLocaleString('es-CL')}}</td>`;
      h += '</tr>';
    }});
    // Fila TOTAL
    h += '<tr style="background:#1A5276;color:white;font-weight:700">';
    h += '<td style="padding:8px 10px">TOTAL FAENA</td>';
    let granTotal = 0;
    mesesArr.forEach(m => {{
      h += `<td style="padding:8px 10px;text-align:right">${{Math.round(m.vol).toLocaleString('es-CL')}}</td>`;
      granTotal += m.vol;
    }});
    h += `<td style="padding:8px 10px;text-align:right;background:#0F3D5E">${{Math.round(granTotal).toLocaleString('es-CL')}}</td>`;
    h += '</tr></tbody></table>';
    return h;
  }};
  document.getElementById('histTablaVol').innerHTML = buildPivot(
    r => r.vol > 0 ? r.vol : null,
    v => Math.round(v).toLocaleString('es-CL'),
    null
  );

  // ── Tabla pivot Cumplimiento % ──────────────────────────
  const cumplColor = r => {{
    if (!r.meta) return '#F1F5F9';
    const p = r.vol/r.meta*100;
    if (p >= 100) return '#DCFCE7';
    if (p >= 90)  return '#ECFCCB';
    if (p >= 70)  return '#FEF9C3';
    if (p >= 50)  return '#FED7AA';
    return '#FEE2E2';
  }};
  let hPct = '<table style="width:100%;border-collapse:collapse;font-size:12px">';
  hPct += '<thead><tr style="background:#1A5276;color:white">';
  hPct += '<th style="padding:8px;text-align:left">Equipo</th>';
  mesesArr.forEach(m => hPct += `<th style="padding:8px;text-align:right">${{m.label}}</th>`);
  hPct += '<th style="padding:8px;text-align:right;background:#0F3D5E">Promedio</th></tr></thead><tbody>';
  equiposSet.forEach((eq, i) => {{
    hPct += `<tr style="background:${{i%2?'#F8FAFC':'#FFFFFF'}}">`;
    hPct += `<td style="padding:6px 10px;font-weight:600;color:#1A5276">${{eq}}</td>`;
    let sumP = 0, n = 0;
    mesesArr.forEach(m => {{
      const r = m.equipos[eq];
      if (r && r.meta > 0) {{
        const p = r.vol/r.meta*100;
        hPct += `<td style="padding:6px 10px;text-align:right;background:${{cumplColor(r)}};font-weight:600">${{p.toFixed(1)}}%</td>`;
        sumP += p; n++;
      }} else {{
        hPct += `<td style="padding:6px 10px;text-align:right;color:#CBD5E1">—</td>`;
      }}
    }});
    const avgP = n > 0 ? (sumP/n).toFixed(1) : '—';
    hPct += `<td style="padding:6px 10px;text-align:right;font-weight:700;background:#EBF5FB">${{avgP}}${{n>0?'%':''}}</td>`;
    hPct += '</tr>';
  }});
  hPct += '<tr style="background:#1A5276;color:white;font-weight:700">';
  hPct += '<td style="padding:8px 10px">CUMPL FAENA</td>';
  let sumFaena = 0;
  mesesArr.forEach(m => {{
    const p = m.meta > 0 ? m.vol/m.meta*100 : 0;
    hPct += `<td style="padding:8px 10px;text-align:right">${{p.toFixed(1)}}%</td>`;
    sumFaena += p;
  }});
  const avgFaena = (sumFaena/mesesArr.length).toFixed(1);
  hPct += `<td style="padding:8px 10px;text-align:right;background:#0F3D5E">${{avgFaena}}%</td>`;
  hPct += '</tr></tbody></table>';
  document.getElementById('histTablaCumpl').innerHTML = hPct;

  // ── Gráfico líneas: tendencia por equipo ────────────────
  const ctx2 = document.getElementById('chartHistEquipos').getContext('2d');
  new Chart(ctx2, {{
    type: 'line',
    data: {{
      labels: mesesArr.map(m => m.label),
      datasets: equiposSet.map((eq, i) => ({{
        label: eq,
        data: mesesArr.map(m => m.equipos[eq] ? Math.round(m.equipos[eq].vol) : null),
        borderColor: COLORS[i % COLORS.length],
        backgroundColor: COLORS[i % COLORS.length] + '22',
        borderWidth: 2,
        tension: 0.3,
        spanGaps: true,
        pointRadius: 4,
      }}))
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{ legend: {{ position: 'bottom' }} }},
      scales: {{
        y: {{ beginAtZero: true, title: {{ display: true, text: 'm³ SSC' }} }}
      }}
    }}
  }});

  // ── Tiempos Perdidos por Mes (apilado Mant/Oper/Proceso) + tabla ──
  const tmM = D.tmMensual || [];
  if (tmM.length && document.getElementById('chartTmMensual')) {{
    const ctxTm = document.getElementById('chartTmMensual').getContext('2d');
    new Chart(ctxTm, {{
      type: 'bar',
      data: {{
        labels: tmM.map(m => m.mesNombre),
        datasets: [
          {{ label: 'Mantención', data: tmM.map(m => m.mant), backgroundColor: '#C0392BCC', stack: 'tm' }},
          {{ label: 'Operacional', data: tmM.map(m => m.oper), backgroundColor: '#E67E22CC', stack: 'tm' }},
          {{ label: 'Proceso', data: tmM.map(m => m.proc), backgroundColor: '#3498DBCC', stack: 'tm' }}
        ]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ position: 'bottom' }},
          tooltip: {{ callbacks: {{ footer: items => 'Perdido: ' + items.reduce((a,b)=>a+b.raw,0).toFixed(0) + ' h' }} }} }},
        scales: {{ x: {{ stacked: true }}, y: {{ stacked: true, beginAtZero: true, title: {{ display: true, text: 'horas' }} }} }}
      }}
    }});
    let th = '<table style="width:100%;border-collapse:collapse;font-size:12px"><thead><tr style="color:#64748B;text-align:right;border-bottom:2px solid #E2E8F0">'
      + '<th style="text-align:left;padding:6px">Mes</th><th>🔧 Mant.</th><th>⚙️ Oper.</th><th>🔄 Proc.</th><th>Perdido</th><th style="color:#94A3B8">⏸️ Prog.</th><th style="text-align:left;padding-left:14px">Top causa</th></tr></thead><tbody>';
    tmM.forEach(m => {{
      const top = (m.topCausas || '').split(';')[0] || '';
      th += '<tr style="text-align:right;border-bottom:1px solid #EEF2F6">'
        + '<td style="text-align:left;padding:6px;font-weight:600">' + m.mesNombre + '</td>'
        + '<td>' + m.mant.toFixed(0) + '</td><td>' + m.oper.toFixed(0) + '</td><td>' + m.proc.toFixed(0) + '</td>'
        + '<td style="font-weight:700">' + m.perdido.toFixed(0) + '</td>'
        + '<td style="color:#94A3B8">' + m.prog.toFixed(0) + '</td>'
        + '<td style="text-align:left;padding-left:14px;color:#475569">' + top + '</td></tr>';
    }});
    th += '</tbody></table>';
    document.getElementById('tmMensualTabla').innerHTML = th;
  }}
}})();

document.getElementById('genDate').textContent = D.generado;

// ═══════════════════════════════════════════════════════════
// ── SEGMENTADOR GLOBAL DE MES ──────────────────────────────
// ═══════════════════════════════════════════════════════════
(() => {{
  const selMes = document.getElementById('selectMes');
  const pillMes = document.getElementById('filtroMesPill');
  const btnClear = document.getElementById('btnLimpiarFiltro');
  const aviso = document.getElementById('avisoMesHistorico');

  const MESES_NOMBRE = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'];
  const mesesHist = [...new Set((D.historico||[]).map(r => r.anio*100 + r.mes))].sort();
  const mesActualKey = D.cfg.anio*100 + D.cfg.mes;
  // Snapshots disponibles (los inyecta el backend)
  const snapshotsDisponibles = D.snapshotsDisponibles || [];

  selMes.options[0].textContent = `Mes en curso (${{D.cfg.mesNombre}} ${{D.cfg.anio}})`;
  mesesHist.forEach(k => {{
    if (k === mesActualKey) return;
    const anio = Math.floor(k/100), mes = k%100;
    const opt = document.createElement('option');
    opt.value = k;
    const snapKey = `${{anio}}-${{String(mes).padStart(2,'0')}}`;
    const tieneSnapshot = snapshotsDisponibles.includes(snapKey);
    opt.textContent = tieneSnapshot
      ? `${{MESES_NOMBRE[mes]}} ${{anio}} (datos completos)`
      : `${{MESES_NOMBRE[mes]}} ${{anio}} (solo resumen)`;
    opt.dataset.tieneSnapshot = tieneSnapshot ? '1' : '0';
    selMes.appendChild(opt);
  }});

  // Pre-seleccionar el mes actual basado en el archivo cargado
  const archivoActual = window.location.pathname.split('/').pop();
  const matchSnap = archivoActual.match(/Dashboard_Cosecha_(\d{{4}})-(\d{{2}})\.html/);
  if (matchSnap) {{
    const k = parseInt(matchSnap[1])*100 + parseInt(matchSnap[2]);
    // Marcar opción correspondiente
    for (const o of selMes.options) {{ if (parseInt(o.value) === k) {{ o.selected = true; break; }} }}
  }}

  // Etiquetar columnas de tablas del Comparativo con su mes
  ['histTablaVol','histTablaCumpl'].forEach(tid => {{
    const cont = document.getElementById(tid);
    if (!cont) return;
    const headers = cont.querySelectorAll('thead th');
    const colMonthMap = {{}};
    headers.forEach((th, idx) => {{
      const txt = th.textContent.trim();
      const m = mesesHist.find(k => {{
        const mn = MESES_NOMBRE[k%100];
        return txt.includes(mn.slice(0,3));
      }});
      if (m) colMonthMap[idx] = m;
    }});
    cont.querySelectorAll('tr').forEach(tr => {{
      tr.querySelectorAll('th, td').forEach((c, idx) => {{
        if (colMonthMap[idx]) c.setAttribute('data-mes', colMonthMap[idx]);
      }});
    }});
  }});

  function aplicar(mesVal) {{
    const filtMes = (mesVal !== '__current__');
    document.querySelectorAll('[data-mes]').forEach(el => {{
      const m = parseInt(el.getAttribute('data-mes'));
      if (filtMes) {{
        if (m === parseInt(mesVal)) {{
          el.style.background = '#DBEAFE';
          el.style.fontWeight = '700';
        }} else {{
          el.style.opacity = '0.35';
          el.style.background = '';
          el.style.fontWeight = '';
        }}
      }} else {{
        el.style.opacity = '';
        el.style.background = '';
        el.style.fontWeight = '';
      }}
    }});
    pillMes.style.display = filtMes ? 'inline-block' : 'none';
    if (filtMes) {{
      const k = parseInt(mesVal);
      pillMes.textContent = `Mes: ${{MESES_NOMBRE[k%100]}} ${{Math.floor(k/100)}}`;
    }}
    aviso.style.display = (filtMes && parseInt(mesVal) !== mesActualKey) ? 'block' : 'none';
    btnClear.style.display = filtMes ? 'inline-block' : 'none';
  }}

  selMes.addEventListener('change', e => {{
    const v = e.target.value;
    // Si es "Mes en curso" → ir al dashboard del mes actual
    if (v === '__current__') {{
      window.location.href = 'index.html';
      return;
    }}
    // Si el snapshot está disponible → navegar al HTML específico de ese mes
    const opt = e.target.selectedOptions[0];
    if (opt && opt.dataset.tieneSnapshot === '1') {{
      const k = parseInt(v);
      const anio = Math.floor(k/100), mes = String(k%100).padStart(2,'0');
      window.location.href = `Dashboard_Cosecha_${{anio}}-${{mes}}.html`;
      return;
    }}
    // Sin snapshot → solo destacar columna en Comparativo (comportamiento anterior)
    aplicar(v);
  }});
  btnClear.addEventListener('click', () => {{
    selMes.value = '__current__';
    window.location.href = 'index.html';
  }});
}})();
{MODAL_JS}
</script>
{MODAL_HTML}
</body>
</html>"""

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"✅ Dashboard HTML generado: {OUTPUT}")
print(f"   Período: {MESES[MES]} {ANIO}")
print(f"   Equipos: {len(TEAMS)} | Días con datos: {DD}/{DM}")
print(f"   Producción total: {round(daily['Vol'].sum(),1)} m³ SSC")

# ── Subida FTP automática (si está configurado) ────────────
if '--ftp' in sys.argv:
    try:
        from SUBIR_FTP import subir_dashboard
        subir_dashboard()
    except ImportError:
        print("⚠️  SUBIR_FTP.py no encontrado. Omitiendo subida FTP.")
    except Exception as e:
        print(f"⚠️  Error al subir por FTP: {e}")
