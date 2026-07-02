#!/usr/bin/env python3
"""
archivar_mes_anterior.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Guarda el cierre del MES ANTERIOR en historico_cierres_mensuales.csv.

Pensado para la NUBE (GitHub Actions, sin disco persistente): como cada
corrida arranca de cero y baja solo el mes en curso, el mes que recién
cerró se perdería. Este script lo recupera bajándolo directo de Arauco.

Idempotente: si el mes anterior YA está en el histórico, no hace nada
(termina con código 0). Así puede correr todos los días sin duplicar.

Uso:  python3 archivar_mes_anterior.py
Requiere: ARAUCO_USER / ARAUCO_PASS (o .noc_config.json), pandas, requests.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
import os, sys, csv, calendar
from pathlib import Path
from datetime import datetime
import importlib.util
import pandas as pd
import urllib3
urllib3.disable_warnings()

BASE_DIR = Path(__file__).parent
HIST_CSV = BASE_DIR / "historico_cierres_mensuales.csv"
sys.path.insert(0, str(BASE_DIR))
from normalizar_produccion import normalizar  # noqa: E402

TEAM_MAP = {
    'S123':'Millalemu 1.1','S58':'Millalemu 1.2','S223':'Millalemu 1.3',
    'S246':'Millalemu 1.4','MG5':'Millalemu 5','TEA02':'Millalemu 7',
    'TEA08':'Millalemu 9','T125':'Millalemu 11','TEA30':'Millalemu 1.3',
}
ORDEN = ['Millalemu 1.1','Millalemu 1.2','Millalemu 1.3','Millalemu 1.4',
         'Millalemu 5','Millalemu 7','Millalemu 9','Millalemu 11']
METAS_DEFAULT = {'Millalemu 1.1':7000,'Millalemu 1.2':7000,'Millalemu 1.3':7000,
                 'Millalemu 1.4':7000,'Millalemu 5':7000,'Millalemu 7':7000,
                 'Millalemu 9':7000,'Millalemu 11':6000}
MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
         "Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
TM_HIST_CSV = BASE_DIR / "historico_tm_mensual.csv"
CLASIF = {
    1:'Mantención',2:'Mantención',3:'Mantención',4:'Mantención',5:'Mantención',
    6:'Mantención',7:'Mantención',8:'Mantención',10:'Mantención',12:'Mantención',
    58:'Mantención',69:'Mantención',
    13:'Operacional',14:'Operacional',15:'Operacional',20:'Operacional',21:'Operacional',
    22:'Operacional',31:'Operacional',32:'Operacional',33:'Operacional',38:'Operacional',
    41:'Operacional',
    16:'Proceso',17:'Proceso',18:'Proceso',25:'Proceso',26:'Proceso',61:'Proceso',
    65:'Proceso',66:'Proceso',68:'Proceso',
    42:'Programado',43:'Programado',
}

def log(m): print(f"[archivar_mes_anterior] {m}")

def metas_desde_excel():
    """Lee metas de la hoja CONFIGURACIÓN si el Excel existe; si no, defaults."""
    metas = dict(METAS_DEFAULT)
    excel = BASE_DIR / "Dashboard_CosechaForestal.xlsx"
    if excel.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(excel), data_only=True)
            if "CONFIGURACIÓN" in wb.sheetnames:
                ws = wb["CONFIGURACIÓN"]
                for i, t in enumerate(ORDEN):
                    v = ws.cell(24 + i, 5).value
                    if v:
                        metas[t] = float(v)
            wb.close()
        except Exception as e:
            log(f"aviso: no se pudo leer metas del Excel ({e}); uso defaults")
    return metas

def mes_ya_archivado(mes, anio):
    if not HIST_CSV.exists():
        return False
    with open(HIST_CSV, encoding='utf-8-sig', newline='') as f:
        for row in csv.reader(f, delimiter=';'):
            if len(row) >= 2 and row[0].strip() == str(mes) and row[1].strip() == str(anio):
                return True
    return False

def descargar_mes(mes, anio):
    """Baja BN (Base2NOC) del mes desde Arauco y devuelve el DataFrame crudo.
    cierre() lo normaliza con normalizar(), que espera formato BN."""
    spec = importlib.util.spec_from_file_location("dnoc", str(BASE_DIR / "descargar_noc_api.py"))
    dn = importlib.util.module_from_spec(spec); spec.loader.exec_module(dn)
    import requests
    s = requests.Session(); s.verify = False
    s.headers.update({"User-Agent": "Mozilla/5.0"})
    token = dn.obtener_token_arcgis(s)
    ult_dia = calendar.monthrange(anio, mes)[1]
    fi, ff = f"{anio}-{mes:02d}-01", f"{anio}-{mes:02d}-{ult_dia:02d}"
    log(f"descargando {MESES[mes]} {anio} ({fi} → {ff})...")
    pg = dn.descargar_reporte(s, token, "BN", fi, ff)
    if not pg:
        return None
    dn.guardar_csv(pg, "BN", Path("/tmp"))
    return pd.read_csv("/tmp/Base2NOC.csv", sep=';', encoding='utf-8-sig')

def cierre(prod, mes, anio, metas):
    prod = normalizar(prod)
    for c in ['Volumen SSC PU', 'Volumen SSC AS']:
        prod[c] = pd.to_numeric(prod[c].astype(str).str.replace(',', '.'), errors='coerce')
    prod['Vol'] = prod['Volumen SSC PU'].fillna(0) + prod['Volumen SSC AS'].fillna(0)
    _te = pd.to_numeric(prod['Tiempo Efectivo'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
    _te_pos = _te[_te > 0]  # unidad robusta por mediana (no por máximo)
    prod['HrsEf'] = _te / (3600 if (len(_te_pos) > 0 and _te_pos.median() > 1000) else 60)
    prod['Team'] = prod['Equipo'].map(TEAM_MAP)
    prod['Fecha_dt'] = pd.to_datetime(prod['Fecha NOC'], dayfirst=True, errors='coerce')
    prod['Dia'] = prod['Fecha_dt'].dt.day
    prod = prod[prod['Fecha_dt'].dt.month == mes]
    fuente = f"Auto-archivado nube {datetime.now().strftime('%Y-%m-%d')}"
    # Días hábiles del mes cerrado: días corridos menos feriados irrenunciables
    # (regla: faltar/parar no premia; un día de falla cuenta como día trabajado).
    _FER = {'01-01', '05-01', '09-18', '09-19', '12-25'}
    _dm = calendar.monthrange(anio, mes)[1]
    dias_hab = sum(1 for d in range(1, _dm + 1) if f"{mes:02d}-{d:02d}" not in _FER)
    filas = []
    for t in ORDEN:
        td = prod[prod['Team'] == t]
        vol = float(td['Vol'].sum()); hrs = float(td['HrsEf'].sum())
        dias = dias_hab; meta = metas[t]
        cumpl = vol / meta * 100 if meta else 0
        prom = vol / dias if dias else 0
        rend = vol / hrs if hrs else 0
        filas.append(f"{mes};{anio};{MESES[mes]};{t};{meta:.0f};{vol:.1f};{cumpl:.1f};{dias};{dias};{prom:.1f};{rend:.1f};{hrs:.1f};{fuente}")
    return filas

def tm_ya_archivado(mes, anio):
    if not TM_HIST_CSV.exists():
        return False
    with open(TM_HIST_CSV, encoding='utf-8-sig', newline='') as f:
        for row in csv.reader(f, delimiter=';'):
            if len(row) >= 2 and row[0].strip() == str(mes) and row[1].strip() == str(anio):
                return True
    return False

def descargar_tp(mes, anio):
    """Baja Tiempos Perdidos del mes desde Arauco."""
    spec = importlib.util.spec_from_file_location("dnoc", str(BASE_DIR / "descargar_noc_api.py"))
    dn = importlib.util.module_from_spec(spec); spec.loader.exec_module(dn)
    import requests
    s = requests.Session(); s.verify = False
    s.headers.update({"User-Agent": "Mozilla/5.0"})
    token = dn.obtener_token_arcgis(s)
    ud = calendar.monthrange(anio, mes)[1]
    tp = dn.descargar_reporte(s, token, "TP", f"{anio}-{mes:02d}-01", f"{anio}-{mes:02d}-{ud:02d}")
    if not tp:
        return None
    dn.guardar_csv(tp, "TP", Path("/tmp"))
    return pd.read_csv("/tmp/TiemposPerdidos.csv", sep=';', encoding='utf-8-sig')

def archivar_tm(mes, anio):
    """Calcula y guarda los tiempos perdidos por categoría + top causas del mes."""
    if tm_ya_archivado(mes, anio):
        log(f"TM de {MESES[mes]} {anio} ya está — nada que hacer.")
        return
    tp = descargar_tp(mes, anio)
    if tp is None or tp.empty:
        log(f"sin TM de {MESES[mes]} {anio} en Arauco — no se archiva.")
        return
    tp['min'] = pd.to_numeric(tp['Tiempo (Min)'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
    tp['cod'] = pd.to_numeric(tp['Código Tiempo Perdido'], errors='coerce').fillna(0).astype(int)
    tp['cat'] = tp['cod'].map(CLASIF).fillna('Operacional')
    by = tp.groupby('cat')['min'].sum() / 60
    ma, op, pr, pg = by.get('Mantención', 0), by.get('Operacional', 0), by.get('Proceso', 0), by.get('Programado', 0)
    perd = ma + op + pr
    tcp = tp[tp['cat'] != 'Programado'].groupby('Descripción')['min'].sum().sort_values(ascending=False).head(3)
    top = "; ".join(f"{str(k)[:28]} ({v/60:.0f}h)" for k, v in tcp.items())
    # top_causas lleva ';' internos → SIEMPRE entre comillas (con escape) o rompe
    # el read_csv(sep=';') que arma tm_mensual_list en GENERAR_HTML.
    top_q = '"' + top.replace('"', '""') + '"'
    nuevo = not TM_HIST_CSV.exists()
    with open(TM_HIST_CSV, 'a', encoding='utf-8') as f:
        if nuevo:
            f.write("mes;anio;mes_nombre;mantencion_h;operacional_h;proceso_h;programado_h;total_perdido_h;top_causas\n")
        f.write(f"{mes};{anio};{MESES[mes]};{ma:.1f};{op:.1f};{pr:.1f};{pg:.1f};{perd:.1f};{top_q}\n")
    log(f"✅ TM {MESES[mes]} {anio} archivado: perdido {perd:.0f}h (Mant {ma:.0f} / Oper {op:.0f} / Proc {pr:.0f})")

def main():
    hoy = datetime.now()
    primer_dia = hoy.replace(day=1)
    prev = primer_dia.fromordinal(primer_dia.toordinal() - 1)  # último día mes anterior
    mes, anio = prev.month, prev.year

    # ── Producción ──
    if mes_ya_archivado(mes, anio):
        log(f"Producción de {MESES[mes]} {anio} ya está.")
    else:
        log(f"Recuperando producción de {MESES[mes]} {anio} de Arauco...")
        prod = descargar_mes(mes, anio)
        if prod is None or prod.empty:
            log(f"sin datos de producción de {MESES[mes]} {anio}.")
        else:
            filas = cierre(prod, mes, anio, metas_desde_excel())
            total = sum(float(r.split(';')[5]) for r in filas)
            with open(HIST_CSV, 'a', encoding='utf-8') as f:
                f.write("\n".join(filas) + "\n")
            log(f"✅ Producción {MESES[mes]} {anio}: {len(filas)} faenas, total {total:,.0f} m³")

    # ── Tiempos perdidos ──
    try:
        archivar_tm(mes, anio)
    except Exception as e:
        log(f"aviso: archivar_tm falló ({e}) — no crítico.")

    return 0

if __name__ == "__main__":
    sys.exit(main())
