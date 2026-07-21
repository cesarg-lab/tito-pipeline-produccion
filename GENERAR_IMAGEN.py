"""
GENERAR_IMAGEN.py — v4.0 (2026-04-16)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Genera imagen PNG de la Tabla de Producciones con números
GRANDES legibles en WhatsApp celular sin hacer zoom.

Requisitos: Python 3 con pandas, matplotlib
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import calendar
import os, sys
from datetime import datetime, date as dt_date
from pathlib import Path

BASE_DIR = Path(__file__).parent
CSV_PROD = BASE_DIR / "Base2NOC.csv"
CSV_TM   = BASE_DIR / "TiemposPerdidos.csv"
sys.path.insert(0, str(BASE_DIR))
from normalizar_produccion import normalizar  # noqa: E402

TEAM_MAP = {
    'S123':'Millalemu 1.1','S58':'Millalemu 1.2','S223':'Millalemu 1.3',
    'S246':'Millalemu 1.4','MG5':'Millalemu 5','TEA02':'Millalemu 7',
    'TEA08':'Millalemu 9','T125':'Millalemu 11','TEA30':'Millalemu 1.3',
}
# Configuración de grupos
GRUPOS = {
    'general':   {'teams': ['Millalemu 1.1','Millalemu 1.2','Millalemu 1.3','Millalemu 1.4',
                           'Millalemu 5','Millalemu 7','Millalemu 9','Millalemu 11'],
                  'titulo': 'Control de Cosecha Forestal',
                  'output': 'grilla_produccion.png'},
    'aereo':     {'teams': ['Millalemu 5','Millalemu 7','Millalemu 9','Millalemu 11'],
                  'titulo': 'Millalemu Aéreo',
                  'output': 'grilla_produccion_aereo.png'},
    'terrestre': {'teams': ['Millalemu 1.1','Millalemu 1.2','Millalemu 1.3','Millalemu 1.4'],
                  'titulo': 'Millalemu Terrestre',
                  'output': 'grilla_produccion_terrestre.png'},
}

# Leer argumento de línea de comandos: --grupo aereo | terrestre | general
GRUPO = 'general'
for i, a in enumerate(sys.argv):
    if a == '--grupo' and i+1 < len(sys.argv):
        g = sys.argv[i+1].lower()
        if g in GRUPOS: GRUPO = g
TEAMS   = GRUPOS[GRUPO]['teams']
TITULO  = GRUPOS[GRUPO]['titulo']
OUTPUT  = BASE_DIR / GRUPOS[GRUPO]['output']
ABREV = {t: t.replace('Millalemu ', 'M') for t in TEAMS}

CLASIF = {
    # Mantención — falla / reparación / mantención de equipos
    1:'Mantención',2:'Mantención',3:'Mantención',4:'Mantención',5:'Mantención',
    6:'Mantención',7:'Mantención',8:'Mantención',10:'Mantención',12:'Mantención',
    58:'Mantención',69:'Mantención',
    # Operacional — logística / personal / abastecimiento / externo
    13:'Operacional',14:'Operacional',15:'Operacional',20:'Operacional',21:'Operacional',
    22:'Operacional',31:'Operacional',32:'Operacional',33:'Operacional',38:'Operacional',
    41:'Operacional',
    # Proceso — flujo interno de la faena
    16:'Proceso',17:'Proceso',18:'Proceso',25:'Proceso',26:'Proceso',61:'Proceso',
    65:'Proceso',66:'Proceso',68:'Proceso',
    # Programado — NO es pérdida (se muestra aparte, no suma al tiempo perdido)
    42:'Programado',43:'Programado',
}

METAS_DEFAULT = {
    'Millalemu 1.1': 7000, 'Millalemu 1.2': 7000, 'Millalemu 1.3': 7000,
    'Millalemu 1.4': 7000, 'Millalemu 5': 7000, 'Millalemu 7': 7000,
    'Millalemu 9': 7000, 'Millalemu 11': 6000
}
METAS = dict(METAS_DEFAULT)

# Orden canónico de equipos en el Excel (siempre son 8, índices fijos)
_TEAMS_ORIGINAL = ['Millalemu 1.1','Millalemu 1.2','Millalemu 1.3','Millalemu 1.4',
                   'Millalemu 5','Millalemu 7','Millalemu 9','Millalemu 11']
EXCEL = BASE_DIR / "Dashboard_CosechaForestal.xlsx"
if EXCEL.exists():
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(EXCEL), data_only=True)
        if "CONFIGURACIÓN" in wb.sheetnames:
            ws = wb["CONFIGURACIÓN"]
            for i, t in enumerate(_TEAMS_ORIGINAL):
                v = ws.cell(24 + i, 5).value
                if v: METAS[t] = float(v)
        wb.close()
    except Exception:
        pass

MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
DIAS_SEMANA = ['Lu','Ma','Mi','Ju','Vi','Sa','Do']

def fmt(n):
    if n is None or (isinstance(n, float) and np.isnan(n)):
        return '—'
    return f"{n:,.0f}".replace(',', '.')

def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}")

def generate():
    log(f"🖼️  Generando imagen de grilla — grupo: {GRUPO.upper()} ({len(TEAMS)} equipos)...")

    prod = pd.read_csv(CSV_PROD, sep=';', encoding='utf-8-sig')
    prod = normalizar(prod)
    tm   = pd.read_csv(CSV_TM,   sep=';', encoding='utf-8-sig')

    for c in ['Volumen SSC PU','Volumen SSC AS']:
        prod[c] = pd.to_numeric(prod[c].astype(str).str.replace(',','.'), errors='coerce')
    prod['Vol'] = prod['Volumen SSC PU'].fillna(0) + prod['Volumen SSC AS'].fillna(0)

    # ── Manual.csv DESACTIVADO (Cesar 2026-05-02) — Selenium fuente única ───
    from pathlib import Path
    MANUAL_CSV = Path("/__manual_desactivado__")  # path imposible
    vol_oficial_diario = None
    if MANUAL_CSV.exists():
        try:
            mdf = pd.read_csv(MANUAL_CSV, sep=';', encoding='utf-8-sig', decimal=',', thousands='.')
            vol_oficial_diario = {}
            for _, r in mdf.iterrows():
                eq_full = str(r.get('Equipo','')).strip()
                if '-' not in eq_full: continue
                sigla = eq_full.split('-')[0]
                team = TEAM_MAP.get(sigla)
                if not team: continue
                for d in range(1, 32):
                    v = r.get(str(d))
                    if pd.notna(v) and v != 0:
                        vol_oficial_diario[(team, int(d))] = float(v)
            log(f"   📋 Manual.csv cargado: {len(vol_oficial_diario)} día/equipo oficiales")
        except Exception as e:
            log(f"   ⚠️  No se pudo leer Manual.csv: {e}")
            vol_oficial_diario = None
    # Tiempo Efectivo viene en MINUTOS → dividir por 60 para horas
    # Tiempo Efectivo: detección automática de unidad (Selenium=segundos, API=minutos)
    _te_raw = pd.to_numeric(prod['Tiempo Efectivo'].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    _te_pos = _te_raw[_te_raw > 0]  # unidad robusta por mediana (no por máximo)
    prod['HrsEf'] = _te_raw / (3600 if (len(_te_pos) > 0 and _te_pos.median() > 1000) else 60)
    # Turno en segundos (Hora Fin - Hora Inicio) para cálculo de disponibilidad
    for c in ['Hora Inicio', 'Hora Fin']:
        prod[c] = pd.to_numeric(prod[c].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    prod['Turno_seg'] = (prod['Hora Fin'] - prod['Hora Inicio']).clip(lower=0)
    prod['Team'] = prod['Equipo'].map(TEAM_MAP)
    prod['Fecha_dt'] = pd.to_datetime(prod['Fecha NOC'], dayfirst=True, errors='coerce')
    prod['Dia'] = prod['Fecha_dt'].dt.day

    tm['Tiempo (Min)'] = pd.to_numeric(tm['Tiempo (Min)'].astype(str).str.replace(',','.'), errors='coerce').fillna(0)
    tm['Clasif'] = tm['Código Tiempo Perdido'].map(CLASIF).fillna('Operacional')
    tm = tm[tm['Clasif'] != 'Programado'].copy()  # colación/descanso no es pérdida
    tm['Team'] = tm['Código Equipo'].map(TEAM_MAP)
    tm['Fecha_dt'] = pd.to_datetime(tm['Fecha'], dayfirst=True, errors='coerce')
    tm = tm.dropna(subset=['Fecha_dt'])
    tm['Dia'] = tm['Fecha_dt'].dt.day

    MES  = int(prod['Fecha_dt'].dt.month.mode()[0])
    ANIO = int(prod['Fecha_dt'].dt.year.mode()[0])
    prod = prod[prod['Fecha_dt'].dt.month == MES]
    tm   = tm[tm['Fecha_dt'].dt.month == MES]

    DM = calendar.monthrange(ANIO, MES)[1]
    # Días hábiles (regla: faltar no premia; solo descuentan feriados irrenunciables)
    _FERIADOS_IRR = {'01-01', '05-01', '09-18', '09-19', '12-25'}
    _ULT = int(prod['Dia'].max())
    DT = sum(1 for d in range(1, DM + 1) if f"{MES:02d}-{d:02d}" not in _FERIADOS_IRR)
    DD = sum(1 for d in range(1, _ULT + 1) if f"{MES:02d}-{d:02d}" not in _FERIADOS_IRR)
    DR = max(DT - DD, 1)

    # Vol suma por fila (cada producto del folio aporta m³ distinto), pero HrsEf y
    # Turno_seg son valores del FOLIO (se repiten en cada fila-producto). Sumarlos por
    # fila los multiplica por N productos → infla las horas y SUBESTIMA el m³/hora.
    # Fix: tomar 1 valor por folio antes de agregar (igual que GENERAR_RESUMEN).
    if 'Número Noc' in prod.columns:
        _vol = prod.groupby(['Dia','Team']).agg(Vol=('Vol','sum')).reset_index()
        _ht = (prod.groupby(['Dia','Team','Número Noc'])
                   .agg(HrsEf=('HrsEf','first'), Turno_seg=('Turno_seg','first')).reset_index()
                   .groupby(['Dia','Team'])
                   .agg(HrsEf=('HrsEf','sum'), Turno_seg=('Turno_seg','sum')).reset_index())
        daily = _vol.merge(_ht, on=['Dia','Team'], how='left')
        daily[['HrsEf','Turno_seg']] = daily[['HrsEf','Turno_seg']].fillna(0)
    else:
        daily = prod.groupby(['Dia','Team']).agg(
            Vol=('Vol','sum'), HrsEf=('HrsEf','sum'), Turno_seg=('Turno_seg','sum')
        ).reset_index()
    # Filtrar por grupo seleccionado
    daily = daily[daily['Team'].isin(TEAMS)].reset_index(drop=True)
    # También filtrar vol_oficial_diario para que agregue filas solo de los equipos del grupo
    if vol_oficial_diario is not None:
        vol_oficial_diario = {(t, d): v for (t, d), v in vol_oficial_diario.items() if t in TEAMS}

    # Si Manual existe, usarlo como cruce: Manual donde haya valor, Base2NOC donde no.
    # Antes se descartaban los días no cubiertos por Manual; ahora se conservan
    # porque Base2NOC vía Selenium ya viene depurado.
    if vol_oficial_diario is not None:
        import pandas as _pd
        # Sobrescribir Vol: Manual si está, Base2NOC si no
        daily['Vol'] = daily.apply(
            lambda r: vol_oficial_diario.get((r['Team'], int(r['Dia'])), r['Vol']), axis=1
        )
        # Quitar filas con Vol=0 y sin horas (fantasmas depurados por Manual)
        mask_quitar = (daily['Vol'] == 0) & (daily['HrsEf'] == 0)
        daily = daily[~mask_quitar].reset_index(drop=True)
        # Agregar filas del Manual que no están en daily (reasignados por GeoNOC)
        idx_existentes = set((r['Team'], int(r['Dia'])) for _, r in daily.iterrows())
        filas_nuevas = []
        for (t, d), v in vol_oficial_diario.items():
            if (t, d) not in idx_existentes:
                filas_nuevas.append({'Dia': d, 'Team': t, 'Vol': v, 'HrsEf': 0, 'Turno_seg': 0})
        if filas_nuevas:
            daily = _pd.concat([daily, _pd.DataFrame(filas_nuevas)], ignore_index=True)

    piv = tm.groupby(['Dia','Team','Clasif'])['Tiempo (Min)'].sum().unstack(fill_value=0).reset_index()
    piv.columns.name = None
    for c in ['Mantención','Operacional','Proceso']:
        if c not in piv.columns: piv[c] = 0
    piv = piv.rename(columns={'Mantención':'TM_Mant','Operacional':'TM_Oper','Proceso':'TM_PP'})
    daily = daily.merge(piv[['Dia','Team','TM_Mant','TM_Oper','TM_PP']], on=['Dia','Team'], how='left')
    daily['TM_Total'] = daily['TM_Mant'].fillna(0) + daily['TM_Oper'].fillna(0) + daily['TM_PP'].fillna(0)
    for c in ['TM_Mant','TM_Oper','TM_PP']:
        daily[c] = daily[c].fillna(0)

    # KPIs por equipo
    team_data = {}
    total_acum = 0; total_hrs = 0; total_tm_mant = 0
    meta_total = sum(METAS[t] for t in TEAMS)

    # Mantención por equipo desde registros CRUDOS (no la tabla cruzada con
    # producción, que pierde fallas de máquinas paradas todo el día → disp inflada).
    _tm_mant_team = tm[tm['Clasif'] == 'Mantención'].groupby('Team')['Tiempo (Min)'].sum()

    total_turno_seg = 0
    for t in TEAMS:
        td = daily[daily['Team'] == t]
        acum = td['Vol'].sum(); hrs = td['HrsEf'].sum()
        tm_mant = float(_tm_mant_team.get(t, 0))
        turno_seg = td['Turno_seg'].sum()
        dias_t = DD; meta = METAS[t]  # días hábiles (regla: faltar no premia)
        prom = acum / dias_t if dias_t > 0 else 0
        proy = acum + prom * DR
        plan_dia = meta / DT
        avance_plan = plan_dia * DD
        dif_plan = acum - avance_plan
        ritmo = (meta - acum) / max(DR, 1)
        rendimiento = acum / hrs if hrs > 0 else 0
        # Disponibilidad = (1 - TM_Mantención / Turno) × 100
        # tm_mant en minutos, turno_seg en segundos → ambos a minutos
        turno_min = turno_seg / 60  # segundos → minutos
        disp = (1 - tm_mant / turno_min) * 100 if turno_min > 0 else 100
        disp = max(0, min(100, disp))
        total_acum += acum; total_hrs += hrs; total_tm_mant += tm_mant; total_turno_seg += turno_seg
        team_data[t] = {
            'meta': meta, 'plan_dia': plan_dia, 'acum': acum,
            'avance_plan': avance_plan, 'dif_plan': dif_plan,
            'proy': proy, 'ritmo': ritmo, 'rendimiento': rendimiento,
            'tm_mant_hrs': tm_mant / 60, 'disp': disp, 'prom': prom,
        }

    prom_total = total_acum / DD if DD > 0 else 0
    proy_total = total_acum + prom_total * DR
    plan_dia_total = meta_total / DT
    avance_plan_total = plan_dia_total * DD
    dif_total = total_acum - avance_plan_total
    ritmo_total = (meta_total - total_acum) / max(DR, 1)
    rend_total = total_acum / total_hrs if total_hrs > 0 else 0
    turno_min_total = total_turno_seg / 60
    disp_total = (1 - total_tm_mant / turno_min_total) * 100 if turno_min_total > 0 else 100
    disp_total = max(0, min(100, disp_total))

    pct_total = round(proy_total / meta_total * 100, 1) if meta_total > 0 else 0

    # ── Tendencia vs día anterior ─────────────────────────────────────
    # Recalcular % Proy excluyendo el último día para compararlo con el actual
    if DD >= 2:
        ultimo_dia_calc = int(daily['Dia'].max())
        daily_ayer = daily[daily['Dia'] < ultimo_dia_calc]
        DD_ayer = daily_ayer['Dia'].nunique() if len(daily_ayer) > 0 else DD - 1
        DR_ayer = max(DT - DD_ayer, 1)
        acum_ayer = daily_ayer['Vol'].sum()
        prom_ayer = acum_ayer / DD_ayer if DD_ayer > 0 else 0
        proy_ayer = acum_ayer + prom_ayer * DR_ayer
        pct_ayer = round(proy_ayer / meta_total * 100, 1) if meta_total > 0 else 0
        delta_pct = pct_total - pct_ayer
    else:
        delta_pct = 0

    grid = {}; grid_tm = {}
    for _, r in daily.iterrows():
        d = int(r['Dia']); t = r['Team']
        grid.setdefault(d, {})[t] = round(float(r['Vol']), 1)
        grid_tm.setdefault(d, {})[t] = round(float(r['TM_Total']), 0)  # TM total (Mant+Oper+Proceso)

    # ════════════════════════════════════════════════════════════
    # DIBUJAR — Fuentes MUY GRANDES para celular
    # ════════════════════════════════════════════════════════════
    n_teams = len(TEAMS)
    n_kpi = 10
    # Mostrar hasta el último día con datos en Base2NOC (Selenium ya depura).
    # Manual.csv solo se usa como cruce para depurar fantasmas, no para limitar el rango.
    ultimo_dia = int(prod['Dia'].max())
    dias = list(range(1, ultimo_dia + 1))
    n_dias = len(dias)

    # Medidas en pulgadas × DPI = pixels
    # DPI alto para nitidez en pantalla retina celular
    DPI = 288

    W_DIA     = 0.7
    W_TEAM    = 1.0
    W_TOTAL   = 1.1
    ROW_H     = 0.38          # altura filas diarias
    ROW_KPI_H = 0.28          # altura filas KPI (más compacta)
    HDR_H   = 0.42
    TITLE_H = 1.00   # un poco más alto para acomodar el % Cierre Proy. grande
    SUMMARY_H = 0.0  # panel de proyección eliminado (ahora se muestra como fila dentro del bloque KPI)

    # Ancho razonable según cantidad de equipos (mínimo para que quepan KPIs + título)
    fig_w_teams = W_DIA + n_teams * W_TEAM + W_TOTAL
    MIN_FIG_W = 7.5  # mínimo para acomodar título + KPIs
    fig_w = max(fig_w_teams, MIN_FIG_W)
    if fig_w > fig_w_teams:
        extra = fig_w - fig_w_teams
        factor = (n_teams * W_TEAM + W_TOTAL + extra) / (n_teams * W_TEAM + W_TOTAL)
        W_TEAM = W_TEAM * factor
        W_TOTAL = W_TOTAL * factor
    # Factor para escalar KPIs del header según ancho disponible
    kpi_scale = min(1.0, fig_w / 9.8)   # 9.8 era el ancho original (8 equipos)
    # Bloque KPI: 8 filas normales + 1 grande (Acumulado) — el tamaño extra lo compensa el loop
    fig_h = TITLE_H + SUMMARY_H + HDR_H + (n_kpi * ROW_KPI_H) + ROW_KPI_H * 0.5 + ((n_dias + 1) * ROW_H) + 0.30

    # fonts relativos al DPI — estos quedan GRANDES
    F_TITLE  = 15      # título principal
    F_SUB    = 8       # subtítulo
    F_HDR    = 11      # header columnas
    F_NUM    = 11      # números en celdas
    F_KPI_L  = 7.5     # labels KPI
    F_KPI_V  = 10      # valores KPI
    F_TM     = 6       # TM hints
    F_DAY    = 10      # número de día
    F_DOW    = 7       # día de semana
    F_FOOTER = 7

    fig, ax = plt.subplots(1, 1, figsize=(fig_w, fig_h), dpi=DPI)
    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, fig_h)
    ax.axis('off')
    fig.patch.set_facecolor('#FFFFFF')

    PRIMARY = '#1A5276'; PRIMARY_DK = '#154360'
    RED = '#C0392B'; TXT = '#1E293B'; MUTED = '#94A3B8'; BORDER = '#CBD5E1'

    def rect(x, y, w, h, c):
        ax.add_patch(plt.Rectangle((x, y), w, h, fc=c, ec='none'))

    def t(x, y, s, fs=F_NUM, c=TXT, ha='right', bold=False):
        ax.text(x, y, str(s), fontsize=fs, color=c, ha=ha, va='center',
                fontweight='bold' if bold else 'normal', fontfamily='sans-serif')

    def hline(y, c=BORDER, lw=0.5):
        ax.plot([0, fig_w], [y, y], color=c, lw=lw, clip_on=False)

    def col_x(i):
        if i == 0: return 0
        if i <= n_teams: return W_DIA + (i-1) * W_TEAM
        return W_DIA + n_teams * W_TEAM

    def col_w(i):
        if i == 0: return W_DIA
        if i <= n_teams: return W_TEAM
        return W_TOTAL

    P = 0.06  # padding

    # ── BANNER ENCABEZADO (estilo dashboard web) ──
    y = fig_h
    rect(0, y - TITLE_H, fig_w, TITLE_H, PRIMARY)

    # --- Lado izquierdo: título + subtítulo + fecha ---
    # Si el título es largo, reducir tamaño para que quepa
    fs_tit = F_TITLE - 3 if len(TITULO) > 30 else F_TITLE
    t(P + 0.02, y - 0.22,
      TITULO, fs=fs_tit, c='white', ha='left', bold=True)
    t(P + 0.02, y - 0.48,
      f"Forestal Millalemu · {MESES[MES]} {ANIO} · {DD}/{DM} días",
      fs=F_SUB + 1, c='#B0C4DE', ha='left')
    t(P + 0.02, y - 0.68,
      f"Datos al: {datetime.now().strftime('%d-%b-%Y')}", fs=F_SUB, c='#7B9DBF', ha='left')

    # --- Lado derecho: KPIs grandes ---
    brecha = proy_total - meta_total
    # Color condicional para Cierre Proy. (mismos umbrales que dashboard HTML)
    if pct_total >= 80:
        cierre_c = '#A3E635'   # verde
    elif pct_total >= 60:
        cierre_c = '#FDE68A'   # amarillo
    else:
        cierre_c = '#FCA5A5'   # rojo
    # Cada box: (label, valor, unidad, color_valor, sub_texto, ancho, fs_valor)
    kpi_boxes = [
        ('Acumulado',  f"{total_acum:,.0f}".replace(',','.'), 'm³ SSC', 'white', None, 1.00*kpi_scale, int(12*kpi_scale)),
        ('Meta',       f"{meta_total:,.0f}".replace(',','.'), 'm³ SSC', 'white', None, 0.95*kpi_scale, int(12*kpi_scale)),
        ('Cierre Proy.', f"{pct_total:.1f}%", None, cierre_c, None, 2.10*kpi_scale, int(28*kpi_scale)),
        ('Proyección', f"{proy_total:,.0f}".replace(',','.'), None, 'white',
         f"Brecha: {brecha:+,.0f}".replace(',','.'), 1.85*kpi_scale, int(22*kpi_scale)),
    ]
    total_w = sum(b[5] for b in kpi_boxes)
    kpi_start_x = fig_w - total_w - 0.10

    bx = kpi_start_x
    for bi, (lbl, val, unit, val_c, sub, box_w, fs_val) in enumerate(kpi_boxes):
        is_highlight = lbl in ('Cierre Proy.', 'Proyección')
        if bi > 0:
            ax.plot([bx, bx], [y - 0.10, y - TITLE_H + 0.06],
                    color='#2E6B8F', lw=0.8, clip_on=False)
        lbl_fs = 8 if is_highlight else 6.5
        lbl_y = y - 0.18 if is_highlight else y - 0.14
        t(bx + box_w/2, lbl_y, lbl, fs=lbl_fs, c='#B0CBE0' if is_highlight else '#7B9DBF', ha='center', bold=is_highlight)
        val_y = y - 0.52 if is_highlight else y - 0.38
        t(bx + box_w/2, val_y, val, fs=fs_val, c=val_c, ha='center', bold=True)
        # Mini-indicador de tendencia DEBAJO del % Cierre Proy. (en fila inferior, separado)
        if lbl == 'Cierre Proy.' and abs(delta_pct) >= 0.05:
            arrow = '▲' if delta_pct > 0 else '▼'
            trend_c = '#4ADE80' if delta_pct > 0 else '#F87171'
            trend_txt = f"{arrow} {abs(delta_pct):.1f} pts vs ayer"
            t(bx + box_w/2, y - 0.77, trend_txt, fs=7.5, c=trend_c, ha='center', bold=True)
        if unit:
            t(bx + box_w/2, y - 0.52, unit, fs=6, c='#7B9DBF', ha='center')
        if sub:
            sub_c = RED if brecha < 0 else '#10B981'
            sub_y = y - 0.77 if is_highlight else y - 0.52
            t(bx + box_w/2, sub_y, sub, fs=7, c=sub_c, ha='center', bold=True)
        bx += box_w

    # Panel de Proyección Cierre Mes ELIMINADO — ahora se muestra como fila dentro del bloque KPI
    y_sum = y - TITLE_H

    # ── HEADER ──
    y_hdr = y_sum - SUMMARY_H - HDR_H
    rect(0, y_hdr, fig_w, HDR_H, PRIMARY_DK)
    yc = y_hdr + HDR_H/2

    t(col_x(0) + col_w(0)/2, yc, "Día", fs=F_HDR-2, c='white', ha='center', bold=True)
    for i, team in enumerate(TEAMS):
        t(col_x(i+1) + col_w(i+1)/2, yc, ABREV[team], fs=F_HDR, c='white', ha='center', bold=True)
    t(col_x(n_teams+1) + col_w(n_teams+1)/2, yc, "Total", fs=F_HDR, c='white', ha='center', bold=True)

    y_top_grid = y_hdr + HDR_H

    # ── KPI ROWS ──
    # (label, val_fn, total_val, bg, color_fn, bold, total_color)
    kpi_defs = [
        ('% Proy',     lambda tm: f"{(team_data[tm]['proy']/METAS[tm]*100):.0f}%", f"{pct_total:.0f}%", '#EBF0F5', lambda tm: '#059669' if (team_data[tm]['proy']/METAS[tm]*100)>=80 else '#D97706' if (team_data[tm]['proy']/METAS[tm]*100)>=60 else RED, True, '#059669' if pct_total>=80 else '#D97706' if pct_total>=60 else RED),
        ('Meta',       lambda tm: fmt(METAS[tm]),                         fmt(meta_total),           '#F5F7FA', None, True, TXT),
        ('Acumulado',  lambda tm: fmt(team_data[tm]['acum']),             fmt(total_acum),           '#EBF0F5',
            lambda tm: '#059669' if team_data[tm]['avance_plan']>0 and team_data[tm]['acum']/team_data[tm]['avance_plan']>=1.0 else '#D97706' if team_data[tm]['avance_plan']>0 and team_data[tm]['acum']/team_data[tm]['avance_plan']>=0.85 else RED,
            True,
            '#059669' if avance_plan_total>0 and total_acum/avance_plan_total>=1.0 else '#D97706' if avance_plan_total>0 and total_acum/avance_plan_total>=0.85 else RED,
            'big'),
        ('Av.Plan',    lambda tm: fmt(team_data[tm]['avance_plan']),      fmt(avance_plan_total),    '#F5F7FA', None, False, TXT),
        ('Dif.Plan',   lambda tm: fmt(team_data[tm]['dif_plan']),         fmt(dif_total),            '#EBF0F5', lambda tm: RED if team_data[tm]['dif_plan']<0 else TXT, 'red_bold', RED if dif_total<0 else TXT),
        ('Ritmo',      lambda tm: fmt(team_data[tm]['ritmo']),            fmt(ritmo_total),          '#F5F7FA', lambda tm: RED if team_data[tm]['ritmo']>team_data[tm]['prom'] else TXT, 'red_bold', RED if ritmo_total>prom_total else TXT),
        ('m³/hr',      lambda tm: f"{team_data[tm]['rendimiento']:.1f}",  f"{rend_total:.1f}",       '#EBF0F5', lambda tm: RED if team_data[tm]['rendimiento']<15 else TXT, 'red_bold', RED if rend_total<15 else TXT),
        ('TM(h)',      lambda tm: f"{team_data[tm]['tm_mant_hrs']:.1f}",  f"{total_tm_mant/60:.1f}", '#F5F7FA', lambda tm: RED if team_data[tm]['tm_mant_hrs']>33 else TXT, 'red_bold', RED if total_tm_mant/60>33 else TXT),
        ('Proyección', lambda tm: fmt(team_data[tm]['proy']),             fmt(proy_total),           '#EBF0F5', lambda tm: '#059669' if (team_data[tm]['proy']/METAS[tm]*100)>=80 else '#D97706' if (team_data[tm]['proy']/METAS[tm]*100)>=60 else RED, True, '#059669' if pct_total>=80 else '#D97706' if pct_total>=60 else RED),
    ]

    yc = y_hdr
    for ki, kpi in enumerate(kpi_defs):
        # Soportar 7 o 8 elementos (con flag 'big' opcional al final)
        if len(kpi) == 8:
            label, val_fn, total_val, bg, color_fn, bold, total_c, extra = kpi
        else:
            label, val_fn, total_val, bg, color_fn, bold, total_c = kpi
            extra = None

        # Altura mayor para filas destacadas (big)
        row_h = ROW_KPI_H * 1.45 if extra == 'big' else ROW_KPI_H
        yc -= row_h
        ym = yc + row_h/2

        # Fila Proyección o 'big' → fuente un poco más grande
        is_highlighted = (label == 'Proyección') or (extra == 'big')
        fs_val = F_KPI_V + 2.5 if is_highlighted else F_KPI_V
        fs_lbl = F_KPI_L + 1.5 if is_highlighted else F_KPI_L

        rect(0, yc, fig_w, row_h, bg)
        hline(yc, BORDER, 0.3)

        t(col_x(0) + P, ym, label, fs=fs_lbl, c=PRIMARY, ha='left', bold=True)
        for i, team in enumerate(TEAMS):
            cx = col_x(i+1) + col_w(i+1) - P
            c = color_fn(team) if color_fn else TXT
            if bold == 'red_bold':
                is_bold = (c == RED)
            else:
                is_bold = bold
            t(cx, ym, val_fn(team), fs=fs_val, c=c, bold=is_bold)
        rect(col_x(n_teams+1), yc, col_w(n_teams+1), row_h, '#C5DAEC')
        t(col_x(n_teams+1) + col_w(n_teams+1) - P, ym, total_val, fs=fs_val, c=total_c, bold=True)

    hline(yc, PRIMARY, 2.5)

    # ── FILAS DIARIAS ──
    dias_con_datos = set(daily['Dia'].unique().astype(int))
    for di, dia in enumerate(dias):
        yc -= ROW_H
        ym = yc + ROW_H/2

        dow_idx = dt_date(ANIO, MES, dia).weekday()
        dow = DIAS_SEMANA[dow_idx]
        tiene_datos = dia in dias_con_datos

        if not tiene_datos:
            bg = '#F3F4F6' if di%2==0 else '#FAFAFA'
        else:
            bg = '#F1F5F9' if di%2==0 else '#FFFFFF'
        rect(0, yc, fig_w, ROW_H, bg)
        hline(yc, BORDER, 0.2)

        dc = MUTED if not tiene_datos else PRIMARY
        t(col_x(0) + P, ym, f"{dia}{dow}", fs=F_DAY, c=dc, ha='left', bold=tiene_datos)

        total_dia = 0
        for i, team in enumerate(TEAMS):
            v = grid.get(dia, {}).get(team, 0)
            tm_v = grid_tm.get(dia, {}).get(team, 0)
            total_dia += v

            xl = col_x(i+1); cw = col_w(i+1)
            if not tiene_datos:
                cb = bg
            else:
                # Heatmap suave para celular: rojo (sin prod), verde (alto), amarillo (medio), naranja (bajo)
                cb = '#FECACA' if v==0 else '#BBF7D0' if v>=200 else '#FEF08A' if v>=100 else '#FDBA74'
            rect(xl, yc, cw, ROW_H, cb)

            cx = xl + cw - P
            if not tiene_datos:
                pass
            elif tm_v > 0:
                val_txt = fmt(v) if v > 0 else '—'
                t(cx, ym + 0.05, val_txt, fs=F_NUM, c=TXT, bold=True)
                t(cx, ym - 0.10, f"{tm_v/60:.1f}h", fs=F_TM, c=RED)
            else:
                val_txt = fmt(v) if v > 0 else '—'
                t(cx, ym, val_txt, fs=F_NUM, c=TXT, bold=True)

        rect(col_x(n_teams+1), yc, col_w(n_teams+1), ROW_H,
             '#E8EAF0' if not tiene_datos else '#C5DAEC')
        if tiene_datos:
            t(col_x(n_teams+1) + col_w(n_teams+1) - P, ym, fmt(total_dia), fs=F_NUM+2, c=PRIMARY_DK, bold=True)

    # Fila TOTAL removida — el "Acumulado" del bloque KPI arriba ya muestra estos valores

    # ── PROMEDIO ──
    yc -= ROW_H
    ym = yc + ROW_H/2
    rect(0, yc, fig_w, ROW_H, '#1E3A5F')
    t(col_x(0) + P, ym, "PRO", fs=F_NUM-1, c='white', ha='left', bold=True)
    for i, team in enumerate(TEAMS):
        td_c = len(daily[daily['Team'] == team])
        avg = team_data[team]['acum'] / td_c if td_c > 0 else 0
        t(col_x(i+1) + col_w(i+1) - P, ym, fmt(avg), fs=F_NUM-1, c='white', bold=True)
    rect(col_x(n_teams+1), yc, col_w(n_teams+1), ROW_H, '#0F2940')
    t(col_x(n_teams+1) + col_w(n_teams+1) - P, ym,
      fmt(total_acum/DD if DD>0 else 0), fs=F_NUM-1, c='white', bold=True)

    y_bottom = yc

    # Líneas verticales
    for i in range(n_teams + 3):
        x = col_x(min(i, n_teams+1)) if i <= n_teams+1 else fig_w
        lw = 0.8 if i in (0, 1, n_teams+1) else 0.3
        ax.plot([x, x], [y_bottom, y_top_grid], color=BORDER, lw=lw, clip_on=False)
    ax.plot([0, 0], [y_bottom, y_top_grid], color=PRIMARY, lw=1.2, clip_on=False)
    ax.plot([fig_w, fig_w], [y_bottom, y_top_grid], color=PRIMARY, lw=1.2, clip_on=False)

    # Footer
    t(fig_w/2, y_bottom - 0.12, "produccion.millalemu.com", fs=F_FOOTER, c=MUTED, ha='center')

    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    fig.savefig(str(OUTPUT), dpi=DPI, bbox_inches='tight',
                facecolor='white', edgecolor='none', pad_inches=0.03)
    plt.close(fig)

    from PIL import Image
    img = Image.open(str(OUTPUT))
    log(f"   ✅ Imagen: {img.size[0]}×{img.size[1]}px | {OUTPUT.stat().st_size // 1024} KB")
    return str(OUTPUT)

if __name__ == '__main__':
    generate()
